import os

import openpyxl
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import shutil
import webbrowser
import zipfile
import time

CURR_DIR_PATH = os.path.dirname(os.path.realpath(__file__))


class InventoryItems:
    def __init__(self):
        self.abreheim_shop = [
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ]
        self.abreheim_shop_defaulf = [
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ]
        self.store_count = [
            ['town', 'start', 'end', 'count'],
            ["Abreheim's items", 0, 0, ""],
            ["Lucy", 0, 0, ""],
        ]
        self.buy_history = [

            ['Items', 'count', "cost", "temp_cost", "price"],
            ['potion', 0, 0, "", 50],
            ['ether', 0, 0, "", 1500],
            ['Antidote', 0, 0, "", 80],
            ['Phoenix_down', 0, 0, "", 300],
            ['Silver_dust', 0, 0, "", 150],
            ['Tent', 0, 0, "", 3000],

        ]

class Cast_Magic:
    def __init__(self):
        self.magic = [
            ["Magic", "spell power", "mp cost", "gil cost", "Magic Type"],
            ['fire', 8, 4, 600, "black magic"],
            ['fire2', 20, 22, "", "black magic"],
            ['fire3', 64, 52, "", "black magic"],
            ['fire master', 64, 52, "", "black magic"],
            ['ice', 8, 4, 600, "black magic"],
            ['ice2', 20, 22, "", "black magic"],
            ['ice3', 64, 52, "", "black magic"],
            ['ice master', 64, 52, "", "black magic"],
            ['storm', 9, 8, 1200, "black magic"],
            ['storm2', 22, 30, "", "black magic"],
            ['storm3', 70, 60, "", "black magic"],
            ['storm master', 70, 60, "", "black magic"],
            ['quake', 9, 8, 1200, "black magic"],
            ['quake2', 22, 30, "", "black magic"],
            ['quake3', 70, 60, "", "black magic"],
            ['quake master', 70, 60, "", "black magic"],
            ['heal', 1.5, 24, 3000, "white magic"],
            ['heal2', 3, 48, "", "white magic"],
            ['life', 1, 80, "", "white magic"],

        ]

        self.magic_gil = [
            ["Magic", "gil cost"],
            ['fire', 600],
            ['ice', 600],
            ['storm', 1200],
            ['quake', 1200],
            ['heal', 3000],
            ['life', 6000],
        ]

        self.magic_sell = [
            ["Magic", "Sell"],
            ['fire', 300],
            ['fire2', 3000],
            ['fire3', 6000],
            ['fire master', 12000],
            ['ice', 300],
            ['ice2', 3000],
            ['ice3', 6000],
            ['ice master', 12000],
            ['storm', 600],
            ['storm2', 6000],
            ['storm3', 12000],
            ['storm master', 24000],
            ['quake', 600],
            ['quake2', 6000],
            ['quake3', 12000],
            ['quake master', 24000],
            ['heal', 1500],
            ['heal2', 6000],
            ['life', 3000],
        ]


class EnemyName:
    def __init__(self):
        self.enemy_name = [
            ["adjective", "name", "status_give", "immune", "loot", "qty_loot"],
            ['Big', "Giant", "", "", "potion", 1],
            ['Scary',"Spider", "", "", "silver_dust", 1],
            ['Unfatithful', "Dragon", "", "fire", "potion", 1],
            ['Demon',"Snake", "poison", "", "potion", 1],
            ['Brutal',"Lizard", "", "", "antidote", 1],
        ]


class LevelEnemy:
    def __init__(self):
        self.enemy_level = [
            ["Level", "StrBase", "", "Bonus", "Bonus Parameter", "StrBase Parameter", "Strength",
             "AtkFactor", "Atk Factor enemy", "Gil"],
            [1, 2.5, 2.5, 0, 3, "", 2.8, 3.5, 2.5, 32],
            [2, 3.14, 3.14, 3, 3, "", 3.83, 3.5, 2.5, 90],
            [3, 3.82, 3.82, 6, 3, "", 4.91,	3.5, 2.5, 165],
            [4, 4.54, 4.54, 9, 3, "", 6.02, 3.5, 2.5, 253],
            [5, 5.3, 5.3, 12, 3, "", 7.18, 3.5, 2.5, 354],
            [6, 6.1, 6.1, 15, 3, "", 8.37, 3.5, 2.5, 465],
            [7, 6.94, 6.94, 18, 3, "", 9.6, 3.5, 2.5, 586],
            [8, 7.82, 7.82, 21, 3, "", 10.88, 3.5, 2.5, 715],
            [9, 8.74, 8.74,    24, 3, "",  12.19, 3.5, 2.5, 854],
            [10, 9.7, 9.7,    27, 3, "", 13.54, 3.5, 2.5, 1000],
            [11, 10.62, 10.62, 30, 3, "",  14.86, 3.5, 2.5, 1154],
            [12, 11.5, 11.5, 33, 3, "", 16.13, 3.5, 2.5, 1315],
            [13, 12.34, 12.34, 36, 3, "", 17.37, 3.5, 2.5, 1483],
            [14, 13.14, 13.14, 39, 3, "", 18.56, 3.5, 2.5, 1657],
            [15, 13.9, 13.9, 42, 3, "", 19.71, 3.5, 2.5, 1837],
            [16, 14.62, 14.62, 45, 3, "", 20.83, 3.5, 2.5, 2024],
            [17, 15.3, 15.3, 48, 3, "", 21.9, 3.5, 2.5, 2216],
            [18, 15.94, 15.94, 51, 3, "", 22.93, 3.5, 2.5, 2415],
            [19, 16.62, 16.62, 54, 3, "", 24.01, 3.5, 2.5, 2619],
            [20, 17.34, 17.34, 57, 3, "", 25.12, 3.5, 2.5, 2829],

        ]

class WeaponsArmor:
    def __init__(self):
        self.weapon_powers = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots", "Attributes",
             "Protection", "Level", "hp", "mp", "type", "equipped"],
            ['sand storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['wind storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['blade storm', 7, 500, 2, 0, 0, 0, 0, 0, "weapon", 1],
            ['excalibur', 9, 5000, 4, 0, 0, 5, 0, 0, "weapon", 1],
            ['the mark of the serpents', 0, 15000, 0, 0, "poison", 0, 0, 0, "ring", 1],
            ['light armor', 0, 3000, 1, 0, 0, 0, 0, 0, "armor", 1],
            ['heavy armor', 0, 7000, 2, 0, 0, 0, 5, 1.2, "armor", 1],
            ['silver ring', 0, 15000, 0, 0, "paralyzed", 0, 0, 0, "ring", 1], ]

        self.shop_buy = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 250, 1],
            ['wind storm', 6, 250, 1],
            ['blade storm', 7, 500, 2],
            ['excalibur', 9, 5000, 4],
            ['the mark of the serpents', 0, 15000, 0],
            ['light armor', 0, 3000, 1],
            ['heavy armor', 0, 7000, 2],
            ['silver ring', 0, 15000, 0], ]

        self.shop_sell = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 125, 1],
            ['wind storm', 6, 125, 1],
            ['blade storm', 7, 250, 2],
            ['excalibur', 9, 2500, 4],
            ['the mark of the serpents', 0, 7500, 0],
            ['light armor', 0, 1500, 1],
            ['heavy armor', 0, 3500, 2],
            ['silver ring', 0, 7500, 0], ]


create_inventory = InventoryItems()
create_magic = Cast_Magic()
create_enemy = EnemyName()
create_enemy_level = LevelEnemy()
create_weapons_armor = WeaponsArmor()

class make_excel_files:
    def __init__(self):
        # create pandas dataframe from lists of class
        self.df_abreheim_shop = pd.DataFrame(create_inventory.abreheim_shop)
        self.df_abreheim_shop_default = pd.DataFrame(create_inventory.abreheim_shop_defaulf)
        self.df_abreheim_shop_store_count = pd.DataFrame(create_inventory.store_count)
        self.df_buy_history = pd.DataFrame(create_inventory.buy_history)
        self.df_magic = pd.DataFrame(create_magic.magic)
        self.df_magic_gill = pd.DataFrame(create_magic.magic_gil)
        self.df_magic_sell = pd.DataFrame(create_magic.magic_sell)
        self.df_enemy_name = pd.DataFrame(create_enemy.enemy_name)
        self.df_enemy_level = pd.DataFrame(create_enemy_level.enemy_level)
        self.df_weapons_power = pd.DataFrame(create_weapons_armor.weapon_powers)
        self.df_weapons_buy = pd.DataFrame(create_weapons_armor.shop_buy)
        self.df_weapons_sell = pd.DataFrame(create_weapons_armor.shop_sell)

    def create_folders(self):
        folders = ["\\save", "\\enemy", "\\inventory_items", "\\magic", "\\weapons_armor", "\\level"]
       # mode = 0o666
        for i in folders:
            path_folder = (CURR_DIR_PATH + i)
            obj = Path(path_folder)
            if obj.exists():
                continue
            else:
                # print(CURR_DIR_PATH + i)
                os.mkdir(CURR_DIR_PATH + i)


    def create_main_file(self, folder_name, file_name):

        path = (CURR_DIR_PATH + folder_name + file_name)
        obj = Path(path)
        if obj.exists():
            print("")

        else:
            wb = openpyxl.Workbook()
            wb.save(path)
            wb.close()


    def write_data(self, sheet_name, folder_name, work_book_name, df_name):
        path = (CURR_DIR_PATH + folder_name + work_book_name)
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        df_name.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        if "Sheet" in book.sheetnames:
            book.remove(book["Sheet"])
        writer.save()


    def run_main(self):

        # create folders
        make_excel_files.create_folders(self)


        # Creates main files
        make_excel_files.create_main_file(self, "\\enemy", "\\enemy_name.xlsx")
        make_excel_files.create_main_file(self, "\\inventory_items", "\\item_store.xlsx")
        make_excel_files.create_main_file(self, "\\magic", "\\cast_magic.xlsx")
        make_excel_files.create_main_file(self, "\\weapons_armor", "\\weapons_armor.xlsx")
        make_excel_files.create_main_file(self, "\\level", "\\level_up.xlsx")



        # create excel files
        make_excel_files.write_data(self, "Abreheim's items", "\\inventory_items", "\\item_store.xlsx", self.df_abreheim_shop)
        make_excel_files.write_data(self, "Abreheim's items_default", "\\inventory_items", "\\item_store.xlsx", self.df_abreheim_shop_default)
        make_excel_files.write_data(self, "store_count", "\\inventory_items", "\\item_store.xlsx", self.df_abreheim_shop_store_count)
        make_excel_files.write_data(self, "buy_history", "\\inventory_items", "\\item_store.xlsx", self.df_buy_history)
        make_excel_files.write_data(self, "magic", "\\magic", "\\cast_magic.xlsx", self.df_magic)
        make_excel_files.write_data(self, "magic_gil", "\\magic", "\\cast_magic.xlsx", self.df_magic_gill)
        make_excel_files.write_data(self, "magic_sell", "\\magic", "\\cast_magic.xlsx", self.df_magic_sell)
        make_excel_files.write_data(self, "Ark1", "\\enemy", "\\enemy_name.xlsx", self.df_enemy_name)
        make_excel_files.write_data(self, "Strength Mod enemy", "\\level", "\\level_up.xlsx", self.df_enemy_level)
        make_excel_files.write_data(self, "weapon_powers", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_power)
        make_excel_files.write_data(self, "shop_buy", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_buy)
        make_excel_files.write_data(self, "shop_sell", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_sell)


    def download_music_content(self):
        url = "https://doc-14-1k-drive-data-export.googleusercontent.com/download/so05epia28g54q0632d0hm49013e6gkb/dvghur88cplnabsams0g0hms923cp964/1663159500000/e4f5f880-ff5a-43a3-8c27-5e792becd879/109853872616057424140/ADt3v-PvAv6vsDIV7NQDGeGc8pxN-tg46R3UXMQ1r1l7o_fDa4hBGv5rCPeQyNmHaFxj9LCY3GPlQwfWYSgJB_dnp6pK0Gle1kVfgRWehYxkWgG5w7-s0ioK0oy9OdPt1tjuTZ5od3jMQXddGEJk0M9-kEHsEFUb0w3OowdPQ2bf5AH7i3FToFhOJ389CDRRwxdLrCCgIHb7GxTrqtHmNLUDPBMHqaSjMimRM8lVpx4ejHBW7iAn1GHae0BNg4R4f4eSA_BKGrHHFCKgipwpoAlDapew8VIodWwD2xIyq6i5JPMtKMBbnMkxMCB_XtZIhGi1rYFzE2J0?authuser=0&nonce=h2br6ipbc1nni&user=109853872616057424140&hash=p2ljtl0r6q7bcilam7ortfvo3coq9svn"
        content = webbrowser.open(url)
        return content


    def find_files_unzip(self):

        user_id = os.getlogin()
        for filename in os.listdir(f'C:/Users/{user_id}/Downloads'):
            if filename.startswith('code_chris_music'):
                original = (f'C:\\Users\\{user_id}\\Downloads\\{filename}')
                target = CURR_DIR_PATH + "\\music_temp\\"
                with zipfile.ZipFile(original, 'r') as zip_ref:
                    zip_ref.extractall(path=target)


    def move_music_files(self, folder_name):
        original = CURR_DIR_PATH + "\\music_temp" + "\\code_chris_music" + folder_name
        target = CURR_DIR_PATH + "\\music" + folder_name
        shutil.move(original, target)


    def remove_music_temp(self):
        shutil.rmtree(CURR_DIR_PATH + "\\music_temp")



