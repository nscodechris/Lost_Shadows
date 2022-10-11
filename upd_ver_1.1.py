import pandas as pd
import os
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import shutil
from io import BytesIO
import requests

CURR_VERSION = "ver 1.1"

def open_ver_file(file_name):

    result = []
    result2 = []
    for root, dir, files in os.walk("C:\\"):
        if file_name in files:
            result.append(os.path.join(root, file_name))
            result2.append(root)
    file_path = result[0]
    path_root = result2[0]
    file = open(file_path)
    content = file.readlines()
    try:
        ver = content[1]
    except IndexError:
        ver = "ver 1.0"
    return ver, path_root, file_path

print("installing patch....")
CURR_DIR_PATH = open_ver_file("lost_shadow_installed.txt")[1]


def get_git_hub_file():
    import zipfile
    print('Downloading started')

    #Defining the zip file URL
    url = 'https://media.githubusercontent.com/media/nscodechris/Lost_Shadows/main/LostShadows_install.zip'
    os.mkdir(CURR_DIR_PATH + "\\temp_git")
    # Split URL to get the file name
    filename = url.split('/')[-1]

    # Downloading the file by sending the request to the URL
    req = requests.get(url)
    print('Downloading Completed')

    # extracting the zip file contents
    zipfile = zipfile.ZipFile(BytesIO(req.content))
    zipfile.extractall(CURR_DIR_PATH + "\\temp_git")

def copy_replace_git_hub_file():
    shutil.copytree(CURR_DIR_PATH + "\\temp_git", CURR_DIR_PATH, dirs_exist_ok=True)
    shutil.rmtree(CURR_DIR_PATH + "\\temp_git")




class InventoryItemsAdd:
    def __init__(self):

        self.store_count = [
            ["North Cave items", 0, 0, ""],

        ]


create_inventory_add = InventoryItemsAdd()

class MakeExcelFilesAdd:
    def __init__(self):
        # create pandas dataframe from lists of class
        self.store_count_add = pd.DataFrame(create_inventory_add.store_count)

    def write_to_sheet_data(self, sheet_name, folder_name, work_book_name, df_name):
        path = (CURR_DIR_PATH + folder_name + work_book_name)
        # MakeExcelFiles.Remove_password_xlsx(self, path, "five@morning!Mind5")
        wb = load_workbook(path)
        ws = wb[sheet_name]
        for r in dataframe_to_rows(df_name, index=False,
                                   header=False):  # No index and don't append the column headers
            ws.append(r)
        wb.save(path)

    def run_main(self):

        MakeExcelFilesAdd.write_to_sheet_data(self, "store_count", "\\inventory_items", "\\item_store.xlsx",
                                    self.store_count_add)

class InventoryItems:
    def __init__(self, items, items_default):
        self.shop = items
        self.shop_default = items_default
        self.store_count = [
            ['town', 'start', 'end', 'count'],
            ["Abreheim's items", 0, 0, ""],
            ["North Cave items", 0, 0, ""],

        ]
        self.buy_history = [

            ['Items', 'count', "cost", "temp_cost", "price"],
            ['potion', 0, 0, "", 50],
            ['ether', 0, 0, "", 1500],
            ['Antidote', 0, 0, "", 80],
            ['Phoenix_down', 0, 0, "", 300],
            ['Silver_dust', 0, 0, "", 150],
            ['Tent', 0, 0, "", 3000],

        ]


class Cast_Magic:
    def __init__(self):
        self.magic = [
            ["Magic", "spell power", "mp cost", "gil cost", "Magic Type"],
            ['fire', 8, 4, 600, "black magic"],
            ['fire2', 20, 22, "", "black magic"],
            ['fire3', 64, 52, "", "black magic"],
            ['fire master', 64, 52, "", "black magic"],
            ['ice', 8, 4, 600, "black magic"],
            ['ice2', 20, 22, "", "black magic"],
            ['ice3', 64, 52, "", "black magic"],
            ['ice master', 64, 52, "", "black magic"],
            ['storm', 9, 8, 1200, "black magic"],
            ['storm2', 22, 30, "", "black magic"],
            ['storm3', 70, 60, "", "black magic"],
            ['storm master', 70, 60, "", "black magic"],
            ['quake', 9, 8, 1200, "black magic"],
            ['quake2', 22, 30, "", "black magic"],
            ['quake3', 70, 60, "", "black magic"],
            ['quake master', 70, 60, "", "black magic"],
            ['heal', 1.5, 24, 3000, "white magic"],
            ['heal2', 3, 48, "", "white magic"],
            ['life', 1, 80, "", "white magic"],

        ]

        self.magic_gil = [
            ["Magic", "gil cost"],
            ['fire', 600],
            ['ice', 600],
            ['storm', 1200],
            ['quake', 1200],
            ['heal', 3000],
            ['life', 6000],
        ]

        self.magic_sell = [
            ["Magic", "Sell"],
            ['fire', 300],
            ['fire2', 3000],
            ['fire3', 6000],
            ['fire master', 12000],
            ['ice', 300],
            ['ice2', 3000],
            ['ice3', 6000],
            ['ice master', 12000],
            ['storm', 600],
            ['storm2', 6000],
            ['storm3', 12000],
            ['storm master', 24000],
            ['quake', 600],
            ['quake2', 6000],
            ['quake3', 12000],
            ['quake master', 24000],
            ['heal', 1500],
            ['heal2', 6000],
            ['life', 3000],
        ]


class EnemyName:
    def __init__(self):
        self.enemy_name = [
            ["adjective", "name", "status_give", "immune", "loot", "qty_loot"],
            ['Big', "Giant", "", "", "potion", 1],
            ['Scary', "Spider", "", "", "silver_dust", 1],
            ['Unfatithful', "Dragon", "", "fire", "potion", 1],
            ['Demon', "Snake", "poison", "", "potion", 1],
            ['Brutal', "Lizard", "", "", "antidote", 1],
        ]


class LevelEnemy:
    def __init__(self):
        self.enemy_level = [
            ["Level", "StrBase", "", "Bonus", "Bonus Parameter", "StrBase Parameter", "Strength",
             "AtkFactor", "Atk Factor enemy", "Gil"],
            [1, 2.5, 2.5, 0, 3, "", 2.8, 3.5, 2.5, 32],
            [2, 3.14, 3.14, 3, 3, "", 3.83, 3.5, 2.5, 90],
            [3, 3.82, 3.82, 6, 3, "", 4.91, 3.5, 2.5, 165],
            [4, 4.54, 4.54, 9, 3, "", 6.02, 3.5, 2.5, 253],
            [5, 5.3, 5.3, 12, 3, "", 7.18, 3.5, 2.5, 354],
            [6, 6.1, 6.1, 15, 3, "", 8.37, 3.5, 2.5, 465],
            [7, 6.94, 6.94, 18, 3, "", 9.6, 3.5, 2.5, 586],
            [8, 7.82, 7.82, 21, 3, "", 10.88, 3.5, 2.5, 715],
            [9, 8.74, 8.74, 24, 3, "", 12.19, 3.5, 2.5, 854],
            [10, 9.7, 9.7, 27, 3, "", 13.54, 3.5, 2.5, 1000],
            [11, 10.62, 10.62, 30, 3, "", 14.86, 3.5, 2.5, 1154],
            [12, 11.5, 11.5, 33, 3, "", 16.13, 3.5, 2.5, 1315],
            [13, 12.34, 12.34, 36, 3, "", 17.37, 3.5, 2.5, 1483],
            [14, 13.14, 13.14, 39, 3, "", 18.56, 3.5, 2.5, 1657],
            [15, 13.9, 13.9, 42, 3, "", 19.71, 3.5, 2.5, 1837],
            [16, 14.62, 14.62, 45, 3, "", 20.83, 3.5, 2.5, 2024],
            [17, 15.3, 15.3, 48, 3, "", 21.9, 3.5, 2.5, 2216],
            [18, 15.94, 15.94, 51, 3, "", 22.93, 3.5, 2.5, 2415],
            [19, 16.62, 16.62, 54, 3, "", 24.01, 3.5, 2.5, 2619],
            [20, 17.34, 17.34, 57, 3, "", 25.12, 3.5, 2.5, 2829],

        ]


class WeaponsArmor:
    def __init__(self):
        self.weapon_powers = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots", "Attributes",
             "Protection", "Level", "hp", "mp", "type", "equipped"],
            ['sand storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['wind storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['blade storm', 7, 500, 2, 0, 0, 0, 0, 0, "weapon", 1],
            ['excalibur', 9, 5000, 4, 0, 0, 5, 0, 0, "weapon", 1],
            ['the mark of the serpents', 0, 15000, 0, 0, "poison", 0, 0, 0, "ring", 1],
            ['light armor', 0, 3000, 1, 0, 0, 0, 0, 0, "armor", 1],
            ['heavy armor', 0, 7000, 2, 0, 0, 0, 5, 1.2, "armor", 1],
            ['silver ring', 0, 15000, 0, 0, "paralyzed", 0, 0, 0, "ring", 1], ]

        self.shop_buy = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 250, 1],
            ['wind storm', 6, 250, 1],
            ['blade storm', 7, 500, 2],
            ['excalibur', 9, 5000, 4],
            ['the mark of the serpents', 0, 15000, 0],
            ['light armor', 0, 3000, 1],
            ['heavy armor', 0, 7000, 2],
            ['silver ring', 0, 15000, 0], ]

        self.shop_sell = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 125, 1],
            ['wind storm', 6, 125, 1],
            ['blade storm', 7, 250, 2],
            ['excalibur', 9, 2500, 4],
            ['the mark of the serpents', 0, 7500, 0],
            ['light armor', 0, 1500, 1],
            ['heavy armor', 0, 3500, 2],
            ['silver ring', 0, 7500, 0], ]


abreheim = InventoryItems([
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ], [
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ])

north_cave_shop = InventoryItems([
            ['Items', 'Cost', "Quantity"],
            ['potion', 250, 5],
            ['ether', 3000, 5],
            ['antidote', 160, 5],
            ['phoenix_down', 500, 5],
            ['silver_dust', 200, 5],
            ['tent', 4000, 5],
        ], [
            ['Items', 'Cost', "Quantity"],
            ['potion', 250, 5],
            ['ether', 3000, 5],
            ['antidote', 160, 5],
            ['phoenix_down', 500, 5],
            ['silver_dust', 200, 5],
            ['tent', 4000, 5],
        ],)


create_magic = Cast_Magic()
create_enemy = EnemyName()
create_enemy_level = LevelEnemy()
create_weapons_armor = WeaponsArmor()


class MakeExcelFiles:
    def __init__(self):
        # create pandas dataframe from lists of class
        self.df_abreheim_shop = pd.DataFrame(abreheim.shop)
        self.df_abreheim_shop_default = pd.DataFrame(abreheim.shop_default)
        self.df_abreheim_shop_store_count = pd.DataFrame(abreheim.store_count)
        self.df_buy_history = pd.DataFrame(abreheim.buy_history)
        self.df_magic = pd.DataFrame(create_magic.magic)
        self.df_magic_gill = pd.DataFrame(create_magic.magic_gil)
        self.df_magic_sell = pd.DataFrame(create_magic.magic_sell)
        self.df_enemy_name = pd.DataFrame(create_enemy.enemy_name)
        self.df_enemy_level = pd.DataFrame(create_enemy_level.enemy_level)
        self.df_weapons_power = pd.DataFrame(create_weapons_armor.weapon_powers)
        self.df_weapons_buy = pd.DataFrame(create_weapons_armor.shop_buy)
        self.df_weapons_sell = pd.DataFrame(create_weapons_armor.shop_sell)
        # new version 1.2 added shops:
        self.df_north_cave_shop = pd.DataFrame(north_cave_shop.shop)
        self.df_north_cave_shop_default = pd.DataFrame(north_cave_shop.shop_default)


    def create_folders(self):
        folders = ["\\enemy", "\\inventory_items", "\\magic", "\\weapons_armor", "\\level"]
        # mode = 0o666
        for i in folders:
            path_folder = (CURR_DIR_PATH + i)
            obj = Path(path_folder)
            if obj.exists():
                continue
            else:
                # print(CURR_DIR_PATH + i)
                os.mkdir(CURR_DIR_PATH + i)



    def create_main_file(self, folder_name, file_name):

        path = (CURR_DIR_PATH + folder_name + file_name)
        obj = Path(path)
        if obj.exists():
            print("")

        else:
            wb = openpyxl.Workbook()
            wb.save(path)
            wb.close()

    def write_data(self, sheet_name, folder_name, work_book_name, df_name):
        path = (CURR_DIR_PATH + folder_name + work_book_name)
        # MakeExcelFiles.Remove_password_xlsx(self, path, "five@morning!Mind5")
        try:
            with pd.ExcelWriter(path, mode="a", engine="openpyxl") as writer:
                df_name.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        except ValueError:
            print()
    def delete_org_sheet(self, folder_name, file_name):
        path = (CURR_DIR_PATH + folder_name + file_name)
        book = openpyxl.load_workbook(path)
        if "Sheet" in book.sheetnames:
            book.remove(book["Sheet"])
        book.save(path)

    def run_main(self):

        # create folders
        MakeExcelFiles.create_folders(self)
        # Creates main files
        MakeExcelFiles.create_main_file(self, "\\enemy", "\\enemy_name.xlsx")
        MakeExcelFiles.create_main_file(self, "\\inventory_items", "\\item_store.xlsx")
        MakeExcelFiles.create_main_file(self, "\\magic", "\\cast_magic.xlsx")
        MakeExcelFiles.create_main_file(self, "\\weapons_armor", "\\weapons_armor.xlsx")
        MakeExcelFiles.create_main_file(self, "\\level", "\\level_up.xlsx")

        # create excel files
        MakeExcelFiles.write_data(self, "Abreheim's items", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop)
        MakeExcelFiles.write_data(self, "Abreheim's items_default", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop_default)

        MakeExcelFiles.write_data(self, "store_count", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop_store_count)
        MakeExcelFiles.write_data(self, "buy_history", "\\inventory_items", "\\item_store.xlsx", self.df_buy_history)
        MakeExcelFiles.write_data(self, "magic", "\\magic", "\\cast_magic.xlsx", self.df_magic)
        MakeExcelFiles.write_data(self, "magic_gil", "\\magic", "\\cast_magic.xlsx", self.df_magic_gill)
        MakeExcelFiles.write_data(self, "magic_sell", "\\magic", "\\cast_magic.xlsx", self.df_magic_sell)
        MakeExcelFiles.write_data(self, "Ark1", "\\enemy", "\\enemy_name.xlsx", self.df_enemy_name)
        MakeExcelFiles.write_data(self, "Strength Mod enemy", "\\level", "\\level_up.xlsx", self.df_enemy_level)
        MakeExcelFiles.write_data(self, "weapon_powers", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_power)
        MakeExcelFiles.write_data(self, "shop_buy", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_buy)
        MakeExcelFiles.write_data(self, "shop_sell", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_sell)
        MakeExcelFiles.delete_org_sheet(self, "\\enemy", "\\enemy_name.xlsx")
        MakeExcelFiles.delete_org_sheet(self, "\\inventory_items", "\\item_store.xlsx")
        MakeExcelFiles.delete_org_sheet(self, "\\magic", "\\cast_magic.xlsx")
        MakeExcelFiles.delete_org_sheet(self, "\\weapons_armor", "\\weapons_armor.xlsx")
        MakeExcelFiles.delete_org_sheet(self, "\\level", "\\level_up.xlsx")
        # version 1.1
        MakeExcelFiles.write_data(self, "North Cave items", "\\inventory_items", "\\item_store.xlsx",
                                  self.df_north_cave_shop)
        MakeExcelFiles.write_data(self, "North Cave items_default", "\\inventory_items", "\\item_store.xlsx",
                                  self.df_north_cave_shop_default)


def delete_folders_reset():
    folders = ["\\enemy", "\\inventory_items", "\\magic", "\\weapons_armor", "\\level"]
    # mode = 0o666
    for i in folders:
        path_folder = (CURR_DIR_PATH + i)
        obj = Path(path_folder)
        if obj.exists():
            shutil.rmtree(CURR_DIR_PATH + i)
        else:
            continue


def check_files(sheet_name, folder_name, file_name):
    path = (CURR_DIR_PATH + folder_name + file_name)
    obj = Path(path)
    if obj.exists():
        book = openpyxl.load_workbook(path)
        if sheet_name in book.sheetnames:
            if open_ver_file("lost_shadow_installed.txt")[0] == "ver 1.0":
                make_excel_files_add = MakeExcelFilesAdd()
                make_excel_files_add.run_main()
                add_stores = MakeExcelFiles()
                add_stores.write_data("North Cave items", "\\inventory_items", "\\item_store.xlsx",
                                      add_stores.df_north_cave_shop)
                add_stores.write_data("North Cave items_default", "\\inventory_items", "\\item_store.xlsx",
                                      add_stores.df_north_cave_shop_default)
                write_version()
            # for new version updates!
            elif open_ver_file("lost_shadow_installed.txt")[0] == CURR_VERSION:
                print()
        else:
            delete_folders_reset()
            write_version()
            make_excel_files = MakeExcelFiles()
            make_excel_files.run_main()
            make_excel_files_add = MakeExcelFilesAdd()
            make_excel_files_add.run_main()

    else:
        delete_folders_reset()
        write_version()
        make_excel_files = MakeExcelFiles()
        make_excel_files.run_main()
        make_excel_files_add = MakeExcelFilesAdd()
        make_excel_files_add.run_main()



def write_version():
    file = open(CURR_DIR_PATH + "\\lost_shadow_installed.txt")
    content = file.readlines()
    try:
        if type(content[1]) is str:
            file.close()
            my_file = open(CURR_DIR_PATH + "\\lost_shadow_installed.txt")
            string_list = my_file.readlines()
            string_list[-1] = CURR_VERSION + "\n"
            my_file = open(CURR_DIR_PATH + "\\lost_shadow_installed.txt", "w")
            new_file_contents = "".join(string_list)
            my_file.write(new_file_contents)

    except IndexError:
        with open(CURR_DIR_PATH + "\\lost_shadow_installed.txt", "a") as f:
            f.write("\n" + CURR_VERSION)
            f.close()


get_git_hub_file()
copy_replace_git_hub_file()
check_files("store_count", "\\inventory_items", "\\item_store.xlsx")
