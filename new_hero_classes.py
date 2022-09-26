
# Edit version
#

# Hero Classes
import math
import time
from datetime import datetime
import os
import pandas as pd
import random
import sys
import shelve
import glob
import pygame
from art import *
import shutil
import secrets
import openpyxl
import openpyxl as xl
from pathlib import Path
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import hang_man_game

# important file names:
CURR_DIR_PATH = os.path.dirname(os.path.realpath(__file__))



weapon_armor_file = "weapons_armor.xlsx"
weapon_armor_file_sheet_shop_sell = "shop_sell"
weapon_armor_file_sheet_shop_buy = "shop_buy"
weapon_armor_file_sheet_shop_weapon_powers = "weapon_powers"
inventory_item_file = "item_store.xlsx"
magic_file = "cast_magic.xlsx"
magic_file_sheet_magic = "magic"
magic_file_sheet_magic_gill = "magic_gil"
magic_file_sheet_magic_sell = "magic_sell"
enemy_name_file = "enemy_name.xlsx"
level_file = "level_up.xlsx"
level_file_sheet_strenght_mod_enemy = "Strength Mod enemy"

# test for sound



# To be able to display all rows an columns etc in pandas
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', -1)


pygame.init()
pygame.mixer.init()

# A few time_wait variables
time_story = 2.0
time_long_wait = 3.0
time_short_wait = 1.5
time_writing = 0.15
time_load = 1.5
time_save = 0.2


victory_text = text2art("VICTORY!!!")


# The Hero Class


class YourHero:
    def __init__(self):

        # Status
        self.strength = 0
        self.hp = 14
        self.max_hp = 0
        self.enemy_boss_level = 0
        self.level = 1
        self.next_level_exp = 0
        self.exp = 0
        self.exp_show = 0
        self.name = ""
        self.defence = 0
        self.atk = 0
        self.stat = {}
        self.total_stat = {}
        self.win_stat = {}
        self.win_stat_2 = {}
        self.win_stat_3 = {}
        # Magic
        self.magic = 0
        self.mp = 9
        self.max_mp = 0
        self.spirit = 0
        self.magic_atk = 0
        self.cast_magic_atk = 0
        self.magic_cost = 0
        self.magic_gil_cost = 0
        self.magic_heal = 0
        self.magic_use_inv = {}
        self.magic_spells = {"Magic": [], "mp cost": [], "AP": [], "Level": [], "Equipped": [], "Magic Type": []}
        self.magic_gil_sell = 0
        self.magic_slots = 0
        self.magic_slots_max = 0
        self.magic_type = ""
        # Status Effects
        self.status_effect_name = ""
        self.status_effect_2_name = ""
        self.status_poison = 0
        self.status_paralyzed = 0
        # Use of Items
        self.antidote_cure = 0
        self.silver_dust_cure = 0
        self.potion_cure = 0
        self.ether_cure = 0
        self.elixir_cure = 0
        self.phoenix_down_cure = 0
        # Item inventory
        self.item_list = {}
        self.inventory = {"Item": [], "QTY": []}
        # Weapons & Shields
        self.armory_inventory = {"Weapon/Armor": [], "Magic Slots": [], "weapon power": [],
                                 "Protection": [], "Equipped": [], "Type": []}
        self.weapon_gil_cost = 0
        self.weapon_gil_sell = 0
        self.weapon_equipped = 0
        self.weapon_equipped_max = 1
        self.armor_equipped = 0
        self.armor_equipped_max = 1
        self.rings_equipped = 0
        self.rings_equipped_max = 2
        self.weapon_level = 0
        self.protection = ""
        self.weapon_type = ""
        self.weapon_power = 0
        self.weapon_atk = 0

        # Counts
        self.run = 0
        self.dead_times = 0
        self.defence_count = 0
        self.factor_level_up = 2
        self.bonus_factor = 0
        self.level_count = self.level - 1
        self.story_number = 0
        self.help = 0
        self.open_chest_count = 0
        self.chapter_1_round_room_count = 0
        self.chapter_1_ally_woman_wall_away = 0
        self.chapter_1_ally_woman_green_hat_away = 0
        self.sleep_abreheim_inn = 0
        self.inn_room_1 = 0
        self.inn_room_2 = 0
        self.food_1 = 0
        self.food_2 = 0
        self.store_quantity = 0
        self.store_cost = 0
        self.abreheim_enter = 0
        self.store_money_back = 0

        # Key Items_puzzles
        self.key_item_1 = 0
        self.key_item_1_name = ""
        self.key_item_2 = 0
        self.key_item_2_name = ""
        self.key_item_3 = 0
        self.key_item_3_name = ""
        self.key_item_4 = 0
        self.key_item_4_name = ""
        self.story_name = ""
        self.music_name = ""
        self.chapter = ""
        self.progress = ""

        # Installed game
        self.game_installed = 0

    # The formula for leveling up heroes
    def x_level_up(self):
        next_level_exp = math.floor(.1 * (self.factor_level_up ** 4) + 4.2 * (self.factor_level_up ** 3)
                                    + 6.1 * (self.factor_level_up ** 2) + (1.4 * self.factor_level_up) - 11.4)
        x_next_level = next_level_exp
        self.next_level_exp = x_next_level
        return next_level_exp

    # Compute magic for heroes
    def x_magic(self):
        m_magic_base = ([2.0, 2.06, 2.12, 2.19, 2.26, 2.34, 2.42, 2.5, 2.59, 2.68, 2.77, 2.85,
                         2.93, 3.01, 3.08, 3.15, 3.21, 3.27, 3.33, 3.4])
        m_magic_calc = round((m_magic_base[self.level_count] + self.level * 3 / 10 + self.bonus_factor / 32), 1)
        magic = m_magic_calc
        self.magic = magic
        return magic

    # Compute spirit for heroes
    def x_spirit(self):
        s_spirit_base = ([2.5, 3.14, 3.82, 4.54, 5.3, 6.1, 6.94, 7.82, 8.74, 9.7,
                          10.62, 11.5, 12.34, 13.14, 13.9, 14.62, 15.3, 15.94, 16.62, 17.34])
        s_spirit_calc = round((s_spirit_base[self.level_count] + self.level * 3 / 10 * self.bonus_factor / 32), 1)
        spirit = s_spirit_calc
        self.spirit = spirit
        return spirit

    # Compute strength for heroes
    def x_strength(self):
        s_strength_base = ([2.5, 3.14, 3.82, 4.54, 5.3, 6.1, 6.94, 7.82, 8.74, 9.7, 10.62, 11.5,
                            12.34, 13.14, 13.9, 14.62, 15.3, 15.94, 16.62, 17.34])
        s_strength_calc = round(s_strength_base[self.level_count] + (self.level * 3 / 10) + (self.bonus_factor / 32), 1)

        strength = s_strength_calc
        self.strength = strength
        return strength

    # Compute hp for heroes
    def x_hp(self):

        hp_mod = ([250, 314, 382, 454, 530, 610, 694, 782, 874, 970,
                   1062, 1150, 1234, 1314, 1390, 1462, 1530, 1594, 1662, 1734])

        if math.ceil(self.strength * hp_mod[self.level_count] * self.level / 50) > 9999:
            h_hp = 9999
            self.max_hp = h_hp
            return h_hp
        elif math.ceil(self.strength * hp_mod[self.level_count] * self.level / 50) <= 9999:
            hp_calc = math.ceil(self.strength * hp_mod[self.level_count] * self.level / 50)
            h_hp = hp_calc
            self.max_hp = h_hp
            return h_hp

    # Compute mp for heroes
    def x_mp(self):
        mp_mod = ([200, 206, 212, 219, 226, 234, 242, 250, 259, 268,
                   277, 285, 293, 301, 308, 315, 321, 327, 333, 340])
        if math.ceil(self.magic * mp_mod[self.level_count] * self.level / 50) > 999:
            m_mp = 999
            self.max_mp = m_mp
            return m_mp
        elif math.ceil(self.magic * mp_mod[self.level_count] * self.level / 50) <= 999:
            mp_calc = math.floor(self.magic * mp_mod[self.level_count] * self.level / 50)
            m_mp = mp_calc
            self.max_mp = m_mp
            return m_mp

    # Compute magic_attack for heroes
    def x_magic_atk(self):
        magic_atk_calc = math.ceil(self.max_mp * self.spirit / self.level / 2)
        magic_atk = magic_atk_calc
        self.magic_atk = magic_atk
        return magic_atk

    # Compute melee Attack for heroes
    def x_atk(self):

        atk_calc = math.floor((self.max_hp * self.strength / self.level / 4))
        atk = atk_calc
        self.atk = atk
        return atk

    # Compute defence for heroes
    def x_defence(self):
        defence_calc = math.ceil(self.strength + self.level * 3 / 10 + self.atk / 2)
        defence = defence_calc
        self.defence = defence
        return defence

    # Compute weapon_damage for heroes
    def x_weapon_dmg(self, weapon_power):
        w_damage = math.floor(((weapon_power / 16) * (self.level + self.atk + self.strength) * 2))
        self.weapon_atk = w_damage
        return w_damage

    # Getting the weapon power from excel file
    def x_weapon_shop_power(self, name):
        weapon_list = pd.read_excel(dir_path.weapons_armor_path + "weapons_armor.xlsx")
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        type_give = dt.loc[name]["weapon power"]
        type = type_give
        self.weapon_power = type
        return type

    # Getting the weapon type from excel file
    def x_weapon_shop_type(self, name):

        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        type_give = dt.loc[name]["type"]
        type = type_give
        self.weapon_type = type
        return type

    # Getting the weapon protection from excel file
    def x_weapon_shop_protection(self, name):
        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        protection_give = dt.loc[name]["Protection"]
        protection = protection_give
        self.protection = protection
        return protection

    # Getting the magic slots, for the weapon of choice from excel file
    def x_weapon_shop_slots(self, name):
        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        weapon_slots = dt.loc[name]["Magic Slots"]
        slots = weapon_slots
        self.magic_slots_max = slots
        return slots

    # Getting the weapon level from excel file
    def x_weapon_shop_level(self, name):
        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        weapon_level = dt.loc[name]["Level"]
        level = weapon_level
        self.weapon_level = level
        return level

    # Getting cost for the weapon when selling
    def x_weapon_gil_sell(self, name):

        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file,
                                    sheet_name=weapon_armor_file_sheet_shop_sell)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        weapon_gill_sell = dt.loc[name]["gil cost"]
        gil_sell = weapon_gill_sell
        self.weapon_gil_sell = gil_sell
        return gil_sell

    # Getting the cost for the weapon when buying
    def x_weapon_gil_cost(self, name):
        weapon_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file)
        dt = pd.DataFrame(weapon_list)
        dt.set_index("Weapon/armor", inplace=True)
        weapon_gill_cost = dt.loc[name]["gil cost"]
        gil_cost = weapon_gill_cost
        self.weapon_gil_cost = gil_cost
        return gil_cost

    # Getting cost for the magic when selling
    def x_magic_gil_sell(self, name):

        magic_sell_list = pd.read_excel(dir_path.magic_path + magic_file, sheet_name=magic_file_sheet_magic_sell)
        df = pd.DataFrame(magic_sell_list)
        df.set_index("Magic", inplace=True)
        spell_gill_sell = df.loc[name]["Sell"]
        sell_gill = spell_gill_sell
        self.magic_gil_sell = sell_gill
        return sell_gill

    # Getting cost for the magic when buying
    def x_magic_gil_cost(self, name):
        magic_spell_list = pd.read_excel(dir_path.magic_path + magic_file)
        dt = pd.DataFrame(magic_spell_list)
        dt.set_index("Magic", inplace=True)
        spell_gill_cost = dt.loc[name]["gil cost"]
        gil_cost = spell_gill_cost
        self.magic_gil_cost = gil_cost
        return gil_cost

    # Getting magic damage from excel file
    def x_magic_dmg(self, name):
        magic_spell_list = pd.read_excel(dir_path.magic_path + magic_file)
        dt = pd.DataFrame(magic_spell_list)
        dt.set_index("Magic", inplace=True)
        spell_power = dt.loc[name]["spell power"]
        m_damage = ((spell_power / 16) * (self.level + self.magic_atk + self.spirit) * 2)
        self.cast_magic_atk = m_damage
        return m_damage

    # Getting Magic Type from excel file
    def x_magic_type(self, name):
        magic_spell_list = pd.read_excel(dir_path.magic_path + magic_file)
        dt = pd.DataFrame(magic_spell_list)
        dt.set_index("Magic", inplace=True)
        magic_type = dt.loc[name]["Magic Type"]
        self.magic_type = magic_type
        return magic_type

    # Getting magic heal power - if heal, cure is used, how much points to receive
    def x_magic_heal(self, name):
        magic_spell_list = pd.read_excel(dir_path.magic_path + magic_file)
        dt = pd.DataFrame(magic_spell_list)
        dt.set_index("Magic", inplace=True)
        spell_power = dt.loc[name]["spell power"]
        magic_heal = (self.level + self.magic_atk * spell_power)
        self.magic_heal = magic_heal
        return magic_heal

    # Getting mp cost for the magic
    def x_magic_cost(self, name):
        magic_spell_list = pd.read_excel(dir_path.magic_path + magic_file)
        dt = pd.DataFrame(magic_spell_list)
        dt.set_index("Magic", inplace=True)
        spell_power = dt.loc[name]["mp cost"]
        self.magic_cost = spell_power
        return spell_power

    # Item money back when selling, from excel file, by qty of user_input
    def x_item_selling(self, name, shop_name, count):
        YourHero.buying_habits(self, shop_name, count, name)


        # in excel file, make temp buying count, erase and save, not use dt.to_excel!!
        item_list_for_index = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
        wb = openpyxl.load_workbook(dir_path.inventory_items_path + inventory_item_file)
        ws = wb[shop_name]
        row_numbers = item_list_for_index[item_list_for_index['Items'] == name].index[0]
        row_numbers = row_numbers + 2
        qty_column = ws.cell(row=row_numbers, column=3)
        temp_qty = ws.cell(row=row_numbers, column=4)
        temp_qty.value = int(count)
        qty_new_value = qty_column.value - temp_qty.value
        qty_column.value = qty_new_value
        temp_qty.value = None
        wb.save(dir_path.inventory_items_path + inventory_item_file)

    def x_item_quantity(self, name, shop_name):
        item_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
        dt = pd.DataFrame(item_list)
        dt.set_index("Items", inplace=True)
        item_quantity = dt.loc[name]["Quantity"]
        store_quantity = item_quantity
        self.store_quantity = store_quantity
        return store_quantity

    def x_item_cost(self, name, shop_name):
        item_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
        dt = pd.DataFrame(item_list)
        dt.set_index("Items", inplace=True)
        item_cost = dt.loc[name]["Cost"]
        store_cost = item_cost
        self.store_cost = store_cost
        return store_cost

    def x_money_back_store(self, name, shop_name):
        item_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
        dt = pd.DataFrame(item_list)
        dt.set_index("Items", inplace=True)
        item_cost = dt.loc[name]["Cost"]
        store_cost = item_cost
        self.store_money_back = store_cost / 2
        return store_cost

    def x_time_enter_shop(self, shop_name):
        if shop_name == "Abreheim's":
            cloud.abreheim_enter += 1
            if cloud.abreheim_enter >= 4:
                cloud.abreheim_enter = 0
                item_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
                dt = pd.DataFrame(item_list)
                dt.set_index("Items", inplace=True)
                dt_list_items = ["potion", "ether", "antidote", "phoenix_down", "silver_Dust", "tent"]
                for i in dt_list_items:
                    dt.loc[i]["Quantity"] = 5
                dt.to_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)

    def x_potion_hp(self):
        potion = self.max_hp * 0.3
        self.potion_cure = potion
        return potion

    def x_ether_mp(self):
        ether = self.max_mp * 0.3
        self.ether_cure = ether
        return ether

    def x_elixir(self):
        elixir = self.max_hp and self.max_mp
        self.elixir_cure = elixir
        return elixir

    def x_potion_use(self, times):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:
            x = times * self.potion_cure
            if self.hp + x >= self.max_hp:
                self.hp = self.max_hp
                return x
            elif self.hp + x <= self.max_hp:
                self.hp = self.hp + x
                return x

    def x_ether_use(self, times):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:
            x = times * self.elixir_cure
            if self.mp + x > self.max_mp:
                self.mp = self.max_mp
                return x

            elif self.mp + x <= self.max_mp:
                self.mp = self.mp + x
                return x

    def x_elixir_use(self):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:
            if self.hp + self.elixir_cure > self.max_hp:
                self.hp = self.max_hp
            elif self.hp + self.elixir_cure <= self.max_hp:
                self.hp = self.hp + self.elixir_cure
            if self.mp + self.elixir_cure > self.max_mp:
                self.mp = self.max_mp
            elif self.mp + self.elixir_cure <= self.max_mp:
                self.mp = self.mp + self.elixir_cure

    def x_silver_dust_use(self):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:

            if self.status_paralyzed > 0:
                self.status_paralyzed = 0

    def x_antidote_use(self):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:
            if self.status_poison >= 1:
                self.status_poison = 0

    def x_phoenix_down_use(self):
        if self.hp <= 0:
            self.hp = self.max_hp

    def x_tent_use(self):
        if self.hp <= 0:
            print("Sad to say, but the hero is dead, only Phoenix Down can save you")
            print("---------------------------------------------------------")
            input("press enter")
        else:
            self.hp = self.max_hp
            self.mp = self.max_mp
            self.status_poison = 0
            self.status_paralyzed = 0
            self.status_effect_name = ""
            self.status_effect_2_name = ""

    def x_status_effect_name(self):
        if self.status_poison == 1:
            self.status_effect_name = "Poison"
        elif self.status_poison == 0:
            self.status_effect_name = ""
        if self.status_paralyzed == 1:
            self.status_effect_2_name = "Paralyzed"
        elif self.status_paralyzed == 0:
            self.status_effect_2_name = ""
            return "Poison" and "Paralyzed" and ""

    # Add items to inventory
    def x_add_items_inventory(self, give_item):
        give_item_list = pd.DataFrame(give_item)
        # print(give_item_list)
        item_list = pd.DataFrame(self.inventory)
        var = len(self.inventory["Item"])
        # give_var = len(give_item["Item"])
        for i in give_item["Item"]:
            if i in item_list.values:
                item_index = item_list[item_list["Item"] == i].index.values
                res = str(item_index)[1:-1]
                give_index = give_item_list[give_item_list["Item"] == i].index.values
                give_res = str(give_index)[1:-1]
                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] + give_item["QTY"][int(give_res)]
            elif i not in item_list.values:
                give_index = give_item_list[give_item_list["Item"] == i].index.values
                give_res = str(give_index)[1:-1]
                # count_list = len(give_item["Item"])
                # for x in range(0, count_list):

                self.inventory["Item"].insert(var + 1, give_item["Item"][int(give_res)])
                self.inventory["QTY"].insert(var + 1, give_item["QTY"][int(give_res)])

    # If mugged or falling from cliff loosing items in some way:
    def x_random_mugging(self):
        mugg = random.randint(0, 10)
        if mugg >= 5:
            stealing = ["gil", "potion", "ether", "silver_dust", "phoenix_down", "tent"]
            stealing_qty = [random.randint(150, 400), random.randint(2, 5), random.randint(1, 2),
                            random.randint(1, 2), random.randint(1, 2), random.randint(1, 2)]
            item_chance_mugging = random.randint(0, 17)
            if item_chance_mugging <= 10:
                mugg_item = random.randint(0, 1)
                if mugg_item == 0:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 1:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
            elif item_chance_mugging > 11 <= 15:
                mugg_item = random.randint(2, 4)
                if mugg_item == 2:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 3:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 4:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
            elif item_chance_mugging >= 16:
                mugg_item = random.randint(0, 5)
                if mugg_item == 1:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 2:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 3:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 4:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
                elif mugg_item == 5:
                    give_item = {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]}
                    return give_item
        else:
            give_item = {"Item": [""], "QTY": [""]}
            return give_item

    # Removing items from inventory
    def x_remove_items_inventory(self, give_item, thief, hero, loose):
        give_item_list = pd.DataFrame(give_item)
        item_list = pd.DataFrame(hero.inventory)
        var = len(hero.inventory["Item"])
        empty = []
        no_integers = []
        yes_integers = []
        str_index = []
        int_index_list = []

        if loose == "yes":
            for i in give_item["Item"]:
                if i in item_list.values:
                    item_index = item_list[item_list["Item"] == i].index.values
                    res = str(item_index)[1:-1]
                    give_index = give_item_list[give_item_list["Item"] == i].index.values
                    give_res = str(give_index)[1:-1]
                    loose = hero.inventory["QTY"][int(res)]
                    hero.inventory["QTY"][int(res)] = hero.inventory["QTY"][int(res)] - give_item["QTY"][int(give_res)]
                    if hero.inventory["QTY"][int(res)] < 0:
                        hero.inventory["QTY"][int(res)] = 0
                        if thief == "yes":
                            print("The thief stool from", hero.name, " you loose", loose, i)
                        elif thief == "no":
                            print(hero.name, " you loose", loose, i)
                    elif hero.inventory["QTY"][int(res)] >= 0:
                        if thief == "yes":
                            print("The thief stool from", hero.name, "you loose", give_item["QTY"][int(give_res)], i)
                        elif thief == "no":
                            print(hero.name, " you loose", give_item["QTY"][int(give_res)], i)
                elif i not in item_list.values:
                    if thief == "yes":
                        if i == "":
                            print(hero.name, "The thief tried too take something but could not")
                        else:
                            print(hero.name, "The thief tried too take", i, "but you didn't have any")
                    elif thief == "no":
                        print(hero.name, "You are lucky, didn't loose anything")
        elif loose == "no":
            for q in give_item_list["QTY"].values:
                empty.append(q)
                no_integers = [x for x in empty if not isinstance(x, int)]
                yes_integers = [x for x in empty if isinstance(x, int)]
            for no_int in no_integers:
                index = give_item_list[give_item_list["QTY"] == no_int].index.values
                res = str(index)[1:-1]
                item_use = give_item_list["Item"][int(res)]
                item_index = item_list[item_list["Item"] == item_use].index.values
                res_item = str(item_index)[1:-1]
                str_index.append(res_item)
                hero.inventory["Item"].pop(int(str_index[0]))
                hero.inventory["QTY"].pop(int(str_index[0]))
            for number_number in yes_integers:
                index = give_item_list[give_item_list["QTY"] == number_number].index.values
                res = str(index)[1:-1]
                int_index_list = [x for x in res]
                int_index_list = [x.replace(" ", "") for x in res]
                int_index_list = list(filter(None, int_index_list))
                for q in int_index_list:
                    hero.inventory["QTY"][int(q)] = hero.inventory["QTY"][int(q)] - give_item["QTY"][int(q)]

    # view total stats of all heroes
    def x_stat_all_hero(self):
        item_list = pd.DataFrame(self.inventory)
        item_use = "gil"
        item_index = item_list[item_list["Item"] == item_use].index.values
        res = str(item_index)[1:-1]

        my_timer.print_time()

        dots = ":     "
        name_dot = self.name + dots

        skills = {name_dot: [""],
                  "Attributes": ["Level", "EXP", "GIL", "HP", "Atk", "Strength", "Def",
                                 "Magic", "MP", "Spirit", "Magic Atk", "Status Effect",
                                 "Status Effect", "Time elapsed"],
                  "Points": [self.level, self.exp_show, self.inventory["QTY"][int(res)], self.hp, self.atk,
                             self.strength, self.defence,
                             self.magic, self.mp, self.spirit,
                             self.magic_atk, self.status_effect_name, self.status_effect_2_name,
                             my_timer.time_stat],
                  "Max Points": ["", "", "", self.max_hp, "", "", "", "", self.max_mp,
                                 "", "", "", "", ""]

                  }
        matrix_total = pd.DataFrame(skills, index=["", "", "", "", "", "", "", "", "", "", "",
                                                   "", "", ""])

        self.total_stat = matrix_total
        return matrix_total

    # For Battles view stats

    def x_stat_load(self):
        item_list = pd.DataFrame(self.inventory)
        item_use = "gil"
        item_index = item_list[item_list["Item"] == item_use].index.values
        res = str(item_index)[1:-1]

        dots = ":     "
        name_dot = "Stats" + dots

        skills = {name_dot: [""],
                  "Name": [cloud.name, elena.name],
                  "Level": [cloud.level, elena.level],
                  "HP": [cloud.hp, elena.hp],
                  "MP": [cloud.mp, elena.mp],
                  "GIL": [cloud.inventory["QTY"][int(res)], elena.inventory["QTY"][int(res)]],
                  "Chapter": [cloud.chapter, elena.chapter],
                  "Place": [cloud.story_name, elena.story_name],
                  "Time": [my_timer.time_stat, my_timer.time_stat],

                  }

        # matrix_skills = pd.DataFrame(skills)
        # matrix_skills = matrix_skills.rename(index=lambda x: "")
        matrix_skills = pd.DataFrame(skills, index=["", ""])
        self.stat = matrix_skills
        return matrix_skills

    def x_stat(self):

        dots = ":     "
        name_dot = self.name + dots

        skills = {name_dot: [""],
                  "Attributes": ["Level", "HP", "MP"],
                  "Points": [self.level, self.hp, self.mp],
                  "Max Points": ["", self.max_hp, self.max_mp],

                  }
        matrix_skills = pd.DataFrame(skills, index=["", "", ""])

        self.stat = matrix_skills
        return matrix_skills

    def x_exit_game(self):
        print("---------------------------------------------------------")
        print("Goodbye!")
        print("---------------------------------------------------------")
        time.sleep(2)
        sys.exit()

    def x_battle_win_3_stat(self):

        enemy_name = enemy.e_adjectives + " " + enemy.e_name
        enemy_name_2 = enemy_2.e_adjectives + " " + enemy_2.e_name
        enemy_name_3 = enemy_3.e_adjectives + " " + enemy_3.e_name
        dots = ":     "
        name_dot = self.name + dots

        skills = {name_dot: [""],
                  "Battle  WON!! ": ["EXP", "AP", "GIL"],
                  enemy_name: [enemy.e_exp_give, enemy.e_exp_give / 2, enemy.e_gil],
                  enemy_name_2: [enemy_2.e_exp_give, enemy_2.e_exp_give / 2, enemy_2.e_gil],
                  enemy_name_3: [enemy_3.e_exp_give, enemy_3.e_exp_give / 2, enemy_3.e_gil]

                  }
        matrix_enemy_3 = pd.DataFrame(skills, index=["", "", ""])

        self.win_stat_3 = matrix_enemy_3
        return matrix_enemy_3

    def x_battle_win_2_stat(self):
        enemy_name = enemy.e_adjectives + " " + enemy.e_name
        enemy_name_2 = enemy_2.e_adjectives + " " + enemy_2.e_name
        dots = ":     "
        name_dot = self.name + dots

        skills = {name_dot: [""],
                  "Battle  WON!! ": ["EXP", "AP", "GIL"],
                  enemy_name: [enemy.e_exp_give, enemy.e_exp_give / 2, enemy.e_gil],
                  enemy_name_2: [enemy_2.e_exp_give, enemy_2.e_exp_give / 2, enemy_2.e_gil]

                  }
        matrix_enemy_2 = pd.DataFrame(skills, index=["", "", ""])

        self.win_stat_2 = matrix_enemy_2
        return matrix_enemy_2

    def x_battle_win_1_stat(self):

        enemy_name = enemy.e_adjectives + " " + enemy.e_name
        dots = ":     "
        name_dot = self.name + dots

        skills = {name_dot: [""],
                  "Battle  WON!! ": ["EXP", "AP", "GIL"],
                  enemy_name: [enemy.e_exp_give, enemy.e_exp_give / 2, enemy.e_gil]
                  }
        matrix_enemy_1 = pd.DataFrame(skills, index=["", "", ""])

        self.win_stat = matrix_enemy_1
        return matrix_enemy_1

    # function for see weapon name, if you have ring equipped and what protection
    # also for see qty of an item, show == "item" "weapon", "ring", item_select == 0 if other then item
    # else name of the item to check for
    def x_see_inventory_name_qty(self, show, item_select):
        # self.armory_inventory = {"Weapon/Armor": ["blade storm", "sand storm"], "Magic Slots": [1, 2],
        # "weapon power": [6, 4],
        # "Protection": [0, 0], "Equipped": ["True", "True"], "Type": ["weapon", "armor"]}
        # item_list = pd.DataFrame(self.inventory)
        if show == "weapon":
            weapon_list = pd.DataFrame(self.armory_inventory)
            weapon_list.loc[(weapon_list['Equipped'] == 'True') &
                            (weapon_list['Type'] == 'weapon'), 'weapon_use'] = weapon_list["Weapon/Armor"]
            # weapon_list[weapon_list.notnull().all(1)]
            weapon_list_new = weapon_list[weapon_list.notnull().all(1)]
            index_weapon_list = weapon_list_new.index.values
            return self.armory_inventory["Weapon/Armor"][int(index_weapon_list)]
        elif show == "item":
            item_list = pd.DataFrame(self.inventory)
            item_use = item_select
            item_index = item_list[item_list["Item"] == item_use].index.values
            res = str(item_index)[1:-1]
            return self.inventory["QTY"][int(res)]
        elif show == "ring":
            if item_select == 0:
                ring_list = pd.DataFrame(self.armory_inventory)
                ring_list.loc[(ring_list['Equipped'] == 'True') &
                              (ring_list['Type'] == 'ring'), 'ring_use'] = ring_list["Weapon/Armor"]
                # weapon_list[weapon_list.notnull().all(1)]
                ring_list_new = ring_list[ring_list.notnull().all(1)]
                index_list_new = ring_list_new.index.values

                for i in index_list_new:
                    if "poison" in self.armory_inventory["Protection"][i]:
                        poison = "poison"
                        return poison
                    elif "paralyzed" in self.armory_inventory["Protection"][i]:
                        paralyzed = "paralyzed"
                        return paralyzed

            elif item_select == 1:
                ring_list = pd.DataFrame(self.armory_inventory)
                ring_list.loc[(ring_list['Equipped'] == 'True') &
                              (ring_list['Type'] == 'ring'), 'ring_use'] = ring_list["Weapon/Armor"]
                # weapon_list[weapon_list.notnull().all(1)]
                ring_list_new = ring_list[ring_list.notnull().all(1)]
                index_list_new = ring_list_new.index.values

                for i in index_list_new:
                    if "The mark of the serpents" in self.armory_inventory["Weapon/Armor"][i] or "Silver ring" \
                            in self.armory_inventory["Weapon/Armor"][i]:
                        snake_ring = "The mark of the serpents"
                        silver_ring = "Silver ring"

                        return snake_ring, silver_ring
                    else:
                        return "Check code!!"
            else:
                return "Check code!!"

    # View all items per Hero
    def x_items_list(self):
        dots = ":     "
        name_dot = self.name + dots
        item_list = pd.DataFrame(self.inventory)
        item_list = item_list.rename(index=lambda x: "")
        item_list = item_list.rename_axis(name_dot, axis=1)
        return item_list

    def x_battle_white_magic_use(self):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint("White Magic")
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            magic_list = pd.DataFrame(self.magic_spells)
            magic_list = magic_list.rename(index=lambda x: "")
            is_true = (magic_list["Equipped"] == "True") & (magic_list["Magic Type"] == "white magic")
            magic_list_true = magic_list[is_true]
            choice = input("Do you want to use magic, or exit inventory?")
            print("---------------------------------------------------------")
            if choice.lower() == "use magic" or choice.lower() == "magic":
                if self.magic_spells == {"Magic": [], "mp cost": [], "AP": [], "Level": [], "Equipped": [],
                                         "Magic Type": []}:
                    print("You don't have any magic")
                    print("Visit a material shop and buy some magic")
                    print("---------------------------------------------------------")
                elif len(magic_list_true) == 0:
                    print("No magic is added to you magic slots!")
                else:
                    magic_list = pd.DataFrame(self.magic_spells)
                    magic_list = magic_list.rename(index=lambda x: "")
                    is_true = (magic_list["Equipped"] == "True") & (magic_list["Magic Type"] == "white magic")
                    magic_list_true = magic_list[is_true]
                    print("---------------------------------------------------------")
                    print(self.name, "Your magic")
                    print("---------------------------------------------------------")
                    print(magic_list_true.head())
                    print("---------------------------------------------------------")
                    magic_select = input("Choose a spell, 'all lower cases please!!' ")
                    if magic_list_true.Magic.isin([magic_select]).any:
                        YourHero.x_magic_cost(self, magic_select)
                        if self.mp < self.magic_cost:
                            print("Sorry you don't have enough MP")
                            input("press enter")
                            print("---------------------------------------------------------")
                            YourHero.x_battle_two(self)
                        else:
                            if magic_select.lower() == "heal" or magic_select.lower() == "heal2":
                                print("Select whom you want to use it on", cloud.name, "or", elena.name)
                                print("---------------------------------------------------------")
                                player_select = input()
                                if player_select == cloud.name and cloud.hp > 0:
                                    give_to = cloud
                                    YourHero.x_magic_heal(self, magic_select)
                                    YourHero.x_magic_cost(self, magic_select)
                                    if give_to.hp + self.magic_heal > give_to.max_hp:
                                        give_to.hp = give_to.max_hp
                                        self.mp = self.mp - self.magic_cost
                                        print("You cast", magic_select, "at the", give_to.name,
                                              "with healing power", self.magic_heal)
                                        print("---------------------------------------------------------")
                                        break
                                    elif give_to.hp + self.magic_heal <= give_to.max_hp:
                                        give_to.hp = give_to.hp + self.magic_heal
                                        self.mp = self.mp - self.magic_cost
                                        print("You cast", magic_select, "at the", give_to.name,
                                              "with healing power", self.magic_heal)
                                        print("---------------------------------------------------------")
                                        break
                                elif player_select == elena.name and elena.hp > 0:
                                    give_to = elena
                                    YourHero.x_magic_heal(self, magic_select)
                                    YourHero.x_magic_cost(self, magic_select)
                                    if give_to.hp + self.magic_heal > give_to.max_hp:
                                        give_to.hp = give_to.max_hp
                                        self.mp = self.mp - self.magic_cost
                                        print("You cast", magic_select, "at the", give_to.name,
                                              "with healing power", self.magic_heal)
                                        print("---------------------------------------------------------")
                                        break
                                    elif give_to.hp + self.magic_heal <= give_to.max_hp:
                                        give_to.hp = give_to.hp + self.magic_heal
                                        self.mp = self.mp - self.magic_cost
                                        print("You cast", magic_select, "at the", give_to.name,
                                              "with healing power", self.magic_heal)
                                        print("---------------------------------------------------------")
                                        break
                                else:
                                    print("Player is dead, need to use phoenix down or magic life")
                                    print("---------------------------------------------------------")
                            elif magic_select.lower() == "life":
                                print("Select whom you want to use it on", cloud.name, "or", elena.name)
                                print("---------------------------------------------------------")
                                player_select = input()
                                if player_select == cloud.name and cloud.hp < 0:
                                    give_to = cloud
                                    give_to.hp = give_to.max_hp
                                    YourHero.x_magic_cost(self, magic_select)
                                    self.mp = self.mp - self.magic_cost
                                    print("You cast", magic_select, "at the", give_to.name,
                                          "and hero returns back to life with full HP")
                                    print("---------------------------------------------------------")
                                    break
                                elif player_select == elena.name and elena.hp < 0:
                                    give_to = elena
                                    give_to.hp = give_to.max_hp
                                    YourHero.x_magic_cost(self, magic_select)
                                    self.mp = self.mp - self.magic_cost
                                    print("You cast", magic_select, "at the", give_to.name,
                                          "and hero returns back to life with full HP")
                                    print("---------------------------------------------------------")
                                    break
                                else:
                                    print("Player is not dead, use a healing magic or potions")
                                    print("---------------------------------------------------------")
                    else:
                        print("Choose a valid magic")
                        print("---------------------------------------------------------")
            elif choice.lower() == "exit" or choice.lower() == "exit inventory":
                YourHero.x_battle_two(self)
                break

    def x_battle_black_magic_use(self, opponent):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint("Black Magic")
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            magic_list = pd.DataFrame(self.magic_spells)
            magic_list = magic_list.rename(index=lambda x: "")
            is_true = (magic_list["Equipped"] == "True") & (magic_list["Magic Type"] == "black magic")
            magic_list_true = magic_list[is_true]
            choice = input("Do you want to use magic, or exit inventory?")
            print("---------------------------------------------------------")
            if choice.lower() == "use magic" or choice.lower() == "magic":
                if self.magic_spells == {"Magic": [], "mp cost": [], "AP": [], "Level": [],
                                         "Equipped": [], "Magic Type": []}:
                    print("You don't have any magic")
                    print("Visit a material shop and buy some magic")
                    print("---------------------------------------------------------")
                elif len(magic_list_true) == 0:
                    print("No magic is added to you magic slots!")
                else:
                    magic_list = pd.DataFrame(self.magic_spells)
                    magic_list = magic_list.rename(index=lambda x: "")
                    is_true = (magic_list["Equipped"] == "True") & (magic_list["Magic Type"] == "black magic")
                    magic_list_true = magic_list[is_true]
                    print("---------------------------------------------------------")
                    print(self.name, "Your magic")
                    print("---------------------------------------------------------")
                    print(magic_list_true.head())
                    print("---------------------------------------------------------")
                    magic_select = input("Choose a spell, 'all lower cases please!!' ")
                    if magic_list_true.Magic.isin([magic_select]).any:
                        YourHero.x_magic_cost(self, magic_select)
                        if self.mp < self.magic_cost:
                            print("Sorry you don't have enough MP")
                            input("press enter")
                            print("---------------------------------------------------------")
                            YourHero.x_battle_two(self)
                        else:
                            hit = random.randint(0, 10)
                            if hit >= 2:
                                YourHero.x_magic_dmg(self, magic_select)
                                YourHero.x_magic_cost(self, magic_select)
                                opponent.e_hp = opponent.e_hp - self.cast_magic_atk
                                self.mp = self.mp - self.magic_cost
                                print("You cast", magic_select, "at the", opponent.e_adjectives, opponent.e_name,
                                      "with magic power", self.cast_magic_atk)
                                print("---------------------------------------------------------")
                                YourHero.x_initials_stats(self)
                                opponent.ex_initials_stats()
                                if opponent.e_hp <= 0:
                                    print("---------------------------------------------------------")
                                    print("Enemy", opponent.e_adjectives, opponent.e_name, "is dead")
                                    input("Press enter")
                                break
                            elif hit <= 1:
                                print("You miss the spell at the enemy", opponent.e_adjectives, opponent.e_name)
                                print("---------------------------------------------------------")
                                break
                    else:
                        print("Choose a valid magic")
                        print("---------------------------------------------------------")
            elif choice.lower() == "exit" or choice.lower() == "exit inventory":
                YourHero.x_battle_two(self)
                break

    def x_battle_item_use(self):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint("Battle Inventory")
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            choice = input("Do you want to use an item, or leave inventory?")
            print("---------------------------------------------------------")
            if choice.lower() == "use item" or choice.lower() == "item":
                item_list = pd.DataFrame(self.inventory)
                item_list = item_list.rename(index=lambda x: "")
                print("---------------------------------------------------------")
                print(YourHero.x_stat(cloud))
                print("---------------------------------------------------------")
                print(YourHero.x_stat(elena))
                print("---------------------------------------------------------")
                print(self.name, "Your items")
                print("---------------------------------------------------------")
                print(item_list)
                print("---------------------------------------------------------")
                print("Select whom you want to use it on", cloud.name, "or", elena.name)
                player_select = input()
                if player_select == cloud.name and cloud.hp > 0:
                    give_to = cloud
                    item_use = input("What do you want to use, all lower cases!?\n")
                    print("---------------------------------------------------------")
                    if item_use.lower() == "potion":
                        item_list = pd.DataFrame(self.inventory)
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.potion_cure * int(use_count) <= give_to.max_hp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_potion_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.potion_cure * int(use_count), "HP",
                                      give_to.name, "your HP is now", give_to.hp, "HP")
                                print("---------------------------------------------------------")
                            elif give_to.potion_cure * int(use_count) > give_to.max_hp:
                                use_count = give_to.max_hp / give_to.potion_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_potion_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.potion_cure * use_count, "HP",
                                      give_to.name, "your HP is now", give_to.hp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, item_use)
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "ether":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.ether_cure * int(use_count) <= give_to.max_mp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_ether_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.ether_cure * int(use_count), "MP", give_to.name, "your MP is now",
                                      give_to.mp)
                                print("---------------------------------------------------------")
                            elif give_to.ether_cure * int(use_count) > give_to.max_mp:
                                use_count = give_to.max_mp / give_to.ether_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_ether_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.ether_cure * use_count, "MP",
                                      give_to.name, "your MP is now", give_to.mp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Ethers")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "antidote":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_antidote_use(give_to)
                            print("You used an antidote and cured", give_to.name, "from poison")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any antidote")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "phoenix down":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "elixir":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_elixir(give_to)
                            print("You used an elixir")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any elixir")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "silver dust":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_silver_dust_use(give_to)
                            print("You used an silver dust and cured", give_to.name, "from paralysed")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any silver dust")
                            print("---------------------------------------------------------")
                    else:
                        print("You don't have that/or can't use in battle")
                        print("---------------------------------------------------------")
                elif player_select == cloud.name and cloud.hp <= 0:
                    give_to = cloud
                    print("Hero is dead you can only use a phoenix down")
                    item_use = input("Do you want to use one?")
                    if item_use.lower() == "yes" or item_use.lower() == "Yes":
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "phoenix down"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                elif player_select == elena.name and elena.hp > 0:
                    give_to = elena
                    item_use = input("What do you want to use?\n")
                    print("---------------------------------------------------------")
                    if item_use.lower() == "potion":
                        item_list = pd.DataFrame(self.inventory)
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.potion_cure * int(use_count) <= give_to.max_hp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_potion_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.potion_cure * int(use_count), "HP",
                                      give_to.name, "your HP is now", give_to.hp, "HP")
                                print("---------------------------------------------------------")
                            elif give_to.potion_cure * int(use_count) > give_to.max_hp:
                                use_count = give_to.max_hp / give_to.potion_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_potion_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.potion_cure * use_count, "HP",
                                      give_to.name, "your HP is now", give_to.hp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, item_use)
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "ether":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.ether_cure * int(use_count) <= give_to.max_mp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_ether_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.ether_cure * int(use_count), "MP", give_to.name, "your MP is now",
                                      give_to.mp)
                                print("---------------------------------------------------------")
                            elif give_to.ether_cure * int(use_count) > give_to.max_mp:
                                use_count = give_to.max_mp / give_to.ether_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_ether_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.ether_cure * use_count, "MP",
                                      give_to.name, "your MP is now", give_to.mp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Ethers")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "antidote":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_antidote_use(give_to)
                            print("You used an antidote and cured", give_to.name, "from poison")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any antidote")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "phoenix down":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "elixir":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_elixir(give_to)
                            print("You used an elixir")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any elixir")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "silver dust":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_silver_dust_use(give_to)
                            print("You used an silver dust and cured", give_to.name, "from paralysed")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any silver dust")
                            print("---------------------------------------------------------")
                    else:
                        print("You don't have that/or cant use in battle")
                        print("---------------------------------------------------------")
                elif player_select == elena.name and elena.hp <= 0:
                    give_to = elena
                    print("Hero is dead you can only use a phoenix down")
                    item_use = input("Do you want to use one?")
                    if item_use.lower() == "yes" or item_use.lower() == "Yes":
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "phoenix down"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
            elif choice.lower() == "leave" or choice.lower() == "leave inventory":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    YourHero.x_battle_two(self)
                    break

    def x_inventory(self):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint("Inventory")
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            print("Welcome to the", self.name, "inventory menu")
            print("---------------------------------------------------------")
            input("press enter")
            print("---------------------------------------------------------")
            choice = input("Equip, Unequipped, use item, magic, or leave?\n")
            if choice.lower() == "equip" or choice.lower() == "Equip":
                if self.armory_inventory == {"Weapon/Armor": [], "Magic Slots": [], "Protection": [],
                                             "Equipped": [], "Type": []}:
                    print("You don't have any weapons/armor")
                    print("Visit a weapon/armor shop and buy some magic")
                    print("---------------------------------------------------------")
                else:
                    weapon_list = pd.DataFrame(self.armory_inventory)
                    weapon_list = weapon_list.rename(index=lambda x: x + 1)
                    print("---------------------------------------------------------")
                    print(self.name, "Your weapon/armor")
                    print("---------------------------------------------------------")
                    print(weapon_list)
                    print("---------------------------------------------------------")
                    type_choice = input("What do you want to equipped, weapon, armor or ring?")
                    if type_choice == "weapon":
                        if self.weapon_equipped == self.weapon_equipped_max:
                            print("You already have a weapon equipped, choose remove and try again")
                            print("---------------------------------------------------------")
                        elif self.weapon_equipped < self.weapon_equipped_max:
                            is_true = (weapon_list["Type"] == "weapon")
                            weapon_list_true = weapon_list[is_true]
                            #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                            print("What weapon do you want to equipped?")
                            print(weapon_list_true)
                            weapon_choice = input("Select by number")
                            print("---------------------------------------------------------")
                            if weapon_choice.isdigit():
                                if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                    print("That weapon has already been equipped, choose another one")
                                elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                    self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                    weapon_inv = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                    # self.magic_spells["Equipped"].append(int(magic_choice) - 1
                                    self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "True")
                                    # magic_list.loc[int(magic_choice), 'Equipped'] = "True"
                                    print("You have now equipped", weapon_inv)
                                    print("---------------------------------------------------------")
                                    self.weapon_equipped += 1
                                    self.magic_slots_max += YourHero.x_weapon_shop_slots(self, weapon_inv)
                            else:
                                print("Please choose a number")
                                print("---------------------------------------------------------")
                    elif type_choice == "ring":
                        if self.rings_equipped == self.rings_equipped_max:
                            print("You already have a weapon equipped, choose remove and try again")
                            print("---------------------------------------------------------")
                        elif self.rings_equipped < self.rings_equipped_max:
                            is_true = (weapon_list["Type"] == "ring")
                            weapon_list_true = weapon_list[is_true]
                            #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                            print("What ring do you want to equipped?")
                            print(weapon_list_true)
                            weapon_choice = input("Select by number")
                            print("---------------------------------------------------------")
                            if weapon_choice.isdigit():
                                if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                    print("That ring has already been equipped, choose another one")
                                elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                    self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                    weapon_inv = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                    # self.magic_spells["Equipped"].append(int(magic_choice) - 1
                                    self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "True")
                                    # magic_list.loc[int(magic_choice), 'Equipped'] = "True"
                                    print("You have now equipped", weapon_inv)
                                    print("---------------------------------------------------------")
                                    self.rings_equipped += 1
                            else:
                                print("Please choose a number")
                                print("---------------------------------------------------------")
                    elif type_choice == "armor":
                        if self.armor_equipped == self.armor_equipped_max:
                            print("You already have a armor equipped, choose remove and try again")
                            print("---------------------------------------------------------")
                        elif self.armor_equipped < self.armor_equipped_max:
                            is_true = (weapon_list["Type"] == "armor")
                            weapon_list_true = weapon_list[is_true]
                            #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                            print("What armor do you want to equipped?")
                            print(weapon_list_true)
                            weapon_choice = input("Select by number")
                            print("---------------------------------------------------------")
                            if weapon_choice.isdigit():
                                if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                    print("That armor has already been equipped, choose another one")
                                elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                    self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                    weapon_inv = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                    # self.magic_spells["Equipped"].append(int(magic_choice) - 1
                                    self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "True")
                                    # magic_list.loc[int(magic_choice), 'Equipped'] = "True"
                                    print("You have now equipped", weapon_inv)
                                    print("---------------------------------------------------------")
                                    self.rings_equipped += 1
                                    self.magic_slots_max += YourHero.x_weapon_shop_slots(self, weapon_inv)
                            else:
                                print("Please choose a number")
                                print("---------------------------------------------------------")
            elif choice.lower() == "Unequipped" or choice.lower() == "unequipped":
                if self.armory_inventory == {"Weapon/Armor": [], "Magic Slots": [], "Protection": [],
                                             "Equipped": [], "Type": []}:
                    print("You don't have any weapons/armor")
                    print("Visit a weapon/armor shop and buy some magic")
                    print("---------------------------------------------------------")
                else:
                    weapon_list = pd.DataFrame(self.armory_inventory)
                    weapon_list = weapon_list.rename(index=lambda x: x + 1)
                    print("---------------------------------------------------------")
                    print(self.name, "Your weapon/armor")
                    print("---------------------------------------------------------")
                    print(weapon_list)
                    print("---------------------------------------------------------")
                    type_choice = input("What do you want to unequipped, weapon, armor or ring?")
                    spells_count = 0
                    if type_choice == "weapon":
                        is_true = (weapon_list["Type"] == "weapon")
                        weapon_list_true = weapon_list[is_true]
                        #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                        print("What weapon do you want to unequipped?")
                        print(weapon_list_true)
                        weapon_choice = input("Select by number")
                        print("---------------------------------------------------------")
                        if weapon_choice.isdigit():
                            if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                print("That is already unequipped")
                                print("---------------------------------------------------------")
                            elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                first_value = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "False")
                                print("You have now removed", first_value)
                                print("---------------------------------------------------------")
                                self.weapon_equipped -= 1
                                self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, first_value)
                                magic_delete = self.magic_slots_max - self.magic_slots
                                if magic_delete <= 0:
                                    magic_spell_list = pd.DataFrame(self.magic_spells)
                                    name_count = len(magic_spell_list["Equipped"])
                                    positivenum = abs(magic_delete)
                                    self.magic_slots -= YourHero.x_weapon_shop_slots(self, first_value)
                                    self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, first_value)
                                    print(positivenum, "Magics have been removed go to magic inventory")
                                    print("---------------------------------------------------------")
                                    for z in range(name_count):
                                        if magic_spell_list["Equipped"].values[z] == "True":
                                            spells_count += 1
                                            self.magic_spells["Equipped"].pop(z)
                                            self.magic_spells["Equipped"].insert(z, "False")
                                            if spells_count == positivenum:
                                                break
                                else:
                                    print("Your magic will stay where they are")
                                    print("---------------------------------------------------------")
                        else:
                            print("Please choose a number")
                            print("---------------------------------------------------------")
                    elif type_choice == "armor":
                        is_true = (weapon_list["Type"] == "armor")
                        weapon_list_true = weapon_list[is_true]
                        #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                        print("What armor do you want to unequipped?")
                        print(weapon_list_true)
                        weapon_choice = input("Select by number")
                        print("---------------------------------------------------------")
                        if weapon_choice.isdigit():
                            if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                print("That is already unequipped")
                            elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                first_value = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "False")
                                print("You have now removed", first_value)
                                self.armor_equipped -= 1
                                self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, first_value)
                                magic_delete = self.magic_slots_max - self.magic_slots
                                if magic_delete <= 0:
                                    magic_spell_list = pd.DataFrame(self.magic_spells)
                                    name_count = len(magic_spell_list["Equipped"])
                                    positivenum = abs(magic_delete)
                                    self.magic_slots -= YourHero.x_weapon_shop_slots(self, first_value)
                                    self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, first_value)
                                    print(positivenum, "Magics have been removed go to magic inventory")
                                    print("---------------------------------------------------------")
                                    for z in range(name_count):
                                        if magic_spell_list["Equipped"].values[z] == "True":
                                            spells_count += 1
                                            self.magic_spells["Equipped"].pop(z)
                                            self.magic_spells["Equipped"].insert(z, "False")
                                            if spells_count == positivenum:
                                                break
                                else:
                                    print("Your magic will stay where they are")
                        else:
                            print("Please choose a number")
                            print("---------------------------------------------------------")
                    elif type_choice == "ring":
                        is_true = (weapon_list["Type"] == "ring")
                        weapon_list_true = weapon_list[is_true]
                        #  weapon_list_true = weapon_list_true.rename(index=lambda x: x + 1)
                        print("What armor do you want to unequipped?")
                        print(weapon_list_true)
                        weapon_choice = input("Select by number")
                        print("---------------------------------------------------------")
                        if weapon_choice.isdigit():
                            if weapon_list['Equipped'].values[int(weapon_choice) - 1] == "False":
                                print("That is already unequipped")
                            elif weapon_list['Equipped'].values[int(weapon_choice) - 1] == "True":
                                self.armory_inventory["Equipped"].pop(int(weapon_choice) - 1)
                                first_value = weapon_list['Weapon/Armor'].values[int(weapon_choice) - 1]
                                self.armory_inventory["Equipped"].insert(int(weapon_choice) - 1, "False")
                                print("You have now removed", first_value)
                                self.rings_equipped -= 1
            elif choice.lower() == "magic":
                if self.magic_spells == {"Magic": [], "mp cost": [], "AP": [], "Level": [], "Equipped": []}:
                    print("You don't have any magic")
                    print("Visit a material shop and buy some magic")
                    print("---------------------------------------------------------")
                else:
                    magic_list = pd.DataFrame(self.magic_spells)
                    magic_list = magic_list.rename(index=lambda x: x + 1)
                    print("---------------------------------------------------------")
                    print(self.name, "Your magic, magic slots:", self.magic_slots, "/", self.magic_slots_max)
                    print("---------------------------------------------------------")
                    print(magic_list)
                    print("---------------------------------------------------------")
                    add_choice = input("Do you want to add materia or remove?")
                    if add_choice == "add" or add_choice.lower() == "add materia":
                        if self.magic_slots == self.magic_slots_max:
                            print("Sorry your magic slots are full, need to remove before adding")
                            print("Or you need to buy a weapon that has magic slots")
                            print("---------------------------------------------------------")
                        elif self.magic_slots < self.magic_slots_max:
                            print("What magic do you want to add?")
                            magic_choice = input("Select by number")
                            if magic_choice.isdigit():
                                if magic_list['Equipped'].values[int(magic_choice) - 1] == "True":
                                    print("That materia has already been equipped, choose another one")
                                elif magic_list['Equipped'].values[int(magic_choice) - 1] == "False":
                                    self.magic_spells["Equipped"].pop(int(magic_choice) - 1)
                                    magic_inv = magic_list['Magic'].values[int(magic_choice) - 1]
                                    # self.magic_spells["Equipped"].append(int(magic_choice) - 1
                                    self.magic_spells["Equipped"].insert(int(magic_choice) - 1, "True")
                                    # magic_list.loc[int(magic_choice), 'Equipped'] = "True"
                                    print("You have now equipped", magic_inv)
                                    self.magic_slots += 1
                            else:
                                print("Please choose a number")
                    elif add_choice.lower() == "remove" or add_choice.lower() == "remove materia":
                        print("What magic do you want to remove?")
                        magic_choice = input("Select by number")
                        if magic_choice.isdigit():
                            if magic_list['Equipped'].values[int(magic_choice) - 1] == "False":
                                print("That materia is already unequipped")
                            elif magic_list['Equipped'].values[int(magic_choice) - 1] == "True":
                                self.magic_spells["Equipped"].pop(int(magic_choice) - 1)
                                first_value = magic_list['Magic'].values[int(magic_choice) - 1]
                                self.magic_spells["Equipped"].insert(int(magic_choice) - 1, "False")
                                print("You have now removed", first_value)
                                self.magic_slots -= 1
                        else:
                            print("Please choose a number")
                    else:
                        print("Don't understand what you want")
            elif choice.lower() == "use item" or choice.lower() == "item":
                item_list = pd.DataFrame(self.inventory)
                item_list = item_list.rename(index=lambda x: "")
                print("---------------------------------------------------------")
                print(YourHero.x_stat(cloud))
                print("---------------------------------------------------------")
                print(YourHero.x_stat(elena))
                print("---------------------------------------------------------")
                print(self.name, "Your items")
                print("---------------------------------------------------------")
                print(item_list)
                print("---------------------------------------------------------")
                print("Select whom you want to use it on", cloud.name, "or", elena.name)
                player_select = input()
                if player_select == cloud.name and cloud.hp > 0:
                    give_to = cloud
                    item_use = input("What do you want to use, all lower cases!?\n")
                    print("---------------------------------------------------------")
                    if item_use.lower() == "potion":
                        item_list = pd.DataFrame(self.inventory)
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.potion_cure * int(use_count) <= give_to.max_hp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_potion_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.potion_cure * int(use_count), "HP",
                                      give_to.name, "your HP is now", give_to.hp, "HP")
                                print("---------------------------------------------------------")
                            elif give_to.potion_cure * int(use_count) > give_to.max_hp:
                                use_count = give_to.max_hp / give_to.potion_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_potion_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.potion_cure * use_count, "HP",
                                      give_to.name, "your HP is now", give_to.hp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, item_use)
                            print("---------------------------------------------------------")
                    elif item_use == "gil":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How much do you want to give?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                            give_to.inventory["QTY"][int(res)] = give_to.inventory["QTY"][int(res)] + int(use_count)
                            print("You gave", use_count, "GIL", "to", give_to.name)
                            print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Gil")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "ether":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.ether_cure * int(use_count) <= give_to.max_mp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_ether_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.ether_cure * int(use_count), "MP", give_to.name, "your MP is now",
                                      give_to.mp)
                                print("---------------------------------------------------------")
                            elif give_to.ether_cure * int(use_count) > give_to.max_mp:
                                use_count = give_to.max_mp / give_to.ether_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_ether_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.ether_cure * use_count, "MP",
                                      give_to.name, "your MP is now", give_to.mp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Ethers")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "antidote":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_antidote_use(give_to)
                            print("You used an antidote and cured", give_to.name, "from poison")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any antidote")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "phoenix down":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "elixir":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_elixir(give_to)
                            print("You used an elixir")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any elixir")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "silver dust":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_silver_dust_use(give_to)
                            print("You used an silver dust and cured", give_to.name, "from paralysed")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any silver dust")
                            print("---------------------------------------------------------")
                    else:
                        print("You don't have that, what do you want to use?")
                        print("---------------------------------------------------------")
                elif player_select == cloud.name and cloud.hp <= 0:
                    give_to = cloud
                    print("Hero is dead you can only use a phoenix down")
                    item_use = input("Do you want to use one?")
                    if item_use.lower() == "yes" or item_use.lower() == "Yes":
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "phoenix down"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                elif player_select == elena.name and elena.hp > 0:
                    give_to = elena
                    item_use = input("What do you want to use, all lower cases!?\n")
                    print("---------------------------------------------------------")
                    if item_use.lower() == "potion":
                        item_list = pd.DataFrame(self.inventory)
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.potion_cure * int(use_count) <= give_to.max_hp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_potion_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.potion_cure * int(use_count), "HP",
                                      give_to.name, "your HP is now", give_to.hp, "HP")
                                print("---------------------------------------------------------")
                            elif give_to.potion_cure * int(use_count) > give_to.max_hp:
                                use_count = give_to.max_hp / give_to.potion_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_potion_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.potion_cure * use_count, "HP",
                                      give_to.name, "your HP is now", give_to.hp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, item_use)
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "gil":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How much do you want to give?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                            give_to.inventory["QTY"][int(res)] = give_to.inventory["QTY"][int(res)] + int(use_count)
                            print("You gave", use_count, "GIL", "to", give_to.name)
                            print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Gil")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "ether":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        use_count = input("How many do you want to use?\n")
                        print("---------------------------------------------------------")
                        if int(use_count) <= self.inventory["QTY"][int(res)]:
                            if give_to.ether_cure * int(use_count) <= give_to.max_mp:
                                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(use_count)
                                YourHero.x_ether_use(give_to, int(use_count))
                                print("You used", use_count, item_use, "and", give_to.name, "got",
                                      give_to.ether_cure * int(use_count), "MP", give_to.name, "your MP is now",
                                      give_to.mp)
                                print("---------------------------------------------------------")
                            elif give_to.ether_cure * int(use_count) > give_to.max_mp:
                                use_count = give_to.max_mp / give_to.ether_cure
                                self.inventory["QTY"][int(res)] = \
                                    self.inventory["QTY"][int(res)] - math.floor(use_count)
                                YourHero.x_ether_use(give_to, use_count)
                                print("You only need to use", math.floor(use_count),
                                      item_use, "and", give_to.name, "got", give_to.ether_cure * use_count, "MP",
                                      give_to.name, "your MP is now", give_to.mp)
                                print("---------------------------------------------------------")
                        elif int(use_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have", use_count, "Ethers")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "antidote":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_antidote_use(give_to)
                            print("You used an antidote and cured", give_to.name, "from poison")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any antidote")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "phoenix down":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "elixir":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_elixir(give_to)
                            print("You used an elixir")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any elixir")
                            print("---------------------------------------------------------")
                    elif item_use.lower() == "silver dust":
                        item_list = pd.DataFrame(self.inventory)
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_silver_dust_use(give_to)
                            print("You used an silver dust and cured", give_to.name, "from paralysed")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any silver dust")
                            print("---------------------------------------------------------")
                    else:
                        print("You don't have that, what do you want to use?")
                        print("---------------------------------------------------------")
                elif player_select == elena.name and elena.hp <= 0:
                    give_to = elena
                    print("Hero is dead you can only use a phoenix down")
                    item_use = input("Do you want to use one?")
                    if item_use.lower() == "yes" or item_use.lower() == "Yes":
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "phoenix down"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        if self.inventory["QTY"][int(res)] > 0:
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - 1
                            YourHero.x_phoenix_down_use(give_to)
                            print("You used an phoenix down and", give_to.name, "Is back to life!!")
                            print("---------------------------------------------------------")
                        elif self.inventory["QTY"][int(res)] <= 0:
                            print("You don't have any phoenix down")
                            print("---------------------------------------------------------")
            elif choice.lower() == "leave":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    break
            else:
                choice = ""
                print("---------------------------------------------------------")

    def x_weapon_armory_shop(self, store_name, clerk_name):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint(store_name)
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            print("Welcome to the", store_name, "shop")
            print("Hi", self.name, "my names is", clerk_name, "how can i help you?")
            print("---------------------------------------------------------")
            input("press enter")
            if self.armory_inventory == {"Weapon/Armor": [], "Magic Slots": [], "weapon power": [],
                                         "Protection": [], "Equipped": [], "Type": []}:
                print("OH, no weapons or armor, I see, that why you are here!")
                print("---------------------------------------------------------")
            else:
                weapon_list = pd.DataFrame(self.armory_inventory)
                weapon_list = weapon_list.rename(index=lambda x: x + 1)
                print("---------------------------------------------------------")
                print(self.name, "Your Weapons/Armor")
                print("---------------------------------------------------------")
                print(weapon_list)
                print("---------------------------------------------------------")
            choice = input("Buy or sell, or leave?\n")
            if choice.lower() == "buy" or choice.lower() == "Buy":

                weapon_buy_list = pd.read_excel(dir_path.weapons_armor_path + weapon_armor_file,
                                                sheet_name=weapon_armor_file_sheet_shop_buy)
                df_weapon_buy = pd.DataFrame(weapon_buy_list)
                df_weapon_buy = df_weapon_buy.rename(index=lambda x: "")
                print("---------------------------------------------------------")
                item_list = pd.DataFrame(self.inventory)
                item_list = item_list.rename(index=lambda x: "")
                print("Your items:")
                print(item_list)
                print("---------------------------------------------------------")
                print(df_weapon_buy)
                print("---------------------------------------------------------")
                buy_weapon = input("What do you want to buy?\n")
                print("---------------------------------------------------------")
                if buy_weapon in df_weapon_buy.values:
                    item_list = pd.DataFrame(self.inventory)
                    item_use = "gil"
                    item_index = item_list[item_list["Item"] == item_use].index.values
                    res = str(item_index)[1:-1]
                    YourHero.x_weapon_gil_cost(self, buy_weapon)
                    if self.weapon_gil_cost > self.inventory["QTY"][int(res)]:
                        print("Sorry you don't have enough Gil")
                        input("press enter")
                        print("---------------------------------------------------------")
                    else:
                        if self.level >= YourHero.x_weapon_shop_level(self, buy_weapon):
                            print("YES YOU CAN BUY")
                            print("You buy", buy_weapon)
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - self.weapon_gil_cost
                            self.armory_inventory["Weapon/Armor"].append(buy_weapon)
                            self.armory_inventory["Magic Slots"].append(YourHero.x_weapon_shop_slots
                                                                        (self, buy_weapon))
                            self.armory_inventory["Protection"].append(YourHero.x_weapon_shop_protection
                                                                       (self, buy_weapon))
                            self.armory_inventory["Type"].append(YourHero.x_weapon_shop_type(self, buy_weapon))
                            self.armory_inventory["weapon power"] \
                                .append(YourHero.x_weapon_shop_power(self, buy_weapon))

                            print("Do you want to equip the", buy_weapon, "?")
                            equipped_choice = input("yes or no?")
                            print("---------------------------------------------------------")
                            if equipped_choice == "yes":
                                # x_weapon_shop_type
                                if YourHero.x_weapon_shop_type(self, buy_weapon) == "weapon":
                                    if self.weapon_equipped == self.weapon_equipped_max:
                                        self.armory_inventory["Equipped"].append("False")
                                        print("Sorry you need to unmount your weapon")
                                        print("You can do this in the inventory menu")
                                        print("The weapon will be placed in inventory, unequipped")
                                        print("---------------------------------------------------------")
                                    else:
                                        self.weapon_equipped += 1
                                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, buy_weapon)
                                        self.armory_inventory["Equipped"].append("True")
                                        print("You have now equipped", buy_weapon, "it can be used right away!")
                                        print("---------------------------------------------------------")
                                elif YourHero.x_weapon_shop_type(self, buy_weapon) == "ring":
                                    if self.rings_equipped == self.rings_equipped_max:
                                        self.armory_inventory["Equipped"].append("False")
                                        print("Sorry you need to unmount your ring")
                                        print("You can do this in the inventory menu")
                                        print("The ring will be placed in inventory, unequipped")
                                        print("---------------------------------------------------------")
                                    else:
                                        self.rings_equipped += 1
                                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, buy_weapon)
                                        self.armory_inventory["Equipped"].append("True")
                                        print("You have now equipped", buy_weapon, "it can be used right away!")
                                        print("---------------------------------------------------------")
                                elif YourHero.x_weapon_shop_type(self, buy_weapon) == "armor":
                                    if self.armor_equipped == self.armor_equipped_max:
                                        self.armory_inventory["Equipped"].append("False")
                                        print("Sorry you need to unmount your armor")
                                        print("You can do this in the inventory menu")
                                        print("The armor will be placed in inventory, unequipped")
                                        print("---------------------------------------------------------")
                                    else:
                                        self.armor_equipped += 1
                                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, buy_weapon)
                                        self.armory_inventory["Equipped"].append("True")
                                        print("You have now equipped", buy_weapon, "it can be used right away!")
                                        print("---------------------------------------------------------")
                            elif equipped_choice == "no":
                                self.armory_inventory["Equipped"].append("False")
                                print("You can equipped the", buy_weapon, "in the inventory menu")
                        else:
                            print("Sorry you ar not at the right level,\nAnd can not buy this!!")
                else:
                    print("Sorry we don't have that")
                    print("---------------------------------------------------------")
            elif choice.lower() == "sell" or choice.lower() == "Sell":
                magic_delete = self.magic_slots_max - self.magic_slots
                spells_count = 0
                weapon_list = pd.DataFrame(self.armory_inventory)
                weapon_list = weapon_list.rename(index=lambda x: x + 1)
                print("---------------------------------------------------------")
                print(self.name, "Your weapons/armor")
                print("---------------------------------------------------------")
                print(weapon_list)
                print("---------------------------------------------------------")
                sell_material = input("What do you want to sell?\n")
                if sell_material.isdigit():
                    print("---------------------------------------------------------")
                    choice = input("Are you sure??")
                    if choice.lower() == "yes" or choice.lower() == "Yes":
                        first_value = weapon_list['Weapon/armor'].values[int(sell_material) - 1]
                        YourHero.x_weapon_gil_sell(self, first_value)
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "gil"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] + self.weapon_gil_sell
                        print("You sell the", first_value, "and receive", self.weapon_gil_sell)
                        print("---------------------------------------------------------")
                        if weapon_list['Equipped'].values[int(sell_material) - 1] == "True":
                            if weapon_list["Type"].values[int(sell_material) - 1] == "weapon":
                                self.weapon_equipped -= 1
                                self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, sell_material)
                                if magic_delete < 0:
                                    print("Go to inventory and add your choice of magic")
                                    print("---------------------------------------------------------")
                                    magic_spell_list = pd.DataFrame(self.magic_spells)
                                    name_count = len(magic_spell_list["Equipped"])
                                    positivenum = abs(magic_delete)
                                    for z in range(name_count):
                                        if magic_spell_list["Equipped"].values[z] == "True":
                                            spells_count += 1
                                            magic_spell_list["Equipped"].pop(z)
                                            magic_spell_list["Equipped"].insert(z, "False")
                                            if spells_count == positivenum:
                                                break
                                else:
                                    print("Your magic will stay where they are")
                                    print("---------------------------------------------------------")
                                self.armory_inventory["Weapon/Armor"].pop(int(sell_material) - 1)
                                self.armory_inventory["Magic Slots"].pop(int(sell_material) - 1)
                                self.armory_inventory["Protection"].pop(int(sell_material) - 1)
                                self.armory_inventory["Equipped"].pop(int(sell_material) - 1)
                                self.armory_inventory["Type"].pop(int(sell_material) - 1)
                                self.armory_inventory["weapon power"].pop(int(sell_material) - 1)
                            elif weapon_list["Type"].values[int(sell_material) - 1] == "ring":
                                self.rings_equipped -= 1
                                self.armory_inventory["Weapon/Armor"].pop(int(sell_material) - 1)
                                self.armory_inventory["Magic Slots"].pop(int(sell_material) - 1)
                                self.armory_inventory["Protection"].pop(int(sell_material) - 1)
                                self.armory_inventory["Equipped"].pop(int(sell_material) - 1)
                                self.armory_inventory["Type"].pop(int(sell_material) - 1)
                                self.armory_inventory["weapon power"].pop(int(sell_material) - 1)
                            elif weapon_list["Type"].values[int(sell_material) - 1] == "armor":
                                self.armor_equipped -= 1
                                self.magic_slots_max -= YourHero.x_weapon_shop_slots(self, sell_material)
                                if magic_delete < 0:
                                    print("Go to inventory and add your choice of magic")
                                    print("---------------------------------------------------------")
                                    magic_spell_list = pd.DataFrame(self.magic_spells)
                                    name_count = len(magic_spell_list["Equipped"])
                                    positivenum = abs(magic_delete)
                                    for z in range(name_count):
                                        if magic_spell_list["Equipped"].values[z] == "True":
                                            spells_count += 1
                                            magic_spell_list["Equipped"].pop(z)
                                            magic_spell_list["Equipped"].insert(z, "False")
                                            if spells_count == positivenum:
                                                break
                                else:
                                    print("Your magic will stay where they are")
                                    print("---------------------------------------------------------")
                                self.armory_inventory["Weapon/Armor"].pop(int(sell_material) - 1)
                                self.armory_inventory["Magic Slots"].pop(int(sell_material) - 1)
                                self.armory_inventory["Protection"].pop(int(sell_material) - 1)
                                self.armory_inventory["Equipped"].pop(int(sell_material) - 1)
                                self.armory_inventory["Type"].pop(int(sell_material) - 1)
                                self.armory_inventory["weapon power"].pop(int(sell_material) - 1)
                        elif weapon_list['Equipped'].values[int(sell_material) - 1] == "False":
                            self.armory_inventory["Weapon/Armor"].pop(int(sell_material) - 1)
                            self.armory_inventory["Magic Slots"].pop(int(sell_material) - 1)
                            self.armory_inventory["Protection"].pop(int(sell_material) - 1)
                            self.armory_inventory["Equipped"].pop(int(sell_material) - 1)
                            self.armory_inventory["Type"].pop(int(sell_material) - 1)
                            self.armory_inventory["weapon power"].pop(int(sell_material) - 1)
                else:
                    print("Sorry pick a number")
                    print("---------------------------------------------------------")
            elif choice.lower() == "leave":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    break
                else:
                    choice = ""
                    print("---------------------------------------------------------")

    def x_finding_weapon_armory_on_journey(self, find_weapon_armory_ring):
        if self.level >= YourHero.x_weapon_shop_level(self, find_weapon_armory_ring):
            print("You found", find_weapon_armory_ring)

            self.armory_inventory["Weapon/Armor"].append(find_weapon_armory_ring)
            self.armory_inventory["Magic Slots"].append(YourHero.x_weapon_shop_slots
                                                        (self, find_weapon_armory_ring))
            self.armory_inventory["Protection"].append(YourHero.x_weapon_shop_protection
                                                       (self, find_weapon_armory_ring))
            self.armory_inventory["Type"].append(YourHero.x_weapon_shop_type(self, find_weapon_armory_ring))
            self.armory_inventory["weapon power"] \
                .append(YourHero.x_weapon_shop_power(self, find_weapon_armory_ring))

            print("Do you want to equip the", find_weapon_armory_ring, "?")
            equipped_choice = input("yes or no?")
            print("---------------------------------------------------------")
            if equipped_choice == "yes":
                # x_weapon_shop_type
                if YourHero.x_weapon_shop_type(self, find_weapon_armory_ring) == "weapon":
                    if self.weapon_equipped == self.weapon_equipped_max:
                        self.armory_inventory["Equipped"].append("False")
                        print("Sorry you need to unmount your weapon")
                        print("You can do this in the inventory menu")
                        print("The weapon will be placed in inventory, unequipped")
                        print("---------------------------------------------------------")
                    else:
                        self.weapon_equipped += 1
                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, find_weapon_armory_ring)
                        self.armory_inventory["Equipped"].append("True")
                        print("You have now equipped", find_weapon_armory_ring, "it can be used right away!")
                        print("---------------------------------------------------------")
                elif YourHero.x_weapon_shop_type(self, find_weapon_armory_ring) == "ring":
                    if self.rings_equipped == self.rings_equipped_max:
                        self.armory_inventory["Equipped"].append("False")
                        print("Sorry you need to unmount your ring")
                        print("You can do this in the inventory menu")
                        print("The ring will be placed in inventory, unequipped")
                        print("---------------------------------------------------------")
                    else:
                        self.rings_equipped += 1
                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, find_weapon_armory_ring)
                        self.armory_inventory["Equipped"].append("True")
                        print("You have now equipped", find_weapon_armory_ring, "it can be used right away!")
                        print("---------------------------------------------------------")
                elif YourHero.x_weapon_shop_type(self, find_weapon_armory_ring) == "armor":
                    if self.armor_equipped == self.armor_equipped_max:
                        self.armory_inventory["Equipped"].append("False")
                        print("Sorry you need to unmount your armor")
                        print("You can do this in the inventory menu")
                        print("The armor will be placed in inventory, unequipped")
                        print("---------------------------------------------------------")
                    else:
                        self.armor_equipped += 1
                        self.magic_slots_max += YourHero.x_weapon_shop_slots(self, find_weapon_armory_ring)
                        self.armory_inventory["Equipped"].append("True")
                        print("You have now equipped", find_weapon_armory_ring, "it can be used right away!")
                        print("---------------------------------------------------------")
            elif equipped_choice == "no":
                self.armory_inventory["Equipped"].append("False")
                print("You can equipped the", find_weapon_armory_ring, "in the inventory menu")
        else:
            print("You found the", find_weapon_armory_ring, "but your level is to weak, you can not carry it")

    def shop_count(self, shop_name, shop_default, enter_count):
        if enter_count == "yes":
            minutes_to_new_items = 25
            path = dir_path.inventory_items_path + inventory_item_file
            wb = openpyxl.load_workbook(path)
            ws = wb["store_count"]
            df_town_count = pd.read_excel(path, "store_count")
            shop_index = df_town_count[df_town_count["town"] == shop_name].index.values
            end_time = df_town_count.loc[shop_index, "end"].values
            start_time = df_town_count.loc[shop_index, "start"].values
            count_times = df_town_count.loc[shop_index, "count"].values
            # print(df_town_count.dtypes)
            end = pd.to_datetime(end_time[0])
            start = pd.to_datetime(start_time[0])
            diff = (end - start)
            minute_diff = diff.seconds / 60
            # print(start)
            row_numbers = df_town_count[df_town_count['town'] == shop_name].index[0]
            row_numbers = row_numbers + 2
            if start_time[0] == 0 and end_time[0] == 0:
                # df_town_count["start"].replace([0], my_timer.get_time_hhmmss)
                b_colm = ws.cell(row=row_numbers, column=2)
                b_colm.value = my_timer.get_time_hhmmss()
                wb.save(path)
            elif start_time[0] != 0 and end_time[0] == 0:
                # df_town_count["end"].replace([0], my_timer.get_time_hhmmss())
                c_colm = ws.cell(row=row_numbers, column=3)
                c_colm.value = my_timer.get_time_hhmmss()
                d_colm = ws.cell(row=row_numbers, column=4)
                d_colm.value = 1
                wb.save(path)
                # df_town_count.to_excel(path)
            elif start_time[0] != 0 and end_time[0] != 0 and count_times[0] > 1:
                if minute_diff >= minutes_to_new_items:
                    ws.cell(row=row_numbers, column=4).value = None
                    ws.cell(row=row_numbers, column=3).value = "0"
                    ws.cell(row=row_numbers, column=2).value = "0"
                    wb.save(path)
                    workbook = xl.load_workbook(path)
                    sheet1 = workbook[shop_default]
                    sheet2 = workbook[shop_name]
                    # mac rows and columns
                    maxr = sheet1.max_row
                    maxc = sheet1.max_column
                    for r in range(1, maxr + 1):
                        for c in range(1, maxc + 1):
                            sheet2.cell(row=r, column=c).value = sheet1.cell(row=r, column=c).value
                    workbook.save(path)
                elif minute_diff < minutes_to_new_items:
                    c_colm = ws.cell(row=row_numbers, column=3)
                    c_colm.value = my_timer.get_time_hhmmss()
                    d_colm = ws.cell(row=row_numbers, column=4)
                    d_colm.value += 1
                    wb.save(path)
            elif start_time[0] != 0 and end_time[0] != 0 and count_times[0] == 1:
                c_colm = ws.cell(row=row_numbers, column=3)
                c_colm.value = my_timer.get_time_hhmmss()
                d_colm = ws.cell(row=row_numbers, column=4)
                d_colm.value += 1
                wb.save(path)

    def buying_habits(self, shop_name, buy_count, buy_items):
        path = dir_path.inventory_items_path + inventory_item_file
        dataset = pd.read_excel(path, "buy_history")
        dataset['count'] = dataset['count'].fillna(0)
        dataset['cost'] = dataset['cost'].fillna(0)
        # print(dataset)
        dataset = dataset.sort_values(['count'], ascending=[False])
        dataset = dataset.reset_index(drop=True)
        dataset.loc[dataset['count'] > 0, 'sold'] = 'True'
        dataset.loc[dataset['count'] == 0, 'sold'] = 'False'

        most_sold_item = dataset.loc[0, "Items"]
        most_sold_item_qty = dataset.loc[0, "count"]
        second_most_sold_item = dataset.loc[1, "Items"]
        second_sold_item_qty = dataset.loc[1, "count"]

        item_index_1 = dataset[dataset["Items"] == most_sold_item].index.values
        item_index_2 = dataset[dataset["Items"] == second_most_sold_item].index.values
        # dataset = pd.read_excel(path, "buy_history")

        wb = openpyxl.load_workbook(path)
        ws = wb["buy_history"]
        df_buy_count = pd.read_excel(path, shop_name)
        if buy_items == most_sold_item or buy_items == second_most_sold_item:
            # print("You buy this a lot, i give you a discount")
            item_index = df_buy_count[df_buy_count["Items"] == buy_items].index.values
            row_numbers = df_buy_count[df_buy_count['Items'] == buy_items].index[0]
            row_numbers = row_numbers + 2
            count_colm_temp = ws.cell(row=row_numbers, column=4)
            count_colm = ws.cell(row=row_numbers, column=2)
            price_colm = ws.cell(row=row_numbers, column=5)
            cost_colm = ws.cell(row=row_numbers, column=3)
            count_colm_temp.value = int(buy_count)
            count_value = count_colm.value + count_colm_temp.value
            count_colm.value = count_value
            total_cost = count_colm.value * price_colm.value
            cost_colm.value = total_cost
            count_colm_temp.value = None
            wb.save(path)
        else:
            item_index = df_buy_count[df_buy_count["Items"] == buy_items].index.values
            row_numbers = df_buy_count[df_buy_count['Items'] == buy_items].index[0]
            row_numbers = row_numbers + 2
            count_colm_temp = ws.cell(row=row_numbers, column=4)
            count_colm = ws.cell(row=row_numbers, column=2)
            price_colm = ws.cell(row=row_numbers, column=5)
            cost_colm = ws.cell(row=row_numbers, column=3)
            count_colm_temp.value = int(buy_count)
            count_value = count_colm.value + count_colm_temp.value
            count_colm.value = count_value
            total_cost = count_colm.value * price_colm.value
            cost_colm.value = total_cost
            count_colm_temp.value = None
            wb.save(path)

    def x_materia_shop(self, shop_name, clerk_name):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint(shop_name)
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            print("Welcome to the", shop_name, "shop")
            print("Hi", self.name, "my name is", clerk_name, "how can i help you?")
            print("---------------------------------------------------------")
            input("press enter")
            if self.magic_spells == {"Magic": [], "mp cost": [], "AP": [], "Level": [],
                                     "Equipped": [], "Magic Type": []}:
                print("no magic, but that's why you are here!")
                print("---------------------------------------------------------")
            else:
                magic_list = pd.DataFrame(self.magic_spells)
                magic_list = magic_list.rename(index=lambda x: x + 1)
                print("---------------------------------------------------------")
                print(self.name, "Your magic")
                print("---------------------------------------------------------")
                print(magic_list)
                print("---------------------------------------------------------")
            choice = input("Buy or sell, or leave?\n")
            if choice.lower() == "buy" or choice.lower() == "Buy":

                magic_buy_list = pd.read_excel(dir_path.magic_path + magic_file, sheet_name=magic_file_sheet_magic_gill)
                df_magic_buy = pd.DataFrame(magic_buy_list)
                df_magic_buy = df_magic_buy.rename(index=lambda x: "")
                item_list = pd.DataFrame(self.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                print("---------------------------------------------------------")
                print(self.name, "you have:", self.inventory["QTY"][int(res)], "Gil")
                print("---------------------------------------------------------")
                print(df_magic_buy)
                print("---------------------------------------------------------")
                buy_material = input("What do you want to buy?\n")
                print("---------------------------------------------------------")
                if buy_material in df_magic_buy.values:
                    YourHero.x_magic_gil_cost(self, buy_material)
                    if self.inventory["QTY"][int(res)] < self.magic_gil_cost:
                        print("Sorry you don't have enough Gil")
                        input("press enter")
                        print("---------------------------------------------------------")
                    else:
                        print("You buy materia", buy_material)
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - self.magic_gil_cost
                        self.magic_spells["Magic"].append(buy_material)
                        self.magic_spells["mp cost"].append(YourHero.x_magic_cost(self, buy_material))
                        self.magic_spells["AP"].append(0)
                        self.magic_spells["Level"].append(1)
                        self.magic_spells["Magic Type"].append(YourHero.x_magic_type(self, buy_material))
                        print("Do you want to equip the magic?")
                        equipped_choice = input("yes or no?")
                        print("---------------------------------------------------------")
                        if equipped_choice == "yes":
                            if self.magic_slots == self.magic_slots_max:
                                self.magic_spells["Equipped"].append("False")
                                print("Sorry all your magic slots are in use, handle this in the inventory menu")
                                print("Or you need to buy a weapon that has magic slots")
                                print("Your new magic will be placed in you inventory, not in use")
                                print("---------------------------------------------------------")
                            elif self.magic_slots < self.magic_slots_max:
                                print("You have now equipped", buy_material, "it can be used right away!")
                                print("---------------------------------------------------------")
                                self.magic_spells["Equipped"].append("True")
                                self.magic_slots += 1
                        else:
                            self.magic_spells["Equipped"].append("False")
                            print("You can go to the inventory, and do this later")
                            print("---------------------------------------------------------")
                else:
                    print("Sorry we don't have that")
                    print("---------------------------------------------------------")
            elif choice.lower() == "sell" or choice.lower() == "Sell":
                magic_list = pd.DataFrame(self.magic_spells)
                magic_list = magic_list.rename(index=lambda x: x + 1)
                print("---------------------------------------------------------")
                print(self.name, "Your magic")
                print("---------------------------------------------------------")
                print(magic_list)
                print("---------------------------------------------------------")
                sell_material = input("What do you want to sell?\n")
                if sell_material.isdigit():
                    print("---------------------------------------------------------")
                    choice = input("Are you sure??")
                    if choice.lower() == "yes" or choice.lower() == "Yes":
                        item_list = pd.DataFrame(self.inventory)
                        item_use = "gil"
                        item_index = item_list[item_list["Item"] == item_use].index.values
                        res = str(item_index)[1:-1]
                        first_value = magic_list['Magic'].values[int(sell_material) - 1]
                        YourHero.x_magic_gil_sell(self, first_value)
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] + self.magic_gil_sell
                        print("You sell the", first_value, "and receive", self.magic_gil_sell)
                        print("---------------------------------------------------------")
                        # magic_list.drop(index=int(sell_material), inplace=True)
                        if magic_list['Equipped'].values[int(sell_material) - 1] == "True":
                            self.magic_slots -= 1
                            self.magic_spells["Magic"].pop(int(sell_material) - 1)
                            self.magic_spells["mp cost"].pop(int(sell_material) - 1)
                            self.magic_spells["AP"].pop(int(sell_material) - 1)
                            self.magic_spells["Level"].pop(int(sell_material) - 1)
                            self.magic_spells["Equipped"].pop(int(sell_material) - 1)
                            self.magic_spells["Magic Type"].pop(int(sell_material) - 1)
                        elif magic_list['Equipped'].values[int(sell_material) - 1] == "False":
                            self.magic_spells["Magic"].pop(int(sell_material) - 1)
                            self.magic_spells["mp cost"].pop(int(sell_material) - 1)
                            self.magic_spells["AP"].pop(int(sell_material) - 1)
                            self.magic_spells["Level"].pop(int(sell_material) - 1)
                            self.magic_spells["Equipped"].pop(int(sell_material) - 1)
                            self.magic_spells["Magic Type"].pop(int(sell_material) - 1)
                else:
                    print("Sorry pick a number")
                    print("---------------------------------------------------------")
            elif choice.lower() == "leave":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    break
                else:
                    choice = ""
                    print("---------------------------------------------------------")

    def x_shop(self, shop_name, clerk_name, shop_default, enter_count):
        YourHero.shop_count(self, shop_name, shop_default, enter_count)
        var = len(self.inventory["Item"])
        print("---------------------------------------------------------")
        tprint(shop_name)
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            print("Welcome to the", shop_name, "shop")
            print("Hi", self.name, "my name is", clerk_name, "how can i help you?")
            print("---------------------------------------------------------")
            input("press enter")
            print("---------------------------------------------------------")
            item_list = pd.DataFrame(self.inventory)
            item_list = item_list.rename(index=lambda x: "")
            print("Your items:")
            print(item_list)
            print("---------------------------------------------------------")
            choice = input("Buy or sell, or leave?\n")
            if choice.lower() == "buy" or choice.lower() == "Buy":

                item_buy_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
                df_item_buy = pd.DataFrame(item_buy_list)
                df_item_buy = df_item_buy.rename(index=lambda x: "")
                print("---------------------------------------------------------")
                # print(self.name, "you have:", item_list["Item"], "Gil")
                print("---------------------------------------------------------")
                print(df_item_buy)
                print("---------------------------------------------------------")
                buy_items = input("What do you want to buy, lower cases!?\n")
                print("---------------------------------------------------------")
                item_list = pd.DataFrame(self.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                if buy_items in df_item_buy.values:
                    YourHero.x_item_quantity(self, buy_items, shop_name)
                    if self.store_quantity == 0:
                        print("Sorry we are out of ", buy_items)
                    elif self.store_quantity > 0:
                        buy_count = input("How many do you want?\n")
                        print("---------------------------------------------------------")
                        if int(buy_count) > self.store_quantity:
                            print("We only have", self.store_quantity)
                            print("---------------------------------------------------------")
                        elif int(buy_count) <= self.store_quantity:
                            YourHero.x_item_cost(self, buy_items, shop_name)
                            if self.store_cost * int(buy_count) > self.inventory["QTY"][int(res)]:
                                print("Sorry you don't have enough money")
                                print("---------------------------------------------------------")
                            elif self.store_cost <= self.inventory["QTY"][int(res)]:
                                if buy_items in item_list.values:
                                    item_index = item_list[item_list["Item"] == buy_items].index.values
                                    res = str(item_index)[1:-1]
                                    self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] + int(buy_count)
                                    item_use = "gil"
                                    item_index = item_list[item_list["Item"] == item_use].index.values
                                    res = str(item_index)[1:-1]
                                    self.inventory["QTY"][int(res)] = \
                                        self.inventory["QTY"][int(res)] - int(buy_count) * self.store_cost
                                    # potion_qty = potion_qty - int(buy_count)
                                    if int(buy_count) >= 2:
                                        print("You bought", buy_count, buy_items)
                                        YourHero.x_item_selling(self, buy_items, shop_name, buy_count)
                                        print("---------------------------------------------------------")
                                    else:
                                        print("You bought", buy_count, buy_items)
                                        YourHero.x_item_selling(self, buy_items, shop_name, buy_count)
                                        print("---------------------------------------------------------")
                                else:
                                    self.inventory["QTY"].insert(var + 1, int(buy_count))
                                    self.inventory["Item"].insert(var + 1, buy_items)
                                    item_use = "gil"
                                    item_index = item_list[item_list["Item"] == item_use].index.values
                                    res = str(item_index)[1:-1]
                                    self.inventory["QTY"][int(res)] = \
                                        self.inventory["QTY"][int(res)] - int(buy_count) * self.store_cost
                                    # potion_qty = potion_qty - int(buy_count)
                                    if int(buy_count) >= 2:
                                        print("You bought", buy_count, buy_items)
                                        YourHero.x_item_selling(self, buy_items, shop_name, buy_count)
                                        print("---------------------------------------------------------")
                                    else:
                                        print("You bought", buy_count, buy_items)
                                        YourHero.x_item_selling(self, buy_items, shop_name, buy_count)
                                        print("---------------------------------------------------------")
                else:
                    print("Sorry we dont reconize the product")
            elif choice.lower() == "sell" or choice.lower() == "Sell":
                print("---------------------------------------------------------")
                print(item_list)
                print("---------------------------------------------------------")
                sell_items = input("What do you want to sell, lower cases!?\n")
                print("---------------------------------------------------------")
                item_list = pd.DataFrame(self.inventory)
                item_index = item_list[item_list["Item"] == sell_items].index.values
                res = str(item_index)[1:-1]
                if sell_items in item_list.values:
                    sell_count = input("How many do you want to sell?\n")
                    print("---------------------------------------------------------")
                    try:
                        if int(sell_count) > self.inventory["QTY"][int(res)]:
                            print("You don't have that many to sell")
                        elif int(sell_count) <= self.inventory["QTY"][int(res)]:
                            item_list = pd.read_excel(dir_path.inventory_items_path + inventory_item_file,
                                                      sheet_name=shop_name)
                            dt = pd.DataFrame(item_list)
                            dt.set_index("Items", inplace=True)
                            dt.loc[sell_items]["Quantity"] += int(sell_count)
                            dt.to_excel(dir_path.inventory_items_path + inventory_item_file, sheet_name=shop_name)
                            YourHero.x_money_back_store(self, sell_items, shop_name)
                            self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - int(sell_count)
                            item_list = pd.DataFrame(self.inventory)
                            item_use = "gil"
                            item_index = item_list[item_list["Item"] == item_use].index.values
                            res = str(item_index)[1:-1]
                            self.inventory["QTY"][int(res)] = \
                                self.inventory["QTY"][int(res)] + int(sell_count) * self.store_money_back
                            print(f"You sold {sell_count}, {sell_items}, and received {self.store_money_back} gil")
                            print("---------------------------------------------------------")
                        else:
                            print("Write a number!!")
                    except:
                        print("---------------------------------------------------------")
                        print("Write a number!!")
                else:
                    print("You don't have that in you inventory!!")
            elif choice.lower() == "leave":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    break
                else:
                    choice = ""
                    print("---------------------------------------------------------")

    def x_sleep_inn(self, inn_name, name, room_1, room_2, room_1_cost,
                    room_2_cost, food_1, food_1_cost, food_2, food_2_cost):
        YourHero.x_initials_stats(self)
        print("---------------------------------------------------------")
        tprint(inn_name)
        print("---------------------------------------------------------")
        choice = ""
        while not choice.lower() == "leave":
            print("Welcome to the", inn_name)
            print("Hi", self.name, "my names is", name, "how can i help you?")
            print("---------------------------------------------------------")
            input("press enter")
            print("---------------------------------------------------------")
            print(YourHero.x_items_list(self))
            print("---------------------------------------------------------")
            choice = input("Sleep, eat, or leave?\n")
            if choice.lower() == "sleep" or choice.lower() == "Sleep":
                dots = ":     "
                name_dot = name + " " + "Inn" + dots

                skills = {name_dot: [""],
                          "Rooms": [room_1, room_2],
                          "Cost": [room_1_cost, room_2_cost],
                          "": ["Gil", "gil"]

                          }
                inn_rooms = pd.DataFrame(skills, index=["", ""])
                item_list = pd.DataFrame(self.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                print("---------------------------------------------------------")
                print(self.name, "you have:", self.inventory["QTY"][int(res)], "Gil")
                print("---------------------------------------------------------")
                print(inn_rooms)
                print("---------------------------------------------------------")
                room_to_sleep = input("What room do you wish to sleep in?\n")
                print("---------------------------------------------------------")

                if room_to_sleep == room_1:
                    if room_1_cost <= self.inventory["QTY"][int(res)]:
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - room_1_cost
                        print("You pay the", room_1_cost, "Gil")
                        print("You start walking to ", room_1)
                        self.inn_room_1 = 1
                        print("---------------------------------------------------------")
                        break
                    else:
                        print("You don't have enough money")
                        print("---------------------------------------------------------")
                elif room_to_sleep == room_2:
                    if room_2_cost <= self.inventory["QTY"][int(res)]:
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - room_2_cost
                        print("You pay the", room_2_cost, "Gil")
                        print("You start walking to ", room_2)
                        self.inn_room_2 = 1
                        print("---------------------------------------------------------")
                        break
                    else:
                        print("You don't have enough money")
                        print("---------------------------------------------------------")
            elif choice.lower() == "eat" or choice.lower() == "Eat":
                dots = ":     "
                name_dot = name + " " + "restaurant" + dots

                skills = {name_dot: [""],
                          "Food": [food_1, food_2],
                          "Cost": [food_1_cost, food_2_cost],
                          "": ["Gil", "gil"]

                          }
                inn_restaurant = pd.DataFrame(skills, index=["", ""])
                print("---------------------------------------------------------")
                print(inn_restaurant)
                print("---------------------------------------------------------")
                eat_choice = input("What do you want to eat?\n")
                print("---------------------------------------------------------")
                item_list = pd.DataFrame(self.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                if eat_choice == food_1:
                    if food_1_cost <= self.inventory["QTY"][int(res)]:
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - food_1_cost
                        print("You pay the", food_1_cost, "Gil")
                        print("You start walking to the restaurant.")
                        self.food_1 = 1
                        print("---------------------------------------------------------")
                        break
                    else:
                        print("You don't have enough money")
                        print("---------------------------------------------------------")
                elif eat_choice == food_2:
                    if food_2_cost <= self.inventory["QTY"][int(res)]:
                        self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] - food_2_cost
                        print("You pay the", food_2_cost, "Gil")
                        print("You start walking to the restaurant.")
                        self.food_2 = 1
                        print("---------------------------------------------------------")
                        break
                    else:
                        print("You don't have enough money")
                        print("---------------------------------------------------------")
            elif choice.lower() == "leave":
                leave = input("Are you sure?\n")
                print("---------------------------------------------------------")
                if leave.lower() == "yes" or leave.lower() == "y":
                    break
                else:
                    choice = ""
                    print("---------------------------------------------------------")

    def x_level_up_f(self):
        while self.exp >= self.next_level_exp:
            self.next_level_exp = YourHero.x_level_up(self)
            if self.exp >= self.next_level_exp:
                self.exp = self.exp - self.next_level_exp
                self.factor_level_up += 1
                self.next_level_exp = YourHero.x_level_up(self)
                self.level += 1
                self.level_count += 1
                self.bonus_factor += 3
                YourHero.x_initials_stats(self)
                tprint("New Level!!")
                tprint(f"{self.name} level {self.level}")
                print("---------------------------------------------------------")
                input("press enter")
                # print(self.name, "exp ", self.exp, "next level: ", self.next_level_exp)
        else:
            print("")
            # print(self.name, "exp: ", self.exp, "next level exp: ", self.next_level_exp)
            # elif self.exp < self.next_level_exp:
            # print(self.name, "exp: ", self.exp, "next level exp: ", self.next_level_exp)

    def x_initials_stats(self):
        self.strength = YourHero.x_strength(self)
        self.max_hp = YourHero.x_hp(self)
        self.atk = YourHero.x_atk(self)
        self.defence = YourHero.x_defence(self)
        self.magic = YourHero.x_magic(self)
        self.max_mp = YourHero.x_mp(self)
        self.spirit = YourHero.x_spirit(self)
        self.magic_atk = YourHero.x_magic_atk(self)
        self.potion_cure = YourHero.x_potion_hp(self)
        self.ether_cure = YourHero.x_ether_mp(self)
        self.elixir_cure = YourHero.x_elixir(self)
        self.stat = YourHero.x_stat(self)
        self.total_stat = YourHero.x_stat_all_hero(self)
        self.item_list = YourHero.x_items_list(self)
        self.win_stat = YourHero.x_battle_win_1_stat(self)
        self.win_stat_2 = YourHero.x_battle_win_2_stat(self)
        self.win_stat_3 = YourHero.x_battle_win_3_stat(self)
        self.status_effect_name = YourHero.x_status_effect_name(self)
        self.status_effect_2_name = YourHero.x_status_effect_name(self)



    def x_story_chapter_count(self):
        if elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[0]:
            intro()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[1]:
            story1()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[2]:
            story1_first_crossing()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[3]:
            story1_inn_at_abreheim_town()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[4]:
            story1_ally_in_abreheim(1)
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[5]:
            story1_ally_in_abreheim(2)
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[6]:
            story1_round_room()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[7]:
            story1_basement()
        elif elena.chapter == story.chapter_names[0] and elena.story_name == story.chap_1_names[8]:
            story1_snow_in_south()
        else:
            print("Something got wrong", elena.chapter)
            exit()

    def x_win_battle(self, hero1, hero2, boss):
        if boss == "yes":
            level = self.enemy_boss_level
            enemy_level = pd.read_excel(dir_path.level_path + level_file, sheet_name=level_file_sheet_strenght_mod_enemy)
            dt = pd.DataFrame(enemy_level)
            dt.set_index("Level", inplace=True)
            new_enemy_strength = dt.loc[level]["Bonus"]
            bonus = new_enemy_strength
            # enemy.e_strength = bonus
            enemy.e_factor_for_level_up += level
            enemy.e_level_count += level
            enemy.e_bonus_factor = bonus
            enemy.e_level += level
            enemy.ex_initials_stats()

        if opponents.count == 1:
            op_count = opponents.count
            if hero1 == cloud and hero2 == cloud:
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                enemy.e_hp = enemy.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                print(cloud.win_stat)
                print("")
                input("Press enter to continue")
                print(cloud.item_list)
                print("")
                input("Press enter to continue")
                print("")
                print(cloud.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            elif hero1 == elena and hero2 == elena:
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                enemy.e_hp = enemy.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                print(elena.win_stat)
                print("")
                input("Press enter to continue")
                print(elena.item_list)
                print("")
                input("Press enter to continue")
                print(elena.total_stat)
                print("")
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            else:
                Enemy.ex_initials_stats(enemy)
                enemy.e_hp = enemy.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                print(cloud.win_stat)
                print("")
                print(elena.win_stat)
                input("Press enter to continue")
                print(cloud.item_list)
                print("")
                print(elena.item_list)
                input("Press enter to continue")
                print("")
                print(cloud.total_stat)
                print("")
                print(elena.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
        elif opponents.count == 2:
            op_count = opponents.count
            if hero1 == cloud and hero2 == cloud:
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                cloud.exp = cloud.exp + enemy_2.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_2.e_exp_give
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_2.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_gen_loot(cloud, enemy_2.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                print(cloud.win_stat_2)
                input("Press enter to continue")
                print("")
                print(cloud.item_list)
                print("")
                input("Press enter to continue")
                print("")
                print(cloud.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            elif hero1 == elena and hero2 == elena:
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                elena.exp = elena.exp + enemy_2.e_exp_give
                elena.exp_show = elena.exp_show + enemy_2.e_exp_give
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_2.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_gen_loot(elena, enemy_2.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                print(elena.win_stat_2)
                print("")
                input("Press enter to continue")
                print(elena.item_list)
                print("")
                input("Press enter to continue")
                print(elena.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            else:
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                cloud.exp = cloud.exp + enemy_2.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_2.e_exp_give
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_2.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_gen_loot(cloud, enemy_2.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                elena.exp = elena.exp + enemy_2.e_exp_give
                elena.exp_show = elena.exp_show + enemy_2.e_exp_give
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_2.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_gen_loot(elena, enemy_2.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                print(elena.win_stat_2)
                print("")
                print(cloud.win_stat_2)
                print("")
                input("Press enter to continue")
                print(elena.item_list)
                print("")
                print(cloud.item_list)
                print("")
                input("Press enter to continue")
                print(elena.total_stat)
                print("")
                print(cloud.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
        elif opponents.count == 3:
            op_count = opponents.count
            if hero1 == cloud and hero2 == cloud:
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy_3.e_hp = enemy_3.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                enemy_3.e_exp = enemy_3.e_exp + enemy_3.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_3)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                cloud.exp = cloud.exp + enemy_2.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_2.e_exp_give
                cloud.exp = cloud.exp + enemy_3.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_3.e_exp_give
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_2.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_3.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_gen_loot(cloud, enemy_2.e_name)
                YourHero.x_gen_loot(cloud, enemy_3.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                Enemy.ex_level_up_f(enemy_3)
                print(cloud.win_stat_3)
                print("")
                input("Press enter to continue")
                print("")
                print(cloud.item_list)
                print("")
                input("Press enter to continue")
                print("")
                print(cloud.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            elif hero1 == elena and hero2 == elena:
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy_3.e_hp = enemy_3.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                enemy_3.e_exp = enemy_3.e_exp + enemy_3.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_3)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                elena.exp = elena.exp + enemy_2.e_exp_give
                elena.exp_show = elena.exp_show + enemy_2.e_exp_give
                elena.exp = elena.exp + enemy_3.e_exp_give
                elena.exp_show = elena.exp_show + enemy_3.e_exp_give
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_2.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_3.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_gen_loot(elena, enemy_2.e_name)
                YourHero.x_gen_loot(elena, enemy_3.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                Enemy.ex_level_up_f(enemy_3)
                print(elena.win_stat_3)
                print("")
                input("Press enter to continue")
                print(elena.item_list)
                print("")
                input("Press enter to continue")
                print(elena.total_stat)
                print("")
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)
            else:
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                enemy.e_hp = enemy.e_max_hp
                enemy_2.e_hp = enemy_2.e_max_hp
                enemy_3.e_hp = enemy_3.e_max_hp
                enemy.e_exp = enemy.e_exp + enemy.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy)
                enemy_2.e_exp = enemy_2.e_exp + enemy_2.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_2)
                enemy_3.e_exp = enemy_3.e_exp + enemy_3.e_exp_give
                enemy.e_exp_give = Enemy.ex_level_up_give(enemy_3)
                YourHero.x_initials_stats(cloud)
                cloud.exp = cloud.exp + enemy.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy.e_exp_give
                cloud.exp = cloud.exp + enemy_2.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_2.e_exp_give
                cloud.exp = cloud.exp + enemy_3.e_exp_give
                cloud.exp_show = cloud.exp_show + enemy_3.e_exp_give
                item_list = pd.DataFrame(cloud.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_2.e_gil
                cloud.inventory["QTY"][int(res)] = cloud.inventory["QTY"][int(res)] + enemy_3.e_gil
                YourHero.x_gen_loot(cloud, enemy.e_name)
                YourHero.x_gen_loot(cloud, enemy_2.e_name)
                YourHero.x_gen_loot(cloud, enemy_3.e_name)
                YourHero.x_initials_stats(cloud)
                YourHero.x_level_up_f(cloud)
                YourHero.x_magic_ap_level_up(cloud, op_count)
                YourHero.x_initials_stats(elena)
                elena.exp = elena.exp + enemy.e_exp_give
                elena.exp_show = elena.exp_show + enemy.e_exp_give
                elena.exp = elena.exp + enemy_2.e_exp_give
                elena.exp_show = elena.exp_show + enemy_2.e_exp_give
                elena.exp = elena.exp + enemy_3.e_exp_give
                elena.exp_show = elena.exp_show + enemy_3.e_exp_give
                item_list = pd.DataFrame(elena.inventory)
                item_use = "gil"
                item_index = item_list[item_list["Item"] == item_use].index.values
                res = str(item_index)[1:-1]
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_2.e_gil
                elena.inventory["QTY"][int(res)] = elena.inventory["QTY"][int(res)] + enemy_3.e_gil
                YourHero.x_gen_loot(elena, enemy.e_name)
                YourHero.x_gen_loot(elena, enemy_2.e_name)
                YourHero.x_gen_loot(elena, enemy_3.e_name)
                YourHero.x_initials_stats(elena)
                YourHero.x_level_up_f(elena)
                YourHero.x_magic_ap_level_up(elena, op_count)
                Enemy.ex_initials_stats(enemy)
                Enemy.ex_level_up_f(enemy)
                Enemy.ex_initials_stats(enemy_2)
                Enemy.ex_level_up_f(enemy_2)
                Enemy.ex_initials_stats(enemy_3)
                Enemy.ex_level_up_f(enemy_3)
                print(elena.win_stat_3)
                print("")
                print(cloud.win_stat_3)
                print("")
                input("Press enter to continue")
                print(elena.item_list)
                print("")
                print(cloud.item_list)
                print("")
                input("Press enter to continue")
                print(elena.total_stat)
                print("")
                print(cloud.total_stat)
                input("Press enter to continue")
                YourHero.x_story_chapter_count(elena)

    def x_magic_ap_level_up(self, op_count):
        df = pd.DataFrame(self.magic_spells)
        in_use = df[df["Equipped"] == "True"].index.values
        level_up_2 = df["AP"].index.values
        level_up_3 = df["AP"].index.values
        ap_index = df.loc[in_use, "AP"].index.values
        for x in in_use:
            for z in ap_index:
                ap = df.loc[[z], "AP"].values
                new_ap = str(ap)[1:-1]
                ap_add = enemy.e_exp_give / 2 * op_count
                ap_p = int(float(new_ap)) + ap_add  # The enemy exp/2
                self.magic_spells["AP"].pop(x)
                self.magic_spells["AP"].insert(z, ap_p)
        if [500] <= self.magic_spells["AP"] <= [1499]:
            for t in level_up_2:
                for y in df["Magic"]:  # df["Magic"]:
                    level2 = y + "2"
                    self.magic_spells["Magic"].pop(t)
                    self.magic_spells["Magic"].append(level2)
        elif self.magic_spells["AP"] >= [1500]:
            for t in level_up_3:
                for y in df["Magic"].str[:-1]:
                    level3 = y + "3"
                    self.magic_spells["Magic"].pop(t)
                    self.magic_spells["Magic"].append(level3)

    def x_gen_loot(self, name):
        tprint("LOOT")
        print("---------------------------------------------------------")
        var = len(self.inventory["Item"])
        enemy_list = pd.read_excel(dir_path.enemy_path + enemy_name_file)
        df = pd.DataFrame(enemy_list)
        df.set_index("name", inplace=True)
        loot_give = df.loc[name]["loot"]
        loot_qty = df.loc[name]["qty_loot"]
        item_list = pd.DataFrame(self.inventory)
        Enemy.ex_initials_stats(enemy)
        Enemy.ex_initials_stats(enemy_2)
        Enemy.ex_initials_stats(enemy_3)
        loot_chance = random.randint(0, 4)
        if loot_chance >= 2:
            if loot_give in item_list.values:
                item_index = item_list[item_list["Item"] == loot_give].index.values
                res = str(item_index)[1:-1]
                self.inventory["QTY"][int(res)] = self.inventory["QTY"][int(res)] + loot_qty
                YourHero.x_initials_stats(self)
                print(self.name, "you looted the", name, "and found", loot_qty, loot_give)
                print("---------------------------------------------------------")
                input("press enter to continue")
            else:
                self.inventory["QTY"].insert(var + 1, loot_qty)
                self.inventory["Item"].insert(var + 1, loot_give)
                YourHero.x_initials_stats(self)
                print(self.name, "you looted the", name, "and found", loot_qty, loot_give)
                print("---------------------------------------------------------")
                input("press enter to continue")
        else:
            print(self.name, "you looted the", name, "and found... nothing")
            print("---------------------------------------------------------")
            input("press enter to continue")

    def x_status_effect(self, opponent):
        # enemy.ex_initials_stats()
        Enemy.ex_initials_stats(opponent)
        ring_list = pd.DataFrame(self.armory_inventory)
        if "ring" not in ring_list:
            if self.status_poison == 1:
                self.status_poison += 1
                # self.hp -= enemy.e_poison_dmg
                self.hp -= opponent.e_poison_dmg
                print(self.name, "have poison and loose", opponent.e_poison_dmg, "HP")
                time.sleep(time_short_wait)
            elif self.status_poison == 2:
                self.status_poison += 1
                self.hp -= opponent.e_poison_dmg
                print(self.name, "have poison and loose", opponent.e_poison_dmg, "HP")
                time.sleep(time_short_wait)
            elif self.status_poison >= 3:
                self.status_poison = self.status_poison - self.status_poison
                print(self.name, "have now cured from the poison")
                time.sleep(time_short_wait)
            else:
                if self.status_poison == 0:
                    if opponent.e_name == "Snake":
                        chance = random.randint(0, 10)
                        if chance >= 6:
                            self.status_poison += 1
                            self.hp -= opponent.e_poison_dmg
                            print(self.name, "got poison and loose", opponent.e_poison_dmg, "HP")
                            time.sleep(time_short_wait)
        else:
            if "poison" in YourHero.x_see_inventory_name_qty(self, "ring", 0):
                ring_name = YourHero.x_see_inventory_name_qty(self, "ring", 1)[0]
                print(f"The enemy tries to poison you, but you have the {ring_name}"
                      f" and are protected from poison")
                time.sleep(time_short_wait)
        if "ring" not in ring_list:
            if self.status_paralyzed == 0:
                if opponent.e_name == "Giant":
                    chance = random.randint(0, 10)
                    if chance >= 8:
                        self.status_paralyzed += 1
                        print(self.name, "are paralyzed!!")
                        time.sleep(time_short_wait)
        else:
            if "paralyzed" in YourHero.x_see_inventory_name_qty(self, "ring", 0):
                ring_name = YourHero.x_see_inventory_name_qty(self, "ring", 1)[1]
                print(f"The enemy tries to paralyse you, but you have the {ring_name} "
                      f" and are protected from paralyse")
                time.sleep(time_short_wait)

    def x_status_check(self, opponent):
        Enemy.ex_initials_stats(opponent)
        if self.status_poison == 1 or self.status_poison == 2:
            self.status_poison += 1
            self.hp -= opponent.e_poison_dmg
            print(self.name, "Cure your status effect with an item! "
                             "Poison in your blood, you loose", opponent.e_poison_dmg, "HP")
            time.sleep(time_short_wait)
        elif self.status_poison == 3:
            self.status_poison = self.status_poison - self.status_poison
            print(self.name, "finally the poison in your blood is now gone!!")
            time.sleep(time_short_wait)
        elif self.status_poison == 0:
            print(self.name, "you are not poisoned and can continue to fight with out continuously loosing HP")
            time.sleep(time_short_wait)

    def x_game_over(self):
        if cloud.hp <= 0:
            cloud.dead_times += 1
            YourHero.x_initials_stats(cloud)
            if cloud.dead_times == 1:
                print("A sad moment has come, you fight for your last breath as your heart stops bonding")
                print(cloud.name, "is DEAD!!!")
                print("Only the well known bird can send the hero back to life, hurry before you", elena.name, "dies")
                print("Then it is forevermore GAME-OVER, and the world will be lost in the hands of pollution")
                input("Press Enter")
            elif cloud.dead_times > 1:
                print("Phoenix Down will bring", cloud.name, "back to life")
                input("Press Enter")
        elif elena.hp <= 0:
            elena.dead_times += 1
            YourHero.x_initials_stats(elena)
            if elena.dead_times == 1:
                print("A sad moment has come, you fight for your last breath as your heart stops bonding")
                print(elena.name, "is DEAD!!!")
                print("Only the well known bird can send the hero back to life, hurry before you", cloud.name, "dies")
                print("Then it is forevermore GAME-OVER, and the world will be lost in the hands of pollution")
                input("Press Enter")
            elif elena.dead_times > 1:
                print("Phoenix Down will bring", elena.name, "back to life")
                input("Press Enter")
        if elena.hp <= 0 and cloud.hp <= 0:
            print("Its know a fact, the world will stand no chance, "
                  "but rot in the destinies left by the unthoughtful humans")
            play_again = input("Do you want to play again?: Yes/No?")
            if play_again.lower() == "yes" or play_again.lower() == "y":
                main_game()
            else:
                print("Thanks for playing")
                print(elena.stat)
                print(cloud.stat)
                exit()

    def x_run_away(self):
        YourHero.x_initials_stats(self)
        chance = random.randint(0, 4)
        if chance > 1:
            self.run = 1
            print("You ran away from the fight")
            input("Press enter")

        elif chance == 0:
            print("Desperate you try to run, but fail")

    def x_front_battle(self):
        print(self.name, "jump back to the front again, normal enemy atk, and hero atk")
        if self.defence_count > 0:
            self.defence_count = 0
        elif self.defence_count <= 0:
            input("You are already in the front, press enter")

    def x_defence_battle(self):
        print(self.name, " jumps back and raise your shield for defence, reduced enemy atk, and reduced hero atk")
        if self.defence_count <= 0:
            self.defence_count += 1
        elif self.defence_count > 0:
            input("You are already in the back, press enter")

    def x_steal_battle(self):
        stealing_chance = random.randint(0, 10)
        if stealing_chance >= 5:
            stealing = ["gil", "potion", "ether", "silver_dust", "phoenix_down", "tent"]
            stealing_qty = [random.randint(150, 400), random.randint(2, 5), random.randint(1, 2),
                            random.randint(1, 2), random.randint(1, 2), random.randint(1, 2)]
            item_chance_mugging = random.randint(0, 17)
            if item_chance_mugging <= 10:
                mugg_item = random.randint(0, 1)
                if mugg_item == 0:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 1:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
            elif item_chance_mugging > 11 <= 15:
                mugg_item = random.randint(2, 4)
                if mugg_item == 2:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 3:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 4:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
            elif item_chance_mugging >= 16:
                mugg_item = random.randint(0, 5)
                if mugg_item == 1:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 2:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 3:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 4:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
                elif mugg_item == 5:
                    YourHero.x_add_items_inventory(self,
                                                   {"Item": [stealing[mugg_item]], "QTY": [stealing_qty[mugg_item]]})
                    print("You stole", stealing_qty[mugg_item], stealing[mugg_item])
        else:
            print("You could not steal anything")

    def x_attack_battle(self, opponent):
        # {"Weapon/Armor": ["wind storm"], "Magic Slots": [1], "weapon power": [1],
        #                               "Protection": ["none"], "Equipped": ["True"], "Type": ["weapon"]}
        hit = random.randint(0, 10)
        if hit >= 2:
            weapon_list = pd.DataFrame(self.armory_inventory)
            weapon_list = weapon_list.rename(index=lambda x: x + 1)
            weapon_list.loc[(weapon_list['Equipped'] == 'True') &
                            (weapon_list['Type'] == 'weapon'), 'Weapon_Equipped'] = 'Weapon_Selected'
            # if weapon_list["Equipped"] == "True" and weapon_list["Type"] == "weapon":
            if "Weapon_Selected" in weapon_list.values:
                is_true = (weapon_list["Equipped"] == "True") & (weapon_list["Type"] == "weapon")
                weapon_list_true = weapon_list[is_true]
                weapon_select = weapon_list_true[["weapon power"]].values[0][0]

                print(self.name, "You attack the", opponent.e_adjectives,
                      opponent.e_name, "with your", YourHero.x_see_inventory_name_qty(self, "weapon", 0),
                      "with atk power", YourHero.x_weapon_dmg(self, weapon_select))
                opponent.e_hp = opponent.e_hp - YourHero.x_weapon_dmg(self, weapon_select)
                if opponent.e_hp <= 0:
                    print("---------------------------------------------------------")
                    print("Enemy", opponent.e_adjectives, opponent.e_name, "is dead")
                    input("Press enter")
            else:
                print(self.name, "You attack the", opponent.e_adjectives,
                      opponent.e_name, "with atk power", self.atk)
                opponent.e_hp = opponent.e_hp - self.atk
                if opponent.e_hp <= 0:
                    print("---------------------------------------------------------")
                    print("Enemy", opponent.e_adjectives, opponent.e_name, "is dead")
                    input("Press enter")
        elif hit <= 1:
            print("You miss the", opponent.e_adjectives, opponent.e_name)

    # this function is not in use:
    def x_one_enemy_atk(self, opponent, choice):
        choice = choice
        if choice == "1" or choice == "attack":
            YourHero.x_attack_battle(self, opponent)
        elif choice == "2" or choice == "black magic":
            YourHero.x_battle_black_magic_use(self, opponent)
        else:
            print("please choose a valid number")
            input("Press enter to continue")

    def x_battle_two(self):
        enemy.ex_initials_stats()
        enemy_2.ex_initials_stats()
        enemy_3.ex_initials_stats()
        if opponents.count == 1:
            print(enemy.e_stat)
            print("---------------------------------------------------------")
            choice = input(" 1. Attack\n 2. Black Magic \n 3. White Magic \n 4. Item\n 5. Run\n 6. Defence\n"
                           " 7. Front\n 8. Steal\n what is your choice?: ")
            if choice == "1" or choice == "attack":
                YourHero.x_attack_battle(self, enemy)
            elif choice == "2" or choice == "black magic":
                YourHero.x_battle_black_magic_use(self, enemy)
            elif choice == "3" or choice == "white magic":
                YourHero.x_battle_white_magic_use(self)
            elif choice == "4" or choice == "item":
                YourHero.x_battle_item_use(self)
            elif choice == "5" or choice == "run away" or choice == "run":
                YourHero.x_run_away(self)
            elif choice == "6" or choice == "defence":
                YourHero.x_defence_battle(self)
            elif choice == "7" or choice == "front":
                YourHero.x_front_battle(self)
            elif choice == "8" or choice == "steal":
                YourHero.x_steal_battle(self)
            else:
                print("choose a valid choice")
                input("press enter")
        elif opponents.count == 2:
            print("---------------------------------------------------------")
            print(enemy.e_stat)
            print("---------------------------------------------------------")
            print(enemy_2.e_stat)
            choice = input(" 1. Attack\n 2. Black Magic \n 3. White Magic \n 4. Item\n 5. Run\n 6. Defence\n"
                           " 7. Front\n what is your choice?: ")
            if choice == "1" or choice == "attack":
                print("Select your enemy!!")
                print("Choose enemy to fight, enemy nr 1", enemy.e_adjectives, enemy.e_name,
                      "or enemy nr 2", enemy_2.e_adjectives, enemy_2.e_name)
                choose_enemy = input("What is your strategy?, nr 1 or nr 2: ")
                if choose_enemy == "1" and enemy.e_hp <= 0:
                    print("The", enemy.e_adjectives, enemy.e_name,
                          "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                    input("Press enter!")
                    YourHero.x_attack_battle(self, enemy_2)
                elif choose_enemy == "1" and enemy.e_hp > 0:
                    YourHero.x_attack_battle(self, enemy)
                elif choose_enemy == "2" and enemy_2.e_hp <= 0:
                    print("The", enemy_2.e_adjectives, enemy_2.e_name,
                          "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                    input("Press enter!")
                    YourHero.x_attack_battle(self, enemy)
                elif choose_enemy == "2" and enemy_2.e_hp > 0:
                    YourHero.x_attack_battle(self, enemy_2)
            elif choice == "2" or choice == "black magic":
                choose_enemy = input("What is your strategy?, nr 1 or nr 2: ")
                if choose_enemy == "1" and enemy.e_hp <= 0:
                    print("The", enemy.e_adjectives, enemy.e_name,
                          "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                    input("Press enter!")
                    YourHero.x_battle_black_magic_use(self, enemy_2)
                elif choose_enemy == "1" and enemy.e_hp > 0:
                    YourHero.x_battle_black_magic_use(self, enemy)
                elif choose_enemy == "2" and enemy_2.e_hp <= 0:
                    print("The", enemy_2.e_adjectives, enemy_2.e_name,
                          "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                    input("Press enter!")
                    YourHero.x_battle_black_magic_use(self, enemy)
                elif choose_enemy == "2" and enemy_2.e_hp > 0:
                    YourHero.x_battle_black_magic_use(self, enemy_2)
            elif choice == "3" or choice == "white magic":
                YourHero.x_battle_white_magic_use(self)
            elif choice == "4" or choice == "item":
                YourHero.x_battle_item_use(self)
            elif choice == "5" or choice == "run away" or choice == "run":
                YourHero.x_run_away(self)
            elif choice == "6" or choice == "defence":
                YourHero.x_defence_battle(self)
            elif choice == "7" or choice == "front":
                YourHero.x_front_battle(self)
            else:
                print("choose a valid choice")
                input("press enter")
        elif opponents.count == 3:
            print("---------------------------------------------------------")
            print(enemy.e_stat)
            print("---------------------------------------------------------")
            print(enemy_2.e_stat)
            print("---------------------------------------------------------")
            print(enemy_3.e_stat)
            choice = input(" 1. Attack\n 2. Black Magic \n 3. White Magic \n 4. Item\n 5. Run\n 6. Defence\n"
                           " 7. Front\n what is your choice?: ")
            if choice == "1" or choice == "attack":
                print("Select your enemy!!")
                print("Choose enemy to fight, enemy nr 1", enemy.e_adjectives, enemy.e_name,
                      "or enemy nr 2", enemy_2.e_adjectives, enemy_2.e_name,
                      "or enemy nr 3", enemy_3.e_adjectives, enemy_3.name)
                choose_enemy = input("What is your strategy?, nr 1, nr 2 or nr 3: ")
                if choose_enemy == "1":
                    if enemy.e_hp <= 0 and enemy_3.e_hp > 0:
                        print("The", enemy.e_adjectives, enemy.e_name, "and", enemy_2.e_adjectives, enemy_2.name,
                              "is dead. You attack the", enemy_3.e_adjectives, enemy_3.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy_3)
                    elif enemy.e_hp <= 0 and enemy_2.e_hp > 0:
                        print("The", enemy.e_adjectives, enemy.e_name, "and", enemy_3.e_adjectives, enemy_3.name,
                              "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy_2)
                    elif enemy.e_hp > 0:
                        YourHero.x_attack_battle(self, enemy)
                elif choose_enemy == "2":
                    if enemy_2.e_hp <= 0 and enemy.e_hp > 0:
                        print("The", enemy_2.e_adjectives, enemy_2.e_name, "and", enemy_3.e_adjectives, enemy_3.name,
                              "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy)
                    elif enemy_2.e_hp <= 0 and enemy_3.e_hp > 0:
                        print("The", enemy_2.e_adjectives, enemy_2.e_name, "and", enemy.e_adjectives, enemy.name,
                              "is dead. You attack the", enemy_3.e_adjectives, enemy_3.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy_3)
                    elif enemy_2.e_hp > 0:
                        YourHero.x_attack_battle(self, enemy_2)
                elif choose_enemy == "3":
                    if enemy_3.e_hp <= 0 and enemy_2.e_hp > 0:
                        print("The", enemy_3.e_adjectives, enemy_3.e_name, "and", enemy.e_adjectives, enemy.name,
                              "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy_2)
                    elif enemy_3.e_hp <= 0 and enemy.e_hp > 0:
                        print("The", enemy_3.e_adjectives, enemy_3.e_name, "and", enemy_2.e_adjectives, enemy_2.name,
                              "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_attack_battle(self, enemy)
                    elif enemy_3.e_hp > 0:
                        YourHero.x_attack_battle(self, enemy_3)
            elif choice == "2" or choice == "black magic":
                print("Select your enemy!!")
                print("Choose enemy to fight, enemy nr 1", enemy.e_adjectives, enemy.e_name,
                      "or enemy nr 2", enemy_2.e_adjectives, enemy_2.e_name,
                      "or enemy nr 3", enemy_3.e_adjectives, enemy_3.name)
                choose_enemy = input("What is your strategy?, nr 1, nr 2 or nr 3: ")
                if choose_enemy == "1":
                    if enemy.e_hp <= 0 and enemy_3.e_hp > 0:
                        print("The", enemy.e_adjectives, enemy.e_name, "and", enemy_2.e_adjectives, enemy_2.name,
                              "is dead. You attack the", enemy_3.e_adjectives, enemy_3.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy_3)
                    elif enemy.e_hp <= 0 and enemy_2.e_hp > 0:
                        print("The", enemy.e_adjectives, enemy.e_name, "and", enemy_3.e_adjectives, enemy_3.name,
                              "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy_2)
                    elif enemy.e_hp > 0:
                        YourHero.x_battle_black_magic_use(self, enemy)
                elif choose_enemy == "2":
                    if enemy_2.e_hp <= 0 and enemy.e_hp > 0:
                        print("The", enemy_2.e_adjectives, enemy_2.e_name, "and", enemy_3.e_adjectives, enemy_3.name,
                              "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy)
                    elif enemy_2.e_hp <= 0 and enemy_3.e_hp > 0:
                        print("The", enemy_2.e_adjectives, enemy_2.e_name, "and", enemy.e_adjectives, enemy.name,
                              "is dead. You attack the", enemy_3.e_adjectives, enemy_3.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy_3)
                    elif enemy_2.e_hp > 0:
                        YourHero.x_battle_black_magic_use(self, enemy_2)
                elif choose_enemy == "3":
                    if enemy_3.e_hp <= 0 and enemy_2.e_hp > 0:
                        print("The", enemy_3.e_adjectives, enemy_3.e_name, "and", enemy.e_adjectives, enemy.name,
                              "is dead. You attack the", enemy_2.e_adjectives, enemy_2.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy_2)
                    elif enemy_3.e_hp <= 0 and enemy.e_hp > 0:
                        print("The", enemy_3.e_adjectives, enemy_3.e_name, "and", enemy_2.e_adjectives, enemy_2.name,
                              "is dead. You attack the", enemy.e_adjectives, enemy.e_name, " instead!")
                        input("Press enter!")
                        YourHero.x_battle_black_magic_use(self, enemy)
                    elif enemy_3.e_hp > 0:
                        YourHero.x_battle_black_magic_use(self, enemy_3)
            elif choice == "3" or choice == "white magic":
                YourHero.x_battle_white_magic_use(self)
            elif choice == "4" or choice == "item":
                YourHero.x_battle_item_use(self)
            elif choice == "5" or choice == "run away" or choice == "run":
                YourHero.x_run_away(self)
            elif choice == "6" or choice == "defence":
                YourHero.x_defence_battle(self)
            elif choice == "7" or choice == "front":
                YourHero.x_front_battle(self)
            else:
                print("choose a valid choice")
                input("press enter")

    def x_enemy_dead(self, hero1, hero2, boss):
        play.music_loop(play.music_battle[1])
        print("---------------------------------------------------------")
        print(victory_text)
        print("---------------------------------------------------------")
        YourHero.x_win_battle(self, hero1, hero2, boss)

    def x_battle_one(self, hero1, hero2, boss):

        play.music_loop(play.music_battle[0])
        Enemy.ex_gen_enemy(enemy)
        enemy.e_hp = Enemy.ex_hp(enemy)
        enemy.ex_initials_stats()
        Enemy.ex_gen_enemy(enemy_2)
        enemy_2.ex_initials_stats()
        enemy_2.e_hp = Enemy.ex_hp(enemy_2)
        Enemy.ex_gen_enemy(enemy_3)
        enemy_3.ex_initials_stats()
        enemy_3.e_hp = Enemy.ex_hp(enemy_3)
        YourHero.x_initials_stats(self)
        opponents.battle_fights = 1
        if boss == "yes":
            enemy.e_adjectives = "Boss"
            battle_text = text2art("BOSS_BATTLE!!!")
            YourHero.x_battle_one_part_2(self, hero1, hero2, boss, battle_text)
        elif boss == "no":
            battle_text = text2art("BATTLE!!!")
            YourHero.x_battle_one_part_2(self, hero1, hero2, boss, battle_text)

    def x_battle_one_part_2(self, hero1, hero2, boss, battle_text):
        while True:
            enemy.ex_initials_stats()
            enemy_2.ex_initials_stats()
            enemy_3.ex_initials_stats()
            YourHero.x_initials_stats(self)
            if opponents.count == 1:
                opponents.opponent = enemy
            elif opponents.count == 2:
                op1_count = 1
                op2_count = 3
                if opponents.battle_fights == op1_count:
                    if enemy.e_hp <= 0:
                        opponents.opponent = enemy_2
                    elif enemy.e_hp > 0:
                        opponents.opponent = enemy
                        opponents.battle_fights += 2
                elif opponents.battle_fights == op2_count:
                    if enemy_2.e_hp <= 0:
                        opponents.opponent = enemy
                    elif enemy_2.e_hp > 0:
                        opponents.opponent = enemy_2
                        opponents.battle_fights -= 2
            elif opponents.count == 3:
                op1_count = 1
                op2_count = 3
                op3_count = 5
                if opponents.battle_fights == op1_count:
                    if enemy.e_hp <= 0 and enemy_2.e_hp <= 0:
                        opponents.opponent = enemy_3
                    elif enemy.e_hp <= 0 and enemy_3.e_hp <= 0:
                        opponents.opponent = enemy_2
                    elif enemy.e_hp > 0:
                        opponents.opponent = enemy
                        opponents.battle_fights += 2
                elif opponents.battle_fights == op2_count:
                    if enemy_2.e_hp <= 0 and enemy_3.e_hp <= 0:
                        opponents.opponent = enemy
                    elif enemy.e_hp <= 0 and enemy_2.e_hp <= 0:
                        opponents.opponent = enemy_3
                    elif enemy_2.e_hp > 0:
                        opponents.opponent = enemy_2
                        opponents.battle_fights += 2
                elif opponents.battle_fights == op3_count:
                    if enemy_3.e_hp <= 0 and enemy_2.e_hp <= 0:
                        opponents.opponent = enemy
                    elif enemy_3.e_hp <= 0 and enemy.e_hp <= 0:
                        opponents.opponent = enemy_2
                    elif enemy_3.e_hp > 0:
                        opponents.opponent = enemy_3
                        opponents.battle_fights -= 4
            if opponents.opponent.e_hp > 0:
                print("---------------------------------------------------------")
                print(battle_text)
                print("---------------------------------------------------------")
                print("You can look at your stats, write stats, status screen or battle to enter battle")
                user_input = input("What will it be?\n")
                if user_input.lower() == "status screen" or user_input.lower() == "stats":
                    print(hero1.item_list)
                    print("---------------------------------------------------------")
                    print(hero2.item_list)
                    print("---------------------------------------------------------")
                    input("Press enter to continue")
                    print(hero1.total_stat)
                    print("---------------------------------------------------------")
                    print(hero2.total_stat)
                    input("Press enter to continue\n")
                elif user_input == "battle":
                    YourHero.x_status_check(hero1, opponents.opponent)
                    YourHero.x_status_check(hero2, opponents.opponent)
                    if hero1.hp <= 0:
                        if hero2.hp <= 0:
                            YourHero.x_game_over(hero1)
                            YourHero.x_game_over(hero2)
                        elif hero2.hp > 0:
                            print("---------------------------------------------------------")
                            print(hero2.stat)
                            print("---------------------------------------------------------")
                            print(hero1.name, " is DEAD!!")
                            input("Press enter")
                            YourHero.x_battle_two(hero2)
                            if self.run == 1:
                                self.run = 0
                                break
                            if opponents.opponent.e_hp > 0:
                                opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                                # Enemy.ex_one_enemy_atk(opponents.opponent)
                    elif hero1.status_paralyzed >= 1:
                        hero1.status_paralyzed = 0
                        YourHero.x_status_check(hero2, opponents.opponent)
                        if hero2.hp <= 0:
                            YourHero.x_game_over(hero2)
                        else:
                            print("---------------------------------------------------------")
                            print(hero2.stat)
                            print("---------------------------------------------------------")
                            print(hero1.name, " is Paralyzed!!")
                            input("Press enter")
                            YourHero.x_battle_two(hero2)
                            if self.run == 1:
                                self.run = 0
                                break
                            if opponents.opponent.e_hp > 0:
                                opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                    else:
                        print("---------------------------------------------------------")
                        print(hero1.stat)
                        print("---------------------------------------------------------")
                        YourHero.x_battle_two(hero1)
                        if self.run == 1:
                            self.run = 0
                            break
                        if opponents.opponent.e_hp > 0:
                            opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                        opponents.opponent.ex_initials_stats()
                        enemy.ex_initials_stats()
                        enemy_2.ex_initials_stats()
                        enemy_3.ex_initials_stats()
                        YourHero.x_initials_stats(self)
                        if opponents.count == 1:
                            opponents.opponent = enemy
                        elif opponents.count == 2:
                            op1_count = 1
                            op2_count = 3
                            if opponents.battle_fights == op1_count:
                                if enemy.e_hp <= 0:
                                    opponents.opponent = enemy_2
                                elif enemy.e_hp > 0:
                                    opponents.opponent = enemy
                                    opponents.battle_fights += 2
                            elif opponents.battle_fights == op2_count:
                                if enemy_2.e_hp <= 0:
                                    opponents.opponent = enemy
                                elif enemy_2.e_hp > 0:
                                    opponents.opponent = enemy_2
                                    opponents.battle_fights -= 2
                        elif opponents.count == 3:
                            op1_count = 1
                            op2_count = 3
                            op3_count = 5
                            if opponents.battle_fights == op1_count:
                                if enemy.e_hp <= 0 and enemy_2.e_hp <= 0:
                                    opponents.opponent = enemy_3
                                elif enemy.e_hp <= 0 and enemy_3.e_hp <= 0:
                                    opponents.opponent = enemy_2
                                elif enemy.e_hp > 0:
                                    opponents.opponent = enemy
                                    opponents.battle_fights += 2
                            elif opponents.battle_fights == op2_count:
                                if enemy_2.e_hp <= 0 and enemy_3.e_hp <= 0:
                                    opponents.opponent = enemy
                                elif enemy.e_hp <= 0 and enemy_2.e_hp <= 0:
                                    opponents.opponent = enemy_3
                                elif enemy_2.e_hp > 0:
                                    opponents.opponent = enemy_2
                                    opponents.battle_fights += 2
                            elif opponents.battle_fights == op3_count:
                                if enemy_3.e_hp <= 0 and enemy_2.e_hp <= 0:
                                    opponents.opponent = enemy
                                elif enemy_3.e_hp <= 0 and enemy.e_hp <= 0:
                                    opponents.opponent = enemy_2
                                elif enemy_3.e_hp > 0:
                                    opponents.opponent = enemy_3
                                    opponents.battle_fights -= 4
                        if opponents.opponent.e_hp > 0:
                            YourHero.x_status_check(hero1, opponents.opponent)
                            YourHero.x_status_check(hero2, opponents.opponent)
                            if hero2.hp <= 0:
                                if hero1.hp <= 0:
                                    YourHero.x_game_over(hero1)
                                    YourHero.x_game_over(hero2)
                                elif hero1.hp > 0:
                                    print("---------------------------------------------------------")
                                    print(hero1.stat)
                                    print("---------------------------------------------------------")
                                    print(hero2.name, "is DEAD!!")
                                    input("Press enter")
                                    YourHero.x_battle_two(hero1)
                                    if self.run == 1:
                                        self.run = 0
                                        break
                                    if opponents.opponent.e_hp > 0:
                                        opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                            elif hero2.status_paralyzed >= 1:
                                hero2.status_paralyzed = 0
                                YourHero.x_status_check(hero1, opponents.opponent)
                                if hero1.hp <= 0:
                                    YourHero.x_game_over(hero1)
                                else:
                                    print("---------------------------------------------------------")
                                    print(hero1.stat)
                                    print("---------------------------------------------------------")
                                    print(hero2.name, " is Paralyzed!!")
                                    input("Press enter")
                                    # Maybe check if opponent is dead?
                                    YourHero.x_battle_two(hero1)
                                    if self.run == 1:
                                        self.run = 0
                                        break
                                    if opponents.opponent.e_hp > 0:
                                        opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                            else:
                                print("---------------------------------------------------------")
                                print(hero2.stat)
                                print("---------------------------------------------------------")
                                YourHero.x_battle_two(hero2)
                                if self.run == 1:
                                    self.run = 0
                                    break
                                if opponents.opponent.e_hp > 0:
                                    opponents.opponent.ex_one_enemy_atk(hero1, hero2, boss)
                        else:
                            YourHero.x_enemy_dead(self, hero1, hero2, boss)
                else:
                    print("Invalid choice")
            else:
                YourHero.x_enemy_dead(self, hero1, hero2, boss)

    def x_battle(self, level, hero1, hero2, boss):
        if boss == "yes":
            hero1 = hero1
            hero2 = hero2
            self.enemy_boss_level = level
            enemy_level = pd.read_excel(dir_path.level_path + level_file, sheet_name=level_file_sheet_strenght_mod_enemy)
            dt = pd.DataFrame(enemy_level)
            dt.set_index("Level", inplace=True)
            new_enemy_strength = dt.loc[level]["Bonus"]
            bonus = new_enemy_strength
            # enemy.e_strength = bonus
            enemy.e_factor_for_level_up += level
            enemy.e_level_count += level
            enemy.e_bonus_factor = bonus
            enemy.e_level += level
            enemy_2.e_level = enemy_2.e_level
            enemy_3.e_level = enemy_3.e_level
            # enemy.ex_level_up_f()
            # enemy_2.ex_level_up_f()
            # enemy_3.ex_level_up_f()
            enemy.ex_initials_stats()
            enemy_2.ex_initials_stats()
            enemy_3.ex_initials_stats()
            opponents.count = 1
            YourHero.x_battle_one(cloud, hero1, hero2, boss)
            return hero1 and hero2
        elif boss == "no":
            hero1 = hero1
            hero2 = hero2
            enemy.e_level = enemy.e_level + level
            enemy_2.e_level = enemy_2.e_level + level
            enemy_3.e_level = enemy_3.e_level + level
            # enemy.ex_level_up_f()
            # enemy_2.ex_level_up_f()
            # enemy_3.ex_level_up_f()
            enemy.ex_initials_stats()
            enemy_2.ex_initials_stats()
            enemy_3.ex_initials_stats()
            numbers_of_enemies = random.randint(0, 6)
            if numbers_of_enemies <= 2:
                opponents.count = 1
                YourHero.x_battle_one(cloud, hero1, hero2, boss)
            elif numbers_of_enemies == 3 or numbers_of_enemies <= 5:
                opponents.count = 2
                YourHero.x_battle_one(cloud, hero1, hero2, boss)
            elif numbers_of_enemies >= 6:
                opponents.count = 3
                YourHero.x_battle_one(cloud, hero1, hero2, boss)
            return hero1 and hero2


class Enemy(YourHero):
    def __init__(self):
        super().__init__()
        self.e_strength = 0
        self.e_hp = 13
        self.e_max_hp = 0.0
        self.e_level = 1
        self.e_exp = 0
        self.e_next_level_exp = 0
        self.e_exp_give = 0
        self.e_level_count = self.e_level - 1
        self.e_factor_for_level_up = 2
        self.e_exp = 0
        self.e_name = ""
        self.e_adjectives = ""
        self.e_stat = {}
        self.e_defence = 0.0
        self.e_atk = 0.0
        self.e_poison_dmg = 0
        self.e_bonus_factor = 0
        self.e_loot = ""
        self.e_gil = 0
        self.e_special_atk_count = 0

    def ex_gil_give(self):
        # gill_calc = 3 * self.e_bonus_factor ** 1.5 + 10
        gill_calc = math.floor(10 * self.e_factor_for_level_up ** 1.5)
        self.e_gil = gill_calc
        return gill_calc

    def ex_level_up(self):
        next_level_exp = math.floor(.1 * (self.e_factor_for_level_up ** 4) + 4.2 * (self.e_factor_for_level_up ** 3)
                                    + 6.1 * (self.e_factor_for_level_up ** 2)
                                    + (1.4 * self.e_factor_for_level_up) - 11.4)
        x_next_level = next_level_exp
        self.e_next_level_exp = x_next_level
        return next_level_exp

    def ex_level_up_give(self):
        e_next_level_exp = math.floor((.1 * (self.e_factor_for_level_up ** 4) + 4.2 *
                                       (self.e_factor_for_level_up ** 3) + 6.1 *
                                       (self.e_factor_for_level_up ** 2) + 1.4 *
                                       self.e_factor_for_level_up - 11.4) / self.e_factor_for_level_up ** 2)
        exp_give = e_next_level_exp
        self.e_exp_give = exp_give
        return exp_give

    def ex_strength(self):
        s_strength_base = ([2.5, 3.14, 3.82, 4.54, 5.3, 6.1, 6.94, 7.82, 8.74, 9.7, 10.62, 11.5,
                            12.34, 13.14, 13.9, 14.62, 15.3, 15.94, 16.62, 17.34])
        s_strength_calc = round(s_strength_base[self.e_level_count]
                                + (self.e_level * 3 / 10) + (self.e_bonus_factor / 32), 1)

        strength = s_strength_calc
        self.e_strength = strength
        return strength

    def ex_hp(self):

        hp_mod = ([250, 314, 382, 454, 530, 610, 694, 782, 874, 970,
                   1062, 1150, 1234, 1314, 1390, 1462, 1530, 1594, 1662, 1734])

        hp_calc = math.ceil(self.e_strength * hp_mod[self.e_level_count] * self.e_level / 55)
        h_hp = hp_calc
        # self.e_hp = round(random.uniform(h_hp, h_hp * 1.2), 0)
        # self.e_hp = h_hp
        self.e_max_hp = h_hp
        return h_hp

    def ex_atk(self):

        atk_calc = math.floor(self.e_max_hp * self.e_strength / self.e_level / 25)
        e_atk = round(random.uniform(atk_calc, atk_calc * 1.2), 0)
        self.e_atk = e_atk
        if cloud.defence_count > 0:
            atk_calc = math.ceil(self.e_max_hp * self.e_strength / self.e_level / 40)
            e_atk = round(random.uniform(atk_calc, atk_calc * 1.2), 0)
            self.e_atk = e_atk
        elif elena.defence_count > 0:
            atk_calc = math.ceil(self.e_max_hp * self.e_strength / self.e_level / 40)
            e_atk = round(random.uniform(atk_calc, atk_calc * 1.2), 0)
            self.e_atk = e_atk
        return e_atk

    def ex_defence(self):
        defence_calc = math.ceil(self.e_strength + self.e_level * 3 / 10 + self.e_atk / 32)
        defence = defence_calc
        self.e_defence = defence
        return defence

    def ex_poison_dmg(self):
        poison = math.ceil(enemy.e_max_hp * enemy.e_strength / 10 / 6)
        poison_dmg = poison
        self.e_poison_dmg = poison_dmg
        return poison_dmg

    def ex_stat_all_enemy(self):

        dots = ":     "

        name_dot = self.e_adjectives + " " + self.e_name + dots

        skills = {name_dot: [""],
                  "Attributes": ["Level", "HP", "Atk", "Def", "next exp"],
                  "Points": [self.e_level, self.e_hp, self.e_atk, self.e_defence, self.e_next_level_exp], }

        matrix_skills = pd.DataFrame(skills, index=["", "", "", "", ""])

        self.e_stat = matrix_skills
        return matrix_skills

    def ex_stat(self):
        if self.e_hp <= 0:
            dots = ":     "

            name_dot = self.e_adjectives + " " + self.e_name + dots

            skills = {name_dot: [""],
                      "Status": ["Enemy is dead"]}

            matrix_skills = pd.DataFrame(skills, index=[""])

            self.e_stat = matrix_skills
            return matrix_skills
        else:
            dots = ":     "

            name_dot = self.e_adjectives + " " + self.e_name + dots

            skills = {name_dot: [""],
                      "Attributes": ["Level", "HP"],
                      "Points": [self.e_level, self.e_hp]}

            matrix_skills = pd.DataFrame(skills, index=["", ""])

            self.e_stat = matrix_skills
            return matrix_skills

    def ex_level_up_f(self):
        # self.e_exp = self.e_exp + Enemy.ex_level_up(self)
        while self.e_exp >= self.e_next_level_exp:
            self.e_next_level_exp = Enemy.ex_level_up(self)
            if self.e_exp >= self.e_next_level_exp:
                self.e_exp = self.e_exp - self.e_next_level_exp
                self.e_factor_for_level_up += 1
                self.e_next_level_exp = Enemy.ex_level_up(self)
                self.e_level += 1
                self.e_level_count += 1
                self.e_bonus_factor += 3
                Enemy.ex_initials_stats(self)
        # print("Enemy new level!!!")
        else:
            Enemy.ex_initials_stats(self)
            # self.e_exp = self.e_exp + self.ex_level_up_give()
            # print("Enemy need to die more to raise in level")

    def ex_initials_stats(self):
        self.e_strength = Enemy.ex_strength(self)
        # self.e_hp = Enemy.ex_hp(self)
        self.e_max_hp = Enemy.ex_hp(self)
        self.e_atk = Enemy.ex_atk(self)
        self.e_defence = Enemy.ex_defence(self)
        self.e_stat = Enemy.ex_stat(self)
        self.e_gil = Enemy.ex_gil_give(self)

    def ex_gen_enemy(self):

        enemy_list = pd.read_excel(dir_path.enemy_path + enemy_name_file)
        df = pd.DataFrame(enemy_list)
        adjective_list = df["adjective"].to_list()
        name_list = df["name"].to_list()
        adjective = secrets.choice(adjective_list)
        name = secrets.choice(name_list)
        self.e_adjectives = adjective
        self.e_name = name
        Enemy.ex_initials_stats(self)
        return name, adjective

    def ex_weapon_use_boss_special(self):
        tprint("Boss")
        tprint("Super_attack")
        if self.e_name == "Spider":
            attack_text = "summons the fear full webb"
            return attack_text
        elif self.e_name == "Giant":
            attack_text = "jumps high up in the sky, and smash the earth with great power"
            return attack_text
        elif self.e_name == "Dragon":
            attack_text = "eyes turns all red, and the burning flames bursts at you"
            return attack_text
        elif self.e_name == "Snake":
            attack_text = "raises high to the sky, unleashing many power full strikes at you"
            return attack_text
        elif self.e_name == "Lizard":
            attack_text = "blows out the greatest venomus fog at you"
            return attack_text

    def ex_weapon_use_text(self):
        enemy_weapon_battle = ["Brutal Sword", "Battle Axe", "Dagger", "Iron Flail", "Steel War Pike"]
        weapon_of_use = random.randint(0, 4)
        if weapon_of_use == 0:
            attack_text = "swings the" + " " + enemy_weapon_battle[weapon_of_use] + " " + "at"
            return attack_text
        elif weapon_of_use == 1:
            attack_text = "runs at you with its" + " " + enemy_weapon_battle[weapon_of_use] + " " + "and slashing"
            return attack_text
        elif weapon_of_use == 2:
            attack_text = "sneaks up on you and stab you with the" + " " + enemy_weapon_battle[weapon_of_use]
            return attack_text
        elif weapon_of_use == 3:
            attack_text = "starts to swing the" + " " + enemy_weapon_battle[weapon_of_use] + " " + \
                          "the sound from the chains howls in your ear, before it hits"
            return attack_text
        elif weapon_of_use == 4:
            attack_text = "screams at you while running with its" + " " + enemy_weapon_battle[weapon_of_use] + " " \
                          + "pierce"
            return attack_text

    def ex_if_boss(self, boss):
        if boss == "yes":
            if self.e_hp <= 50 % self.e_max_hp:
                self.e_special_atk_count += 1
                if self.e_special_atk_count == 1:
                    return Enemy.ex_weapon_use_boss_special(self)
                elif self.e_special_atk_count >= 2 and self.e_hp < 21 % self.e_max_hp:
                    return Enemy.ex_weapon_use_boss_special(self)
                else:
                    return Enemy.ex_weapon_use_text(self)
            else:
                return Enemy.ex_weapon_use_text(self)
        elif boss == "no":
            return Enemy.ex_weapon_use_text(self)

    def ex_boss_special_atk(self):
        if self.e_hp <= 50 % self.e_max_hp:
            if self.e_special_atk_count == 1:
                return 1.5
            elif self.e_special_atk_count == 2 and self.e_hp <= 20 % self.e_max_hp:
                return 1.8
            else:
                return 1
        else:
            return 1

    def ex_one_enemy_atk(self, hero1, hero2, boss):
        hit = random.randint(0, 10)
        if hit >= 3:
            if hero2.hp <= 0:
                print("The", self.e_adjectives, self.e_name, Enemy.ex_if_boss(self, boss),
                      hero1.name, "with the attack power", self.e_atk if boss == "no"
                      else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self)))
                print("---------------------------------------------------------")
                input("press enter")
                hero1.hp = hero1.hp - self.e_atk if boss == "no" \
                    else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self))
                self.ex_poison_dmg()
                YourHero.x_status_effect(hero1, self)
                YourHero.x_initials_stats(hero1)
                YourHero.x_game_over(hero2)
            elif hero1.hp <= 0:
                print("The", self.e_adjectives, self.e_name, Enemy.ex_if_boss(self, boss),
                      hero2.name, "with the attack power", self.e_atk if boss == "no"
                      else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self)))
                print("---------------------------------------------------------")
                input("press enter")
                hero2.hp = hero2.hp - self.e_atk if boss == "no" \
                    else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self))
                self.ex_poison_dmg()
                YourHero.x_status_effect(hero2, self)
                YourHero.x_initials_stats(hero2)
                YourHero.x_game_over(hero1)
            elif hero2.hp > 0 and hero1.hp > 0:
                hero_choice = random.randint(0, 10)
                if hero_choice <= 5:
                    print("The", self.e_adjectives, self.e_name, Enemy.ex_if_boss(self, boss),
                          hero1.name, "with the attack power", self.e_atk if boss == "no"
                          else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self)))
                    print("---------------------------------------------------------")
                    input("press enter")
                    hero1.hp = hero1.hp - self.e_atk if boss == "no" \
                        else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self))
                    self.ex_poison_dmg()
                    YourHero.x_status_effect(hero1, self)
                    YourHero.x_initials_stats(hero1)
                    YourHero.x_game_over(hero1)
                elif hero_choice >= 6:
                    print("The", self.e_adjectives, self.e_name, Enemy.ex_if_boss(self, boss),
                          hero1.name, "with the attack power", self.e_atk if boss == "no"
                          else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self)))
                    print("---------------------------------------------------------")
                    input("press enter")
                    hero2.hp = hero2.hp - self.e_atk if boss == "no" \
                        else math.floor(self.e_atk * Enemy.ex_boss_special_atk(self))
                    self.ex_poison_dmg()
                    YourHero.x_status_effect(hero2, self)
                    YourHero.x_initials_stats(hero2)
                    YourHero.x_game_over(hero2)
            # else:
        # YourHero.x_game_over(cloud and elena)
        elif hit <= 2:
            print("The", enemy.e_adjectives, enemy.e_name, "MISS")
            print("---------------------------------------------------------")
            input("press enter")


class ChooseEnemy:
    def __init__(self):
        self.count = 0
        self.opponent = self.opponent
        self.battle_fights = 0

    def opponent(self):
        opponent = self.opponent
        return opponent


class Music:


    music_chapter_1_mp3 = None
    music_for_all_mp3 = None
    music_battle_mp3 = None
    music_main_mp3 = None


    def __init__(self):
        self.music_main_mp3 = ["intro.mp3", "prologue.mp3"]
        self.music_for_all_mp3 = ["sleep_song.mp3","sleeping_at_in.mp3"]
        self.music_battle_mp3 = ["battle_one.mp3", "battle_win.mp3"]
        # "chapter_1_the_crossing.mp3"
        self.music_chapter_1_mp3 = ["chapter_1_the_crossing.mp3", "chapter_1_town_abreheim.mp3", "chapter_1_item_shop.mp3",
                                "chapter_1_abrehiem_ally.mp3", "chapter_1_round_room.mp3", "chapter_1_abreheim_inn.mp3",
                                "chapter_1_weapon_shop.mp3", "chapter_1_abrehiem_ally_dark.mp3",
                                "chapter_1_forest.mp3", "chapter_1_snow_area.mp3", "chapter_1_mountain.mp3",
                                    "chapter_1_materia_shop.mp3", "chapter_1_basement.mp3"]

        self.sound_chapter_1 = ["chapter_1_the_crossing", "chapter_1_town_abreheim", "chapter_1_item_shop",
                                "chapter_1_abrehiem_ally", "chapter_1_round_room", "chapter_1_abreheim_inn",
                                "chapter_1_weapon_shop", "chapter_1_abrehiem_ally_dark",
                                "chapter_1_forest", "chapter_1_snow_area", "chapter_1_mountain",
                                "chapter_1_materia_shop", "chapter_1_basement"]


        self.music_main = ["intro", "prologue"]
        self.music_battle = ["x_battle", "battle_win"]
        self.music_for_all = ["sleep_song", "sleeping_at_in"]

    def music_loop(self, music_name):
        elena.music_name = music_name

        if elena.music_name == self.music_main[0]:
            os.chdir(dir_path.music_main_intro)
            pygame.mixer.music.load(play.music_main_mp3[0])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.music_main[1]:
            os.chdir(dir_path.music_main_intro)
            pygame.mixer.music.load(play.music_main_mp3[1])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.music_for_all[1]:
            os.chdir(dir_path.battle_path)
            pygame.mixer.music.load(play.music_for_all[1])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.music_for_all[0]:
            os.chdir(dir_path.for_all_path)
            pygame.mixer.music.load(play.music_for_all_mp3[0])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.music_battle[0]:
            os.chdir(dir_path.battle_path)
            pygame.mixer.music.load(play.music_battle_mp3[0])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.music_battle[1]:
            os.chdir(dir_path.battle_path)
            pygame.mixer.music.load(play.music_battle_mp3[1])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[0]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[0])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[1]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[1])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[2]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[2])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[3]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[3])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[4]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[4])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[5]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[5])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[6]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[6])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[7]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[7])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[8]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[8])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[9]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[9])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[10]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[10])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[11]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[11])
            pygame.mixer.music.play(-1)
        elif elena.music_name == self.sound_chapter_1[12]:
            os.chdir(dir_path.chapter_1_path)
            pygame.mixer.music.load(play.music_chapter_1_mp3[12])
            pygame.mixer.music.play(-1)


        else:
            print("end of music")


class StoryName:
    def __init__(self):
        self.chap_1_names = ["Intro", "", "the_crossing", "abreheim_town_in", "ally_in_abrehiem",
                             "dark_ally_in_abrehiem", "round_room", "basement", "south_snow"]

        self.chapter_names = ["Chapter 1", "Chapter 2", "Chapter 3", "Chapter 4"]
        self.progress = ""

    def story_name_loop(self, chap, name):
        elena.story_name = name
        elena.chapter = chap
        cloud.story_name = name
        cloud.chapter = chap
        return name and chap


class Timer:
    def __init__(self):
        self.start = time.time()
        self.end = time.time()
        self.pause = time.time()
        self.resume = 0.0
        self.time_stat = ""

    def start_time(self):
        my_timer.start = time.time()

    def pause_time(self):
        my_timer.pause = time.time()
        if my_timer.pause > time.time():
            my_timer.pause = my_timer.pause - my_timer.pause
        else:
            my_timer.pause = time.time()

    def restart(self):
        my_timer.resume = time.time() - my_timer.pause
        my_timer.start = my_timer.start + my_timer.resume

    def get_time_hhmmss(self):
        end = time.time()
        m, s = divmod(end - my_timer.start, 60)
        h, m = divmod(m, 60)
        time_str = "%02d:%02d:%02d" % (h, m, s)
        return time_str

    def print_time(self):
        time_hhmmss = my_timer.get_time_hhmmss()
        my_timer.time_stat = time_hhmmss
        # print(time_hhmmss)


class CodeDigits:
    def __init__(self):
        self.code_4_digits_box_1 = random.randint(1001, 9898)  # hero painting
        self.code_1_digits_box_1 = random.randint(0, 9)  # eagle painting
        self.code_1_digits_box_2 = random.randint(0, 9)  # eagle painting
        self.code_1_digits_box_3 = random.randint(0, 9)  # eagle painting
        self.code_1_digits_box_4 = random.randint(0, 9)  # eagle painting
        self.code_box_1_digit_collect_1_2_3_4 = int('%d%d%d%d' % (self.code_1_digits_box_1, self.code_1_digits_box_2,
                                                                  self.code_1_digits_box_3, self.code_1_digits_box_4))
        self.keep_code_4_digits_box_1 = self.code_4_digits_box_1


class GetImportant(Music):
    def __init__(self):
        super().__init__()
        
        # self.install_file = "#_code_chris_098421_path_finder.txt"
        self.important = ""
        self.folder_name = ["save", "enemy", "inventory_items", "magic", "music", "weapons_armor", "level"]
        self.folder_music = ["main_intro", "chapter_1", "for all", "battle"]
        self.music_main_mp3 = Music.music_main_mp3
        self.music_chapter_1_mp3 = Music.music_chapter_1_mp3
        self.music_for_all_mp3 = Music.music_for_all_mp3
        self.music_battle_mp3 = Music.music_battle_mp3
        self.dir_path = []
        self.save = 0
        self.enemy = 1
        self.inventory_items = 2
        self.magic = 3
        self.music = 4
        self.weapons_armor = 4
        self.level = 5
        self.main_intro_intro = 6
        self.chapter_1_path = 7
        self.for_all_path = 8
        self.battle_path = 9

    def important_paths(self):
        # place_find = os.path.abspath(self.install_file)
        use_path_folder = CURR_DIR_PATH # place_find[:-36]
        # print(use_path_folder)
        f = open("lost_shadow_installed.txt", "w+")
        f.write(use_path_folder)
        use_pdf_input = use_path_folder.replace("\\", "\\\\")
        extra_backslash = "\\\\"
        for i in self.folder_name:
            important_directory = use_pdf_input + extra_backslash + i + extra_backslash
            self.dir_path.append(important_directory)
        GetImportant.music_path_2(self)
        # input("press enter")
        f.close()
        return self.dir_path

    def install_game(self):
        print("\nWelcome to install menu")
        print("----------------------------------------------------------------------------------------------")
        GetImportant.important_paths(self)
        self.dir_path.pop(4)
        print("Installing game, at these directories")
        for x in range(0, 9):
            print(self.dir_path[x])
        print("Game installed!!\n Happy Gaming")
        input("press enter to start")
        print("----------------------------------------------------------------------------------------------")
        return self.dir_path
    # music_path was used to get path and file name in same string, didnt work for pygame mixer.
    def music_path(self):
        extra_backslash = "\\\\"
        # "main_intro", "chapter_1", "for all", "battle"]
        for i in self.folder_music:
            if i == "main_intro":
                for x in self.music_main_mp3:
                    main_music = self.dir_path[self.music] + i + extra_backslash + x
                    self.dir_path.append(main_music)
                    # return self.dir_path
            elif i == "chapter_1":
                for x in self.music_chapter_1_mp3:
                    chapter_1_music = self.dir_path[self.music] + i + extra_backslash + x
                    self.dir_path.append(chapter_1_music)
                    # return self.dir_path
            elif i == "for all":
                for x in self.music_for_all_mp3:
                    for_all_music = self.dir_path[self.music] + i + extra_backslash + x
                    self.dir_path.append(for_all_music)
                    # return self.dir_path
            elif i == "battle":
                for x in self.music_battle_mp3:
                    battle_music = self.dir_path[self.music] + i + extra_backslash + x
                    self.dir_path.append(battle_music)
                    # return self.dir_path
    def music_path_2(self):
        extra_backslash = "\\\\"
        # "main_intro", "chapter_1", "for all", "battle"]
        for i in self.folder_music:
            if i == "main_intro":
                main_music = self.dir_path[self.music] + i + extra_backslash
                self.dir_path.append(main_music)
            elif i == "chapter_1":
                chapter_1_music = self.dir_path[self.music] + i + extra_backslash
                self.dir_path.append(chapter_1_music)
            elif i == "for all":
                for_all_music = self.dir_path[self.music] + i + extra_backslash
                self.dir_path.append(for_all_music)
            elif i == "battle":
                battle_music = self.dir_path[self.music] + i + extra_backslash
                self.dir_path.append(battle_music)


class DirectoryPaths:
    def __init__(self):
        self.save_path = ""
        self.enemy_path = ""
        self.inventory_items_path = ""
        self.magic_path = ""
        self.weapons_armor_path = ""
        self.level_path = ""
        self.music_main_intro = ""
        self.chapter_1_path = ""
        self.for_all_path = ""
        self.battle_path = ""





class InventoryItems:
    def __init__(self):
        self.abreheim_shop = [
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ]
        self.abreheim_shop_defaulf = [
            ['Items', 'Cost', "Quantity"],
            ['potion', 50, 5],
            ['ether', 1500, 5],
            ['antidote', 80, 5],
            ['phoenix_down', 300, 5],
            ['silver_dust', 150, 5],
            ['tent', 3000, 5],
        ]
        self.store_count = [
            ['town', 'start', 'end', 'count'],
            ["Abreheim's items", 0, 0, ""],
            ["Lucy", 0, 0, ""],
        ]
        self.buy_history = [

            ['Items', 'count', "cost", "temp_cost", "price"],
            ['potion', 0, 0, "", 50],
            ['ether', 0, 0, "", 1500],
            ['Antidote', 0, 0, "", 80],
            ['Phoenix_down', 0, 0, "", 300],
            ['Silver_dust', 0, 0, "", 150],
            ['Tent', 0, 0, "", 3000],

        ]


class Cast_Magic:
    def __init__(self):
        self.magic = [
            ["Magic", "spell power", "mp cost", "gil cost", "Magic Type"],
            ['fire', 8, 4, 600, "black magic"],
            ['fire2', 20, 22, "", "black magic"],
            ['fire3', 64, 52, "", "black magic"],
            ['fire master', 64, 52, "", "black magic"],
            ['ice', 8, 4, 600, "black magic"],
            ['ice2', 20, 22, "", "black magic"],
            ['ice3', 64, 52, "", "black magic"],
            ['ice master', 64, 52, "", "black magic"],
            ['storm', 9, 8, 1200, "black magic"],
            ['storm2', 22, 30, "", "black magic"],
            ['storm3', 70, 60, "", "black magic"],
            ['storm master', 70, 60, "", "black magic"],
            ['quake', 9, 8, 1200, "black magic"],
            ['quake2', 22, 30, "", "black magic"],
            ['quake3', 70, 60, "", "black magic"],
            ['quake master', 70, 60, "", "black magic"],
            ['heal', 1.5, 24, 3000, "white magic"],
            ['heal2', 3, 48, "", "white magic"],
            ['life', 1, 80, "", "white magic"],

        ]

        self.magic_gil = [
            ["Magic", "gil cost"],
            ['fire', 600],
            ['ice', 600],
            ['storm', 1200],
            ['quake', 1200],
            ['heal', 3000],
            ['life', 6000],
        ]

        self.magic_sell = [
            ["Magic", "Sell"],
            ['fire', 300],
            ['fire2', 3000],
            ['fire3', 6000],
            ['fire master', 12000],
            ['ice', 300],
            ['ice2', 3000],
            ['ice3', 6000],
            ['ice master', 12000],
            ['storm', 600],
            ['storm2', 6000],
            ['storm3', 12000],
            ['storm master', 24000],
            ['quake', 600],
            ['quake2', 6000],
            ['quake3', 12000],
            ['quake master', 24000],
            ['heal', 1500],
            ['heal2', 6000],
            ['life', 3000],
        ]


class EnemyName:
    def __init__(self):
        self.enemy_name = [
            ["adjective", "name", "status_give", "immune", "loot", "qty_loot"],
            ['Big', "Giant", "", "", "potion", 1],
            ['Scary', "Spider", "", "", "silver_dust", 1],
            ['Unfatithful', "Dragon", "", "fire", "potion", 1],
            ['Demon', "Snake", "poison", "", "potion", 1],
            ['Brutal', "Lizard", "", "", "antidote", 1],
        ]


class LevelEnemy:
    def __init__(self):
        self.enemy_level = [
            ["Level", "StrBase", "", "Bonus", "Bonus Parameter", "StrBase Parameter", "Strength",
             "AtkFactor", "Atk Factor enemy", "Gil"],
            [1, 2.5, 2.5, 0, 3, "", 2.8, 3.5, 2.5, 32],
            [2, 3.14, 3.14, 3, 3, "", 3.83, 3.5, 2.5, 90],
            [3, 3.82, 3.82, 6, 3, "", 4.91, 3.5, 2.5, 165],
            [4, 4.54, 4.54, 9, 3, "", 6.02, 3.5, 2.5, 253],
            [5, 5.3, 5.3, 12, 3, "", 7.18, 3.5, 2.5, 354],
            [6, 6.1, 6.1, 15, 3, "", 8.37, 3.5, 2.5, 465],
            [7, 6.94, 6.94, 18, 3, "", 9.6, 3.5, 2.5, 586],
            [8, 7.82, 7.82, 21, 3, "", 10.88, 3.5, 2.5, 715],
            [9, 8.74, 8.74, 24, 3, "", 12.19, 3.5, 2.5, 854],
            [10, 9.7, 9.7, 27, 3, "", 13.54, 3.5, 2.5, 1000],
            [11, 10.62, 10.62, 30, 3, "", 14.86, 3.5, 2.5, 1154],
            [12, 11.5, 11.5, 33, 3, "", 16.13, 3.5, 2.5, 1315],
            [13, 12.34, 12.34, 36, 3, "", 17.37, 3.5, 2.5, 1483],
            [14, 13.14, 13.14, 39, 3, "", 18.56, 3.5, 2.5, 1657],
            [15, 13.9, 13.9, 42, 3, "", 19.71, 3.5, 2.5, 1837],
            [16, 14.62, 14.62, 45, 3, "", 20.83, 3.5, 2.5, 2024],
            [17, 15.3, 15.3, 48, 3, "", 21.9, 3.5, 2.5, 2216],
            [18, 15.94, 15.94, 51, 3, "", 22.93, 3.5, 2.5, 2415],
            [19, 16.62, 16.62, 54, 3, "", 24.01, 3.5, 2.5, 2619],
            [20, 17.34, 17.34, 57, 3, "", 25.12, 3.5, 2.5, 2829],

        ]


class WeaponsArmor:
    def __init__(self):
        self.weapon_powers = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots", "Attributes",
             "Protection", "Level", "hp", "mp", "type", "equipped"],
            ['sand storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['wind storm', 6, 250, 1, 0, 0, 0, 0, 0, "weapon", 1],
            ['blade storm', 7, 500, 2, 0, 0, 0, 0, 0, "weapon", 1],
            ['excalibur', 9, 5000, 4, 0, 0, 5, 0, 0, "weapon", 1],
            ['the mark of the serpents', 0, 15000, 0, 0, "poison", 0, 0, 0, "ring", 1],
            ['light armor', 0, 3000, 1, 0, 0, 0, 0, 0, "armor", 1],
            ['heavy armor', 0, 7000, 2, 0, 0, 0, 5, 1.2, "armor", 1],
            ['silver ring', 0, 15000, 0, 0, "paralyzed", 0, 0, 0, "ring", 1], ]

        self.shop_buy = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 250, 1],
            ['wind storm', 6, 250, 1],
            ['blade storm', 7, 500, 2],
            ['excalibur', 9, 5000, 4],
            ['the mark of the serpents', 0, 15000, 0],
            ['light armor', 0, 3000, 1],
            ['heavy armor', 0, 7000, 2],
            ['silver ring', 0, 15000, 0], ]

        self.shop_sell = [
            ["Weapon/armor", "weapon power", "gil cost", "Magic Slots"],
            ['sand storm', 6, 125, 1],
            ['wind storm', 6, 125, 1],
            ['blade storm', 7, 250, 2],
            ['excalibur', 9, 2500, 4],
            ['the mark of the serpents', 0, 7500, 0],
            ['light armor', 0, 1500, 1],
            ['heavy armor', 0, 3500, 2],
            ['silver ring', 0, 7500, 0], ]


create_inventory = InventoryItems()
create_magic = Cast_Magic()
create_enemy = EnemyName()
create_enemy_level = LevelEnemy()
create_weapons_armor = WeaponsArmor()


class MakeExcelFiles:
    def __init__(self):
        # create pandas dataframe from lists of class
        self.df_abreheim_shop = pd.DataFrame(create_inventory.abreheim_shop)
        self.df_abreheim_shop_default = pd.DataFrame(create_inventory.abreheim_shop_defaulf)
        self.df_abreheim_shop_store_count = pd.DataFrame(create_inventory.store_count)
        self.df_buy_history = pd.DataFrame(create_inventory.buy_history)
        self.df_magic = pd.DataFrame(create_magic.magic)
        self.df_magic_gill = pd.DataFrame(create_magic.magic_gil)
        self.df_magic_sell = pd.DataFrame(create_magic.magic_sell)
        self.df_enemy_name = pd.DataFrame(create_enemy.enemy_name)
        self.df_enemy_level = pd.DataFrame(create_enemy_level.enemy_level)
        self.df_weapons_power = pd.DataFrame(create_weapons_armor.weapon_powers)
        self.df_weapons_buy = pd.DataFrame(create_weapons_armor.shop_buy)
        self.df_weapons_sell = pd.DataFrame(create_weapons_armor.shop_sell)

    def create_folders(self):
        folders = ["\\music", "\\save", "\\enemy", "\\inventory_items", "\\magic", "\\weapons_armor", "\\level"]
        # mode = 0o666
        for i in folders:
            path_folder = (CURR_DIR_PATH + i)
            obj = Path(path_folder)
            if obj.exists():
                continue
            else:
                # print(CURR_DIR_PATH + i)
                os.mkdir(CURR_DIR_PATH + i)

    def create_sub_folders_save(self):
        path_folder = (CURR_DIR_PATH + "\\save" + "\\save_temp")
        obj = Path(path_folder)
        if obj.exists():
            print()
        else:
            os.mkdir(path_folder)

    def create_main_file(self, folder_name, file_name):

        path = (CURR_DIR_PATH + folder_name + file_name)
        obj = Path(path)
        if obj.exists():
            print("")

        else:
            wb = openpyxl.Workbook()
            wb.save(path)
            wb.close()

    def write_data(self, sheet_name, folder_name, work_book_name, df_name):
        path = (CURR_DIR_PATH + folder_name + work_book_name)
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        df_name.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        if "Sheet" in book.sheetnames:
            book.remove(book["Sheet"])
        writer.save()

    def run_main(self):

        # create folders
        MakeExcelFiles.create_folders(self)
        MakeExcelFiles.create_sub_folders_save(self)
        # Creates main files
        MakeExcelFiles.create_main_file(self, "\\enemy", "\\enemy_name.xlsx")
        MakeExcelFiles.create_main_file(self, "\\inventory_items", "\\item_store.xlsx")
        MakeExcelFiles.create_main_file(self, "\\magic", "\\cast_magic.xlsx")
        MakeExcelFiles.create_main_file(self, "\\weapons_armor", "\\weapons_armor.xlsx")
        MakeExcelFiles.create_main_file(self, "\\level", "\\level_up.xlsx")

        # create excel files
        MakeExcelFiles.write_data(self, "Abreheim's items", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop)
        MakeExcelFiles.write_data(self, "Abreheim's items_default", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop_default)
        MakeExcelFiles.write_data(self, "store_count", "\\inventory_items", "\\item_store.xlsx",
                                    self.df_abreheim_shop_store_count)
        MakeExcelFiles.write_data(self, "buy_history", "\\inventory_items", "\\item_store.xlsx", self.df_buy_history)
        MakeExcelFiles.write_data(self, "magic", "\\magic", "\\cast_magic.xlsx", self.df_magic)
        MakeExcelFiles.write_data(self, "magic_gil", "\\magic", "\\cast_magic.xlsx", self.df_magic_gill)
        MakeExcelFiles.write_data(self, "magic_sell", "\\magic", "\\cast_magic.xlsx", self.df_magic_sell)
        MakeExcelFiles.write_data(self, "Ark1", "\\enemy", "\\enemy_name.xlsx", self.df_enemy_name)
        MakeExcelFiles.write_data(self, "Strength Mod enemy", "\\level", "\\level_up.xlsx", self.df_enemy_level)
        MakeExcelFiles.write_data(self, "weapon_powers", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_power)
        MakeExcelFiles.write_data(self, "shop_buy", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_buy)
        MakeExcelFiles.write_data(self, "shop_sell", "\\weapons_armor", "\\weapons_armor.xlsx", self.df_weapons_sell)

class MusicDownload:
    def __init__(self, url):
        self.url = url
        self.folder_url = "https://drive.google.com/drive/folders/15TyBVDzYyffK0SOBLBCFKgtjeaDiFq1Y"
        self.name_list = []
        self.id_list = []
        self.url_name_class = "Q5txwe"
        self.url_name_attribute = "aria-label"
        self.url_id_class = "WYuW0e"
        self.url_id_attribute = "data-id"
        self.music_name_dict = {}
        self.music_id_dict = {}
        self.folder_name_list = []
        self.x = 0

    def get_folder_name(self):
        URL = self.folder_url
        page = requests.get(URL)
        # print(page.text)
        soup = BeautifulSoup(page.content, "html.parser")
        # print(soup)
        results = soup.find(id="drive_main_page")
        # print(results.prettify())
        job_elements_name = results.find_all("div", class_=self.url_name_class)
        for job_elements_name in job_elements_name:
            # title_music = results["div aria-label"]
            # print(job_elements, end="\n"*2)
            music_name = job_elements_name[self.url_name_attribute]
            self.folder_name_list.append(music_name)

    def get_music_name_id(self):
        URL = self.url
        page = requests.get(URL)
        # print(page.text)
        soup = BeautifulSoup(page.content, "html.parser")
        # print(soup)
        results = soup.find(id="drive_main_page")
        # print(results.prettify())
        job_elements_name = results.find_all("div", class_=self.url_name_class)
        for job_elements_name in job_elements_name:
            # title_music = results["div aria-label"]
            # print(job_elements, end="\n"*2)
            music_name = job_elements_name[self.url_name_attribute]
            self.name_list.append(music_name)
        job_elements_id = results.find_all("div", class_=self.url_id_class)
        for job_elements_id in job_elements_id:
            music_id = job_elements_id[self.url_id_attribute]
            self.id_list.append(music_id)
        MusicDownload.create_folders(self)
    def download_file_from_google_drive(self, id, destination):
        URL = "https://docs.google.com/uc?export=download"

        session = requests.Session()

        response = session.get(URL, params={'id': id}, stream=True)
        token = MusicDownload.get_confirm_token(self, response)

        if token:
            params = {'id': id, 'confirm': token}
            response = session.get(URL, params=params, stream=True)

        MusicDownload.save_response_content(self, response, destination)

    def get_confirm_token(self, response):
        for key, value in response.cookies.items():
            if key.startswith('download_warning'):
                return value

        return None

    def save_response_content(self, response, destination):
        CHUNK_SIZE = 32768

        with open(destination, "wb") as f:
            for chunk in response.iter_content(CHUNK_SIZE):
                if chunk:  # filter out keep-alive new chunks
                    f.write(chunk)


    def create_folders(self):

        regex = "([^\\s]+(\\.(?i)(jpe?g|png|gif|bmp|mp3))$)"
        p = re.compile(regex)

        # print(f"self x {main_folder.x}")
        for i in self.name_list:
            if re.search(p, i):
                folder_name = self.folder_name_list[main_folder.x]
                # print(folder_name)
                self.music_name_dict = {folder_name: self.name_list}
                self.music_id_dict = {folder_name: self.id_list}
                main_folder.x += 1
                MusicDownload.download_music(self)
            else:
                for i in self.name_list:
                    rootdir = CURR_DIR_PATH + "\\music"
                    if i not in os.listdir(rootdir):
                        os.mkdir(rootdir + "\\" + i)

            try:
                get_music = MusicDownload(f"https://drive.google.com/drive/folders/{main_folder.id_list[main_folder.x]}")
                # main_folder.x += 1
                get_music.get_folder_name()
                get_music.get_music_name_id()
            except:
                print()


    def download_music(self):
        for key, value in self.music_id_dict.items():
            for x in range(0, len(value)):
                # print(x)
                folder = key
                file_id = value[x]
                # print(file_id)
                song_name = self.music_name_dict[key][x]
                # print(folder, file_id, song_name)
                print(f"Downloading music: {song_name}")
                destination = CURR_DIR_PATH + "\\music" + "\\" + folder + "\\" + song_name
                MusicDownload.download_file_from_google_drive(self, file_id, destination)
                print(f"The music {song_name} has been downloaded")

class TickTack:
    def __init__(self, board):
        self.board = board
        self.win = 0

    def print_board(self, board):
        row1 = "|{}|{}|{}|".format(board[0], board[1], board[2])
        row2 = "|{}|{}|{}|".format(board[3], board[4], board[5])
        row3 = "|{}|{}|{}|".format(board[6], board[7], board[8])
        print()
        print(row1)
        print(row2)
        print(row3)
        print()


    def player_move(self, icon, board):
        if icon == "X":
            number = 1
            print(f"Your turn player {icon}")
            choice = int(input("Enter your move (1-9)").strip())
            if board[choice - 1] == " ":
                board[choice - 1] = icon
            else:
                print("That space is taken!")
                TickTack.player_move(self, icon, self.board)
        elif icon == "O":
            number = 2
            print(f"Your turn player {icon}")
            if board[0] == icon and board[1] == icon or board[0] == "X" and board[1] == "X":
                choice = 3
                TickTack.if_taken(self, choice, self.board)
            elif board[1] == icon and board[2] == icon or board[1] == "X" and board[2] == "X":
                choice = 1
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[2] == icon or board[0] == "X" and board[2] == "X":
                choice = 4
                TickTack.if_taken(self, choice, self.board)
            elif board[3] == icon and board[4] == icon or board[3] == "X" and board[4] == "X":
                choice = 6
                TickTack.if_taken(self, choice, self.board)
            elif board[3] == icon and board[5] == icon or board[3] == "X" and board[5] == "X":
                choice = 5
                TickTack.if_taken(self, choice, self.board)
            elif board[4] == icon and board[5] == icon or board[4] == "X" and board[5] == "X":
                choice = 4
                TickTack.if_taken(self, choice, self.board)
            elif board[6] == icon and board[7] == icon or board[6] == "X" and board[7] == "X":
                choice = 9
                TickTack.if_taken(self, choice, self.board)
            elif board[6] == icon and board[8] == icon or board[6] == "X" and board[8] == "X":
                choice = 8
                TickTack.if_taken(self, choice, self.board)
            elif board[7] == icon and board[8] == icon or board[7] == "X" and board[8] == "X":
                choice = 7
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[3] == icon or board[0] == "X" and board[3] == "X":
                choice = 7
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[6] == icon or board[0] == "X" and board[6] == "X":
                choice = 4
                TickTack.if_taken(self, choice, self.board)
            elif board[3] == icon and board[6] == icon or board[4] == "X" and board[6] == "X":
                choice = 1
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[3] == icon or board[0] == "X" and board[3] == "X":
                choice = 7
                TickTack.if_taken(self, choice, self.board)
            elif board[1] == icon and board[4] == icon or board[1] == "X" and board[4] == "X":
                choice = 8
                TickTack.if_taken(self, choice, self.board)
            elif board[1] == icon and board[7] == icon or board[1] == "X" and board[7] == "X":
                choice = 5
                TickTack.if_taken(self, choice, self.board)
            elif board[4] == icon and board[7] == icon or board[4] == "X" and board[7] == "X":
                choice = 2
                TickTack.if_taken(self, choice, self.board)
            elif board[2] == icon and board[5] == icon or board[2] == "X" and board[5] == "X":
                choice = 9
                TickTack.if_taken(self, choice, self.board)
            elif board[2] == icon and board[8] == icon or board[2] == "X" and board[8] == "X":
                choice = 6
                TickTack.if_taken(self, choice, self.board)
            elif board[5] == icon and board[8] == icon or board[5] == "X" and board[8] == "X":
                choice = 3
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[4] == icon or board[0] == "X" and board[4] == "X":
                choice = 9
                TickTack.if_taken(self, choice, self.board)
            elif board[0] == icon and board[8] == icon or board[0] == "X" and board[8] == "X":
                choice = 5
                TickTack.if_taken(self, choice, self.board)
            elif board[4] == icon and board[8] == icon or board[4] == "X" and board[8] == "X":
                choice = 1
                TickTack.if_taken(self, choice, self.board)
            elif board[2] == icon and board[6] == icon or board[2] == "X" and board[6] == "X":
                choice = 5
                TickTack.if_taken(self, choice, self.board)
            elif board[2] == icon and board[4] == icon or board[2] == "X" and board[4] == "X":
                choice = 7
                TickTack.if_taken(self, choice, self.board)
            elif board[4] == icon and board[6] == icon or board[4] == "X" and board[6] == "X":
                choice = 3
                TickTack.if_taken(self, choice, self.board)
            elif board[4] == " ":
                choice = 5
            else:
                choice = random.randint(1, 9)
                TickTack.if_taken(self, choice, self.board)
            if board[choice - 1] == " ":
                board[choice - 1] = icon
            else:
                print("That space is taken!")
                board[TickTack.if_taken(self, choice, self.board) - 1] = icon
                # if_taken(choice)
                # player_move(icon)


    def if_taken(self, choice, board):
        if board[choice - 1] == "X" or board[choice - 1] == "O":
            empty_space = []
            for i in range(1, 10):
                if board[i - 1] == " ":
                    empty_space.append(i)
            try:
                choice = random.choice(empty_space)
            except:
                print()
        return choice



    def clear_board(self, player_win_count, computer_win_count, turn, player_name, board):
        score_board = {}
        print(f"Player points = {player_win_count}\nComputer points = {computer_win_count}")
        if player_win_count >= 3:
            self.win = 1
            print(f"Congratulations {player_name}, you won!!")
            score_board[player_name] = player_win_count
            score_board["Computer"] = computer_win_count
            with open('score_board.txt', 'a') as file:
                file.write(str(score_board))

        elif computer_win_count >= 3:
            self.win = 0
            print(f"Sorry {player_name}, you loose, the Computer wins!!")
            score_board[player_name] = player_win_count
            score_board["Computer"] = computer_win_count
            with open('score_board.txt', 'a') as file:
                file.write(str(score_board))

        else:
            input(f"Press enter for round {turn}!!")
            print("-"*25)
            for x in range(1, 10):
                if board[x - 1] != " ":
                    board[x - 1] = " "


    def is_victory(self, icon, board):
        if (board[0] == icon and board[1] == icon and board[2] == icon) or \
            (board[3] == icon and board[4] == icon and board[5] == icon) or \
            (board[6] == icon and board[7] == icon and board[8] == icon) or \
            (board[0] == icon and board[3] == icon and board[6] == icon) or \
            (board[1] == icon and board[4] == icon and board[7] == icon) or \
            (board[2] == icon and board[5] == icon and board[8] == icon) or \
            (board[0] == icon and board[4] == icon and board[8] == icon) or \
            (board[2] == icon and board[4] == icon and board[6] == icon):
            return True
        else:
            return False

    def is_draw(self, board):
        if " " not in board:
            return True
        else:
            return False


    def start_game(self):
        turn = 1
        player_win_count = 0
        computer_win_count = 0
        player_name = input("What is your name").lower()
        who_starts = input("Who start?").lower()
        TickTack.clear_board(self, player_win_count, computer_win_count, turn, player_name, self.board)
        if "i" in who_starts:
            print("Ok, you will start")
            player_start = "X"
            second_player = "O"
        else:
            print("Ok, I will start")
            player_start = "O"
            second_player = "X"
        while player_win_count < 3 and computer_win_count < 3:
            if player_win_count < 3 and computer_win_count < 3:
                TickTack.print_board(self, self.board)
                TickTack.player_move(self, player_start, self.board)
                TickTack.print_board(self, self.board)
                if TickTack.is_victory(self, player_start, self.board):
                    if player_start == "X":
                        player_win_count += 1
                    else:
                        computer_win_count += 1
                    turn += 1
                    print(f"{player_start} Wins! Congratulations!")
                    TickTack.clear_board(self, player_win_count, computer_win_count, turn, player_name, self.board)
                elif TickTack.is_draw(self, self.board):
                    print("Its a draw!")
                    turn += 1
                    TickTack.clear_board(self, player_win_count, computer_win_count, turn, player_name, self.board)
                TickTack.player_move(self, second_player, self.board)
                TickTack.print_board(self, self.board)
                if TickTack.is_victory(self, second_player, self.board):
                    if second_player == "X":
                        player_win_count += 1
                    else:
                        computer_win_count += 1
                    turn += 1
                    print(f"{second_player} Wins! Congratulations!")
                    TickTack.clear_board(self, player_win_count, computer_win_count, turn, player_name, self.board)
                elif TickTack.is_draw(self, self.board):
                    turn += 1
                    print("Its a draw!")
                    TickTack.clear_board(self, player_win_count, computer_win_count, turn, player_name, self.board)
            elif player_win_count > 3:
                break
            elif computer_win_count > 3:
                break

game_tick_tack = TickTack([" " for i in range(9)])
main_folder = MusicDownload("https://drive.google.com/drive/folders/15TyBVDzYyffK0SOBLBCFKgtjeaDiFq1Y")
make_excel_files = MakeExcelFiles()
important = GetImportant()
dir_path = DirectoryPaths()
opponents = ChooseEnemy()
my_timer = Timer()
story = StoryName()
code = CodeDigits()
play = Music()
cloud = YourHero()
elena = YourHero()
enemy = Enemy()
enemy_2 = Enemy()
enemy_3 = Enemy()


def x_save_hero():
    my_timer.pause_time()
    my_timer.print_time()
    date = datetime.strftime(datetime.now(), ' %Y_%m_%d_%H_%M_%S')
    file_name = elena.name
    cloud.x_initials_stats()
    elena.x_initials_stats()
    Enemy.ex_initials_stats(enemy)
    Enemy.ex_initials_stats(enemy_2)
    Enemy.ex_initials_stats(enemy_3)
    while True:
        x_save = 1
        if x_save == 1:
            user_input = input("What wo you want to do, new save or save over file? \n")
            print("---------------------------------------------------------")
            if user_input == "save over file":
                temp_folder = "save_temp" + "\\\\"
                temp_save = dir_path.save_path + temp_folder
                # print(temp_save)
                save_file = shelve.open(temp_save + file_name + date)
                # Elena
                save_file['hero.exp_show'] = elena.exp_show
                save_file['hero.exp'] = elena.exp
                save_file['hero.name'] = elena.name
                save_file['hero.level'] = elena.level
                save_file['hero.hp'] = elena.hp
                save_file['hero.mp'] = elena.mp
                save_file['hero.bonus_factor'] = elena.bonus_factor
                save_file['hero.factor_level_up'] = elena.factor_level_up
                # save_file['hero.story_number'] = elena.chapter
                save_file['hero.story_name'] = elena.story_name
                save_file['hero.chapter'] = elena.chapter
                save_file['hero.key_item_1'] = elena.key_item_1
                save_file['hero.key_item_2'] = elena.key_item_2
                save_file['hero.key_item_3'] = elena.key_item_3
                save_file['hero.key_item_4'] = elena.key_item_4
                save_file['hero.key_item_name_1'] = elena.key_item_1_name
                save_file['hero.key_item_name_2'] = elena.key_item_2_name
                save_file['hero.key_item_name_3'] = elena.key_item_3_name
                save_file['hero.key_item_name_4'] = elena.key_item_4_name
                save_file['hero_progress'] = elena.progress
                save_file['hero.magic_spells'] = elena.magic_spells
                save_file['hero.magic_slots'] = elena.magic_slots
                save_file['hero.magic_slots_max'] = elena.magic_slots_max
                save_file['hero.armory_inventory'] = elena.armory_inventory
                save_file['hero.inventory'] = elena.inventory

                # Cloud
                save_file['hero2.exp_show'] = cloud.exp_show
                save_file['hero2.exp'] = cloud.exp
                save_file['hero2.name'] = cloud.name
                save_file['hero2.level'] = cloud.level
                save_file['hero2.hp'] = cloud.hp
                save_file['hero2.mp'] = cloud.mp
                save_file['hero2.bonus_factor'] = cloud.bonus_factor
                save_file['hero2.factor_level_up'] = cloud.factor_level_up
                # save_file['hero2.story_number'] = cloud.story_number
                save_file['hero2.story_name'] = cloud.story_name
                save_file['hero2.chapter'] = cloud.chapter
                save_file['hero2.key_item_1'] = cloud.key_item_1
                save_file['hero2.key_item_2'] = cloud.key_item_2
                save_file['hero2.key_item_3'] = cloud.key_item_3
                save_file['hero2.key_item_4'] = cloud.key_item_4
                save_file['hero2.key_item_name_1'] = cloud.key_item_1_name
                save_file['hero2.key_item_name_2'] = cloud.key_item_2_name
                save_file['hero2.key_item_name_3'] = cloud.key_item_3_name
                save_file['hero2.key_item_name_4'] = cloud.key_item_4_name
                save_file['hero2.chapter_1_round_room_count'] = cloud.chapter_1_round_room_count
                save_file['hero2.chapter_1_ally_woman_wall_away'] = cloud.chapter_1_ally_woman_wall_away
                save_file['hero2.chapter_1_ally_woman_g_hat_away'] = cloud.chapter_1_ally_woman_green_hat_away
                save_file['hero2.sleep_abreheim_inn'] = cloud.sleep_abreheim_inn
                save_file['hero2.inn_room_1'] = cloud.inn_room_1
                save_file['hero2.inn_room_2'] = cloud.inn_room_2
                save_file['hero2.food_1'] = cloud.food_1
                save_file['hero2.food_2'] = cloud.food_2
                save_file['hero2.magic_spells'] = cloud.magic_spells
                save_file['hero2.magic_slots'] = cloud.magic_slots
                save_file['hero2.magic_slots_max'] = cloud.magic_slots_max
                save_file['hero2_progress'] = cloud.progress
                save_file['hero2.armory_inventory'] = cloud.armory_inventory
                save_file['hero2.inventory'] = cloud.inventory

                # Enemies
                save_file['enemy.exp'] = enemy.e_exp
                save_file['enemy.level'] = enemy.e_level
                save_file['enemy.hp'] = enemy.e_hp
                save_file['enemy.factor_level_up'] = enemy.e_factor_for_level_up
                save_file['enemy.bonus_factor'] = enemy.e_bonus_factor
                save_file['enemy2.exp'] = enemy_2.e_exp
                save_file['enemy2.level'] = enemy_2.e_level
                save_file['enemy2.hp'] = enemy_2.e_hp
                save_file['enemy2.factor_level_up'] = enemy_2.e_factor_for_level_up
                save_file['enemy2.bonus_factor'] = enemy_2.e_bonus_factor
                save_file['enemy3.exp'] = enemy_3.e_exp
                save_file['enemy3.level'] = enemy_3.e_level
                save_file['enemy3.hp'] = enemy_3.e_hp
                save_file['enemy3.factor_level_up'] = enemy_3.e_factor_for_level_up
                save_file['enemy3.bonus_factor'] = enemy_3.e_bonus_factor

                # Time
                save_file['time_start'] = my_timer.start
                save_file['time_pause'] = my_timer.pause
                save_file.close()
                x = 0
                path = dir_path.save_path
                extension = 'dat'
                os.chdir(path)
                files = glob.glob('*.{}'.format(extension))
                count_files = len(files)
                file_range = range(count_files + 1)
                for file in files:
                    x += 1
                    i = x
                    save_choice = i
                    your_choice = files[int(save_choice) - 1]
                    load_choice = your_choice.removesuffix('.dat')
                    # print(load_choice)
                    x_load = shelve.open(dir_path.save_path + load_choice)
                    elena.exp_show = x_load['hero.exp_show']
                    elena.level = x_load['hero.level']
                    elena.name = x_load['hero.name']
                    elena.hp = x_load['hero.hp']
                    elena.mp = x_load['hero.mp']
                    elena.chapter = x_load['hero.chapter']
                    # elena.chapter = save_file['hero.story_number']
                    elena.story_name = x_load['hero.story_name']
                    cloud.exp_show = x_load['hero2.exp_show']
                    cloud.level = x_load['hero2.level']
                    cloud.name = x_load['hero2.name']
                    cloud.hp = x_load['hero2.hp']
                    cloud.mp = x_load['hero2.mp']
                    # cloud.chapter = save_file['hero2.story_number']
                    cloud.story_name = x_load['hero2.story_name']
                    cloud.chapter = x_load['hero2.chapter']
                    # Timer
                    my_timer.start = x_load['time_start']
                    my_timer.pause = x_load['time_pause']
                    my_timer.restart()
                    my_timer.print_time()
                    YourHero.x_initials_stats(elena)
                    YourHero.x_initials_stats(cloud)
                    print(str(i) + " " + file)
                    print(YourHero.x_stat_load(cloud and elena))
                    x_load.close()
                    print("---------------------------------------------------------")
                save_choice = int(input("Choose a file to save over: "))
                if save_choice in file_range and save_choice != 0:
                    your_choice = files[int(save_choice) - 1]
                    load_choice = your_choice.removesuffix('.dat')

                    os.remove(dir_path.save_path + load_choice + ".dat")
                    os.remove(dir_path.save_path + load_choice + ".dir")
                    os.remove(dir_path.save_path + load_choice + ".bak")

                    save_file = shelve.open(temp_save + file_name + date)
                    # Elena
                    elena.exp_show = save_file['hero.exp_show']
                    elena.exp = save_file['hero.exp']
                    elena.name = save_file['hero.name']
                    elena.level = save_file['hero.level']
                    elena.hp = save_file['hero.hp']
                    elena.mp = save_file['hero.mp']
                    elena.bonus_factor = save_file['hero.bonus_factor']
                    elena.factor_level_up = save_file['hero.factor_level_up']
                    # elena.chapter = save_file['hero.story_number']
                    elena.story_name = save_file['hero.story_name']
                    elena.key_item_1 = save_file['hero.key_item_1']
                    elena.key_item_2 = save_file['hero.key_item_2']
                    elena.key_item_3 = save_file['hero.key_item_3']
                    elena.key_item_4 = save_file['hero.key_item_4']
                    elena.key_item_1_name = save_file['hero.key_item_name_1']
                    elena.key_item_2_name = save_file['hero.key_item_name_2']
                    elena.key_item_3_name = save_file['hero.key_item_name_3']
                    elena.key_item_4_name = save_file['hero.key_item_name_4']
                    elena.chapter = save_file['hero.chapter']
                    elena.progress = save_file['hero_progress']
                    elena.magic_spells = save_file['hero.magic_spells']
                    elena.magic_slots = save_file['hero.magic_slots']
                    elena.magic_slots_max = save_file['hero.magic_slots_max']
                    elena.progress = save_file['hero_progress']
                    elena.armory_inventory = save_file['hero.armory_inventory']
                    elena.inventory = save_file['hero.inventory']

                    # Cloud
                    cloud.exp_show = save_file['hero2.exp_show']
                    cloud.exp = save_file['hero2.exp']
                    cloud.name = save_file['hero2.name']
                    cloud.level = save_file['hero2.level']
                    cloud.hp = save_file['hero2.hp']
                    cloud.mp = save_file['hero2.mp']
                    cloud.bonus_factor = save_file['hero2.bonus_factor']
                    cloud.factor_level_up = save_file['hero2.factor_level_up']
                    # cloud.story_number = save_file['hero2.story_number']
                    cloud.story_name = save_file['hero2.story_name']
                    cloud.chapter = save_file['hero2.chapter']
                    cloud.key_item_1 = save_file['hero2.key_item_1']
                    cloud.key_item_2 = save_file['hero2.key_item_2']
                    cloud.key_item_3 = save_file['hero2.key_item_3']
                    cloud.key_item_4 = save_file['hero2.key_item_4']
                    cloud.key_item_1_name = save_file['hero2.key_item_name_1']
                    cloud.key_item_2_name = save_file['hero2.key_item_name_2']
                    cloud.key_item_3_name = save_file['hero2.key_item_name_3']
                    cloud.key_item_4_name = save_file['hero2.key_item_name_4']
                    cloud.chapter_1_round_room_count = save_file['hero2.chapter_1_round_room_count']
                    cloud.chapter_1_ally_woman_wall_away = save_file['hero2.chapter_1_ally_woman_wall_away']
                    cloud.chapter_1_ally_woman_green_hat_away = save_file['hero2.chapter_1_ally_woman_g_hat_away']
                    cloud.sleep_abreheim_inn = save_file['hero2.sleep_abreheim_inn']
                    cloud.inn_room_1 = save_file['hero2.inn_room_1']
                    cloud.inn_room_2 = save_file['hero2.inn_room_2']
                    cloud.food_1 = save_file['hero2.food_1']
                    cloud.food_2 = save_file['hero2.food_2']
                    cloud.magic_spells = save_file['hero2.magic_spells']
                    cloud.magic_slots = save_file['hero2.magic_slots']
                    cloud.magic_slots_max = save_file['hero2.magic_slots_max']
                    cloud.progress = save_file['hero2_progress']
                    cloud.armory_inventory = save_file['hero2.armory_inventory']
                    cloud.inventory = save_file['hero2.inventory']

                    # Enemies
                    enemy.e_exp = save_file['enemy.exp']
                    enemy.e_level = save_file['enemy.level']
                    enemy.e_hp = save_file['enemy.hp']
                    enemy.e_factor_for_level_up = save_file['enemy.factor_level_up']
                    enemy.e_bonus_factor = save_file['enemy.bonus_factor']
                    enemy_2.e_exp = save_file['enemy2.exp']
                    enemy_2.e_level = save_file['enemy2.level']
                    enemy_2.e_hp = save_file['enemy2.hp']
                    enemy_2.e_factor_for_level_up = save_file['enemy2.factor_level_up']
                    enemy_2.e_bonus_factor = save_file['enemy2.bonus_factor']
                    enemy_3.e_exp = save_file['enemy3.exp']
                    enemy_3.e_level = save_file['enemy3.level']
                    enemy_3.e_hp = save_file['enemy3.hp']
                    enemy_3.e_factor_for_level_up = save_file['enemy3.factor_level_up']
                    enemy_3.e_bonus_factor = save_file['enemy3.bonus_factor']

                    # Time
                    my_timer.start = save_file['time_start']
                    my_timer.pause = save_file['time_pause']

                    source_dir = temp_save
                    target_dir = dir_path.save_path
                    file_names = os.listdir(source_dir)
                    for file_name in file_names:
                        shutil.move(os.path.join(source_dir, file_name), target_dir)

                    save_text = "Saving...\n"
                    now = time.time()
                    future = now + 3
                    while time.time() < future:
                        for char in save_text:
                            sys.stdout.write(char)
                            sys.stdout.flush()
                            time.sleep(time_save)

                    input("\nGame saved: \nPress enter")
                    my_timer.restart()
                    break
            elif user_input.lower() == "new save":
                save_file = shelve.open(dir_path.save_path + file_name + date)
                # Elena
                save_file['hero.exp_show'] = elena.exp_show
                save_file['hero.exp'] = elena.exp
                save_file['hero.name'] = elena.name
                save_file['hero.level'] = elena.level
                save_file['hero.hp'] = elena.hp
                save_file['hero.mp'] = elena.mp
                save_file['hero.bonus_factor'] = elena.bonus_factor
                save_file['hero.factor_level_up'] = elena.factor_level_up
                # save_file['hero.story_number'] = elena.chapter
                save_file['hero.story_name'] = elena.story_name
                save_file['hero.chapter'] = elena.chapter
                save_file['hero.key_item_1'] = elena.key_item_1
                save_file['hero.key_item_2'] = elena.key_item_2
                save_file['hero.key_item_3'] = elena.key_item_3
                save_file['hero.key_item_4'] = elena.key_item_4
                save_file['hero.key_item_name_1'] = elena.key_item_1_name
                save_file['hero.key_item_name_2'] = elena.key_item_2_name
                save_file['hero.key_item_name_3'] = elena.key_item_3_name
                save_file['hero.key_item_name_4'] = elena.key_item_4_name
                save_file['hero_progress'] = elena.progress
                save_file['hero.magic_spells'] = elena.magic_spells
                save_file['hero.magic_slots'] = elena.magic_slots
                save_file['hero.magic_slots_max'] = elena.magic_slots_max
                save_file['hero.armory_inventory'] = elena.armory_inventory
                save_file['hero.inventory'] = elena.inventory

                # Cloud
                save_file['hero2.exp_show'] = cloud.exp_show
                save_file['hero2.exp'] = cloud.exp
                save_file['hero2.name'] = cloud.name
                save_file['hero2.level'] = cloud.level
                save_file['hero2.hp'] = cloud.hp
                save_file['hero2.mp'] = cloud.mp
                save_file['hero2.bonus_factor'] = cloud.bonus_factor
                save_file['hero2.factor_level_up'] = cloud.factor_level_up
                # save_file['hero2.story_number'] = cloud.story_number
                save_file['hero2.story_name'] = cloud.story_name
                save_file['hero2.chapter'] = cloud.chapter
                save_file['hero2.key_item_1'] = cloud.key_item_1
                save_file['hero2.key_item_2'] = cloud.key_item_2
                save_file['hero2.key_item_3'] = cloud.key_item_3
                save_file['hero2.key_item_4'] = cloud.key_item_4
                save_file['hero2.key_item_name_1'] = cloud.key_item_1_name
                save_file['hero2.key_item_name_2'] = cloud.key_item_2_name
                save_file['hero2.key_item_name_3'] = cloud.key_item_3_name
                save_file['hero2.key_item_name_4'] = cloud.key_item_4_name
                save_file['hero2.chapter_1_round_room_count'] = cloud.chapter_1_round_room_count
                save_file['hero2.chapter_1_ally_woman_wall_away'] = cloud.chapter_1_ally_woman_wall_away
                save_file['hero2.chapter_1_ally_woman_g_hat_away'] = cloud.chapter_1_ally_woman_green_hat_away
                save_file['hero2.sleep_abreheim_inn'] = cloud.sleep_abreheim_inn
                save_file['hero2.inn_room_1'] = cloud.inn_room_1
                save_file['hero2.inn_room_2'] = cloud.inn_room_2
                save_file['hero2.food_1'] = cloud.food_1
                save_file['hero2.food_2'] = cloud.food_2
                save_file['hero2.magic_spells'] = cloud.magic_spells
                save_file['hero2.magic_slots'] = cloud.magic_slots
                save_file['hero2_progress'] = cloud.progress
                save_file['hero2.magic_slots_max'] = cloud.magic_slots_max
                save_file['hero2.armory_inventory'] = cloud.armory_inventory
                save_file['hero2.inventory'] = cloud.inventory

                # Enemies
                save_file['enemy.exp'] = enemy.e_exp
                save_file['enemy.level'] = enemy.e_level
                save_file['enemy.hp'] = enemy.e_hp
                save_file['enemy.factor_level_up'] = enemy.e_factor_for_level_up
                save_file['enemy.bonus_factor'] = enemy.e_bonus_factor
                save_file['enemy2.exp'] = enemy_2.e_exp
                save_file['enemy2.level'] = enemy_2.e_level
                save_file['enemy2.hp'] = enemy_2.e_hp
                save_file['enemy2.factor_level_up'] = enemy_2.e_factor_for_level_up
                save_file['enemy2.bonus_factor'] = enemy_2.e_bonus_factor
                save_file['enemy3.exp'] = enemy_3.e_exp
                save_file['enemy3.level'] = enemy_3.e_level
                save_file['enemy3.hp'] = enemy_3.e_hp
                save_file['enemy3.factor_level_up'] = enemy_3.e_factor_for_level_up
                save_file['enemy3.bonus_factor'] = enemy_3.e_bonus_factor

                # Time
                save_file['time_start'] = my_timer.start
                save_file['time_pause'] = my_timer.pause

                save_text = "Saving...\n"
                now = time.time()
                future = now + 3
                while time.time() < future:
                    for char in save_text:
                        sys.stdout.write(char)
                        sys.stdout.flush()
                        time.sleep(time_save)
                save_file.close()
                input("\nGame saved: \nPress enter")
                my_timer.restart()
                break
            else:
                print("Choose a valid option")


def x_load_hero():
    x_load = 1
    while True:
        while x_load == 1:
            if x_load == 1:
                user_input = input("What wo you want to do, load, go back to main game? \n")
                if user_input == "go back to main game" or user_input == "go back":
                    main_game()
                elif user_input == "load":
                    x = 0
                    path = dir_path.save_path
                    extension = 'dat'
                    os.chdir(path)
                    files = glob.glob('*.{}'.format(extension))
                    count_files = len(files)
                    file_range = range(count_files + 1)
                    for file in files:
                        x += 1
                        i = x
                        save_choice = i
                        your_choice = files[int(save_choice) - 1]
                        load_choice = your_choice.removesuffix('.dat')
                        # print(load_choice)
                        save_file = shelve.open(dir_path.save_path + load_choice)
                        elena.level = save_file['hero.level']
                        elena.name = save_file['hero.name']
                        elena.hp = save_file['hero.hp']
                        elena.mp = save_file['hero.mp']
                        elena.chapter = save_file['hero.chapter']
                        # elena.chapter = save_file['hero.story_number']
                        elena.story_name = save_file['hero.story_name']
                        elena.inventory = save_file['hero.inventory']
                        elena.exp_show = save_file['hero.exp_show']
                        cloud.level = save_file['hero2.level']
                        cloud.name = save_file['hero2.name']
                        cloud.hp = save_file['hero2.hp']
                        cloud.mp = save_file['hero2.mp']
                        # cloud.chapter = save_file['hero2.story_number']
                        cloud.story_name = save_file['hero2.story_name']
                        cloud.chapter = save_file['hero2.chapter']
                        cloud.inventory = save_file['hero2.inventory']
                        cloud.exp_show = save_file['hero2.exp_show']
                        # Timer
                        my_timer.start = save_file['time_start']
                        my_timer.pause = save_file['time_pause']
                        my_timer.restart()
                        my_timer.print_time()
                        YourHero.x_initials_stats(elena)
                        YourHero.x_initials_stats(cloud)
                        print(str(i) + " " + file)
                        # print(elena.load_stat)
                        print(YourHero.x_stat_load(cloud and elena))

                        save_file.close()
                        print("---------------------------------------------------------")
                    save_choice = int(input("Choose a file to load from: "))
                    # if user_input.lower() == "quit" or user_input == "go back" or user_input == "main game":

                    if save_choice in file_range and save_choice != 0:
                        your_choice = files[int(save_choice) - 1]
                        load_choice = your_choice.removesuffix('.dat')
                        save_file = shelve.open(dir_path.save_path + load_choice)
                        # Elena
                        elena.exp_show = save_file['hero.exp_show']
                        elena.exp = save_file['hero.exp']
                        elena.name = save_file['hero.name']
                        elena.level = save_file['hero.level']
                        elena.hp = save_file['hero.hp']
                        elena.mp = save_file['hero.mp']
                        elena.bonus_factor = save_file['hero.bonus_factor']
                        elena.factor_level_up = save_file['hero.factor_level_up']
                        # elena.chapter = save_file['hero.story_number']
                        elena.story_name = save_file['hero.story_name']
                        elena.key_item_1 = save_file['hero.key_item_1']
                        elena.key_item_2 = save_file['hero.key_item_2']
                        elena.key_item_3 = save_file['hero.key_item_3']
                        elena.key_item_4 = save_file['hero.key_item_4']
                        elena.key_item_1_name = save_file['hero.key_item_name_1']
                        elena.key_item_2_name = save_file['hero.key_item_name_2']
                        elena.key_item_3_name = save_file['hero.key_item_name_3']
                        elena.key_item_4_name = save_file['hero.key_item_name_4']
                        elena.chapter = save_file['hero.chapter']
                        elena.magic_spells = save_file['hero.magic_spells']
                        elena.progress = save_file['hero_progress']
                        elena.magic_slots = save_file['hero.magic_slots']
                        elena.magic_slots_max = save_file['hero.magic_slots_max']
                        elena.armory_inventory = save_file['hero.armory_inventory']
                        elena.inventory = save_file['hero.inventory']

                        # Cloud
                        cloud.exp_show = save_file['hero2.exp_show']
                        cloud.progress = save_file['hero2_progress']
                        cloud.exp = save_file['hero2.exp']
                        cloud.name = save_file['hero2.name']
                        cloud.level = save_file['hero2.level']
                        cloud.hp = save_file['hero2.hp']
                        cloud.mp = save_file['hero2.mp']
                        cloud.bonus_factor = save_file['hero2.bonus_factor']
                        cloud.factor_level_up = save_file['hero2.factor_level_up']
                        # cloud.story_number = save_file['hero2.story_number']
                        cloud.story_name = save_file['hero2.story_name']
                        cloud.chapter = save_file['hero2.chapter']
                        cloud.key_item_1 = save_file['hero2.key_item_1']
                        cloud.key_item_2 = save_file['hero2.key_item_2']
                        cloud.key_item_3 = save_file['hero2.key_item_3']
                        cloud.key_item_4 = save_file['hero2.key_item_4']
                        cloud.key_item_1_name = save_file['hero2.key_item_name_1']
                        cloud.key_item_2_name = save_file['hero2.key_item_name_2']
                        cloud.key_item_3_name = save_file['hero2.key_item_name_3']
                        cloud.key_item_4_name = save_file['hero2.key_item_name_4']
                        cloud.chapter_1_round_room_count = save_file['hero2.chapter_1_round_room_count']
                        cloud.chapter_1_ally_woman_wall_away = save_file['hero2.chapter_1_ally_woman_wall_away']
                        cloud.chapter_1_ally_woman_green_hat_away = save_file['hero2.chapter_1_ally_woman_g_hat_away']
                        cloud.sleep_abreheim_inn = save_file['hero2.sleep_abreheim_inn']
                        cloud.inn_room_1 = save_file['hero2.inn_room_1']
                        cloud.inn_room_2 = save_file['hero2.inn_room_2']
                        cloud.food_1 = save_file['hero2.food_1']
                        cloud.food_2 = save_file['hero2.food_2']
                        cloud.magic_spells = save_file['hero2.magic_spells']
                        cloud.magic_slots = save_file['hero2.magic_slots']
                        cloud.magic_slots_max = save_file['hero2.magic_slots_max']
                        cloud.progress = save_file['hero2_progress']
                        cloud.armory_inventory = save_file['hero2.armory_inventory']
                        cloud.inventory = save_file['hero2.inventory']

                        # Enemies
                        enemy.e_exp = save_file['enemy.exp']
                        enemy.e_level = save_file['enemy.level']
                        enemy.e_hp = save_file['enemy.hp']
                        enemy.e_factor_for_level_up = save_file['enemy.factor_level_up']
                        enemy.e_bonus_factor = save_file['enemy.bonus_factor']
                        enemy_2.e_exp = save_file['enemy2.exp']
                        enemy_2.e_level = save_file['enemy2.level']
                        enemy_2.e_hp = save_file['enemy2.hp']
                        enemy_2.e_factor_for_level_up = save_file['enemy2.factor_level_up']
                        enemy_2.e_bonus_factor = save_file['enemy2.bonus_factor']
                        enemy_3.e_exp = save_file['enemy3.exp']
                        enemy_3.e_level = save_file['enemy3.level']
                        enemy_3.e_hp = save_file['enemy3.hp']
                        enemy_3.e_factor_for_level_up = save_file['enemy3.factor_level_up']
                        enemy_3.e_bonus_factor = save_file['enemy3.bonus_factor']

                        # Time
                        my_timer.start = save_file['time_start']
                        my_timer.pause = save_file['time_pause']

                        print("\rLoading.....", end="")
                        time.sleep(time_load)
                        print("\rtips: use antidote to cure from poison", end="")
                        time.sleep(time_load)
                        print("\rLoading.....", end="")
                        time.sleep(time_load)
                        print("\rtips: use silver_dust to cure from paralyzed", end="")
                        time.sleep(time_load)
                        print("\rLoad complete\n", end="")
                        time.sleep(time_load)
                        save_file.close()
                        my_timer.restart()
                        my_timer.print_time()
                        YourHero.x_initials_stats(elena)
                        YourHero.x_initials_stats(cloud)
                        Enemy.ex_initials_stats(enemy)
                        Enemy.ex_initials_stats(enemy_2)
                        Enemy.ex_initials_stats(enemy_3)
                        print("---------------------------------------------------------")
                        print(cloud.stat)
                        print("---------------------------------------------------------")
                        print(elena.stat)
                        input("Press enter to continue")
                        print(elena.item_list)
                        print("---------------------------------------------------------")
                        print(cloud.item_list)
                        print("---------------------------------------------------------")
                        input("Press enter to continue")
                        print(elena.total_stat)
                        print("---------------------------------------------------------")
                        print(cloud.total_stat)
                        print("---------------------------------------------------------")
                        input("Press enter to continue\n")
                        x_load = 0
                        YourHero.x_story_chapter_count(elena)
                    else:
                        print("Please choose a valid number")
                        x_load_hero()
                elif user_input.lower() == "quit" or user_input == "back" or user_input == "main game":
                    main_game()
                else:
                    print("What do you want?")
                    print("---------------------------------------------------------")


def ask_save():
    while True:
        print("---------------------------------------------------------")
        save_game = input("Do you want to save?\n 1. Yes\n 2. No\n").lower()
        if save_game == "yes" or save_game == "y" or save_game == "1":
            tent_use = input("Do you want to use a tent?")
            if tent_use == "yes":
                item_list_cloud = pd.DataFrame(cloud.inventory)
                item_list_elena = pd.DataFrame(elena.inventory)
                item_use = "tent"
                if item_use in item_list_cloud.values or item_use in item_list_elena.values:
                    item_index_cloud = item_list_cloud[item_list_cloud["Item"] == item_use].index.values
                    item_index_elena = item_list_elena[item_list_elena["Item"] == item_use].index.values
                    res_cloud = str(item_index_cloud)[1:-1]
                    res_elena = str(item_index_elena)[1:-1]
                    if cloud.inventory["QTY"][int(res_cloud)] > 0 or elena.inventory["QTY"][int(res_elena)] > 0:
                        print("---------------------------------------------------------")
                        print(cloud.name, "Inventory")
                        print(item_list_cloud)
                        print("---------------------------------------------------------")
                        print(elena.name, "Inventory")
                        print(item_list_elena)
                        print("---------------------------------------------------------")
                        print("Whom is going to use it", cloud.name, "or", elena.name)
                        player_select = input()
                        if player_select == cloud.name:
                            if cloud.inventory["QTY"][int(res_cloud)] <= 0:
                                print("Choose another player")
                                print("---------------------------------------------------------")
                            elif cloud.inventory["QTY"][int(res_cloud)] > 0:
                                cloud.x_remove_items_inventory({"Item": ["tent"], "QTY": [1]}, "no", cloud, "no")
                                cloud.x_tent_use()
                                elena.x_tent_use()
                                x_save_hero()
                                break
                        elif player_select == elena.name:
                            if elena.inventory["QTY"][int(res_elena)] <= 0:
                                print("Choose another player")
                                print("---------------------------------------------------------")
                            elif elena.inventory["QTY"][int(res_elena)] > 0:
                                elena.x_remove_items_inventory({"Item": ["tent"], "QTY": [1]}, "no", elena, "no")
                                cloud.x_tent_use()
                                elena.x_tent_use()
                                x_save_hero()
                                break
                else:
                    print("You don't have any tent, saving the game")
                    x_save_hero()
                    break
            else:
                x_save_hero()
                break
        elif save_game == "no" or save_game == "n" or save_game == "2":
            print("---------------------------------------------------------")
            print(" A dumb choice to not save!")
            print("---------------------------------------------------------")
            break
        else:
            print("---------------------------------------------------------")
            print("Choose a valid choice!")
            print("---------------------------------------------------------")


def game_help():
    while True:
        x_help = 1
        while x_help == 1:
            if x_help == 1:
                print("---------------------------------------------------------")
                print("In this game you will be in in battles, solving puzzles, entering stores, or sleep at an in")
                print(" 'if you wish to go to main game window just write, quit, back or main game' ")
                print("Check out the writing commands, write battles! or solving puzzles")
                print("---------------------------------------------------------")
                user_input = input("What's on your mind, what do you need help with?")
                if user_input.lower() == "quit" or user_input == "back" or user_input == "main game":
                    main_game()
                elif user_input.lower() == "battles" or user_input.lower() == "battle":
                    print("---------------------------------------------------------")
                    tprint("Battle system")
                    print("When entering a battle you will have a choice to see your stats or enter the battle")
                    print("Once in battle you will have a few options like, attack, magic, and so on.\n"
                          "These are for the moment only chosen by entering digits")
                    print("The battle is a turn based system, "
                          "you attack then the enemy attack then you again and so on")
                    print("---------------------------------------------------------")
                    input("press enter")
                elif user_input.lower() == "solving puzzles" or user_input.lower() == "puzzle" \
                        or user_input.lower() == "puzzle":
                    print("---------------------------------------------------------")
                    tprint("Puzzle system")
                    print("When solving puzzles, think outside the box")
                    print("You can navigate in the puzzle, with different commands, like: ")
                    print("Walk over to the....Look at the....Open the...and so on")
                    print("There will also be a command for 'help', write it out and you will receive a clue")
                    print("Be aware though use it to much and something will happened!!")
                    print("---------------------------------------------------------")
                    input("Press enter")
                elif user_input.lower() == "entering stores" or user_input.lower() == "store" \
                        or user_input.lower() == "stores":
                    print("---------------------------------------------------------")
                    tprint("Store system")
                    print("Stores are simple, pick an item buy it with the coins you earn from battles,\n or find"
                          "in the game")
                    print("---------------------------------------------------------")
                    input("press enter")
                elif user_input.lower() == "sleep at an in" or user_input.lower() == "in" \
                        or user_input.lower() == "sleep at in":
                    print("---------------------------------------------------------")
                    tprint("The sleeping system")
                    print("Sleeping will give you full HP and MP and cure from any status effect\n"
                          "The price for a night will differ depending on where you stay")
                    print("If you have a tent then you can sleep for free out in the wild!!")
                    print("---------------------------------------------------------")
                    input("press enter")
                elif user_input.lower() == "writing commands":
                    print("For now there is a hidden command you can use in open world, towns, ally and so on")
                    print("You can not use it during battles and puzzle solving")
                    print("The command is: inventory")
                    print("The inventory will give you access to everything you have, and you can give items\n"
                          " and Gil, with each-other")
                else:
                    print("Don't really understand what you need help with")


def main_game():
    play.music_loop(play.music_main[1])
    print('''


          ''')
    print("---------------------------------------------------------")
    tprint("lost  shadows")
    print("---------------------------------------------------------")
    dots = ":     "
    name_dot = "Welcome to Lost Shadows" + dots

    start_up = {name_dot: [""],
                "#": ["1", "2", "3", "4"],
                "Options": ["Start", "Load", "Help", "Exit"]
                }
    start = pd.DataFrame(start_up, index=["", "", "", ""])
    print("---------------------------------------------------------")
    print(start)
    print("---------------------------------------------------------")
    game_starter = input("What is your choice? ").lower()
    print("---------------------------------------------------------")
    if game_starter == "1" or game_starter == "s" or game_starter == "start":
        intro()
    elif game_starter == "2" or game_starter == "l" or game_starter == "load":
        x_load_hero()
    elif game_starter == "3" or game_starter == "h" or game_starter == "help":
        game_help()
    elif game_starter == "4" or game_starter == "e" or game_starter == "exit":
        print("---------------------------------------------------------")
        print("Goodbye!")
        print("---------------------------------------------------------")
        time.sleep(2)
        sys.exit()
    else:
        main_game()


def story1_round_room():
    while True:
        elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[6])
        play.music_loop(play.sound_chapter_1[4])
        room = 1
        room_time = 0
        # second = ""
        while room == 1:
            if room_time < 25:
                room_time += 1
                cloud.x_initials_stats()
                elena.x_initials_stats()
                print("---------------------------------------------------------")
                print("You find yourself in a mysterious room, with dark-green walls"
                      " and paintings of a dystopic landscape.")
                time.sleep(time_short_wait)
                print("There is a table in the middle of the room, in dark-mahogany.")
                time.sleep(time_short_wait)
                print("And a door is located in the end of the room.")
                time.sleep(time_short_wait)
                print("There is also a statue holding an axe in its hand")
                time.sleep(time_short_wait)
                print("A big chandelier is hanging from the ceiling,"
                      " the candles are burning with a bright warm light")
                second = input("What do you do? ")
                if "table" in second:
                    print("---------------------------------------------------------")
                    print("There is only two boxes in different sizes placed on the table, "
                          "one with a form of a rectangle, one as a triangle.")
                elif second.lower() == "take the rectangle box":
                    print("---------------------------------------------------------")
                    print("Its well accord to the table, cant be removed.")
                elif "paintings" in second:
                    print("---------------------------------------------------------")
                    print(
                        "There is two different paintings, one with a hero standing and holding a flag."
                        " Anther one with"
                        " an eagle flying away from the polluted planet")
                elif "look at the paintings" in second:
                    print("---------------------------------------------------------")
                    print(
                        "There is two painting to look at, one with a hero and one with an eagle")
                elif "hero" in second:
                    print("---------------------------------------------------------")
                    print(
                        "The lonely hero is standing over a pile of mud and dirt, holding the flag, it looks like"
                        " the flag is moving")
                elif "flag" in second:
                    print("---------------------------------------------------------")
                    print("The flag is gray of dust and its impossible to see what it looks like.")
                elif second.lower() == "dust of the flag" or second.lower() == "clean the flag":
                    print("---------------------------------------------------------")
                    print("As you slowly remove the dust from the flag you can see a number that has been written"
                          " upon it, the number is", code.code_4_digits_box_1)
                elif "eagle" in second:
                    print("---------------------------------------------------------")
                    print(
                        "The eagle has its lungs are full of dirt, it tries to fly away from our planet\n"
                        "Desperate to find a new place to live on, its left and right wings are strangely hurt")
                elif "left wing" in second:
                    print("---------------------------------------------------------")
                    print("As you move closer to look at the eagles left wing,"
                          " its missing", code.code_1_digits_box_1, " or is it ", code.code_1_digits_box_2, " fetters")
                    time.sleep(time_short_wait)
                    print("you hear a sound, a sound that could make anyone shiver")
                    input("press enter")
                    YourHero.x_battle(cloud, 0, cloud, elena, "no")
                elif "right wing" in second:
                    print("---------------------------------------------------------")
                    print("The right wing is missing som fetters, is it ", code.code_1_digits_box_3,
                          "or", code.code_1_digits_box_4, " fetters?")
                elif second.lower() == "open rectangle box" or second.lower() == "open the rectangle box":
                    print("---------------------------------------------------------")
                    print("there is a four digit console on the box enter right combination to open")
                    rectangle_open = input("Do you want to enter the digits\n")
                    if rectangle_open.lower() == "yes":
                        print("Enter code...\n")
                        rectangle_code = input()
                        print(code.code_box_1_digit_collect_1_2_3_4)
                        if rectangle_code == str(code.keep_code_4_digits_box_1):
                            print("The code is correct!! You hear a clicking sound an the box opens")
                            print("You take a piece of an object, that looks like a half of a key")
                            print("There is also 200 gil in the box, share them wise")
                            cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [100]})
                            elena.x_add_items_inventory({"Item": ["gil"], "QTY": [100]})
                            cloud.key_item_1 = 1
                            cloud.key_item_1_name = "half of a key"
                            if cloud.key_item_1 == 1 and cloud.key_item_2 == 1:
                                print("Looks like you have both pieces")
                                print("if put together... do you want to try?\n")
                                key_input = input("yes or no\n")
                                if key_input.lower() == "yes":
                                    cloud.key_item_3 = 1
                                    cloud.key_item_3_name = "Key of the serpent"
                                    print("You know have the key of the serpent")
                                else:
                                    print("Even if you don't want to, the pieces are put together")
                                    cloud.key_item_3 = 1
                                    cloud.key_item_3_name = "Key of the serpent"
                                    print("You know have the key of the serpent")
                                    print("---------------------------------------------------------")
                            else:
                                print("You only have one piece, and know needs to find the second part")
                        else:
                            print("When you entered the code", rectangle_code, "a mysterious voice calls your name")
                            print("\ryou", cloud.name, end="")
                            time.sleep(time_long_wait)
                            print("\rhave entered the wrong code...", end="")
                            time.sleep(time_long_wait)
                            print("\rprepare your self...", end="")
                            time.sleep(time_short_wait)
                            YourHero.x_battle(cloud, 0, cloud, elena, "no")
                    elif rectangle_open.lower() == "no":
                        print("---------------------------------------------------------")
                        print("wise choice if you don't know the password")
                elif second.lower() == "take the triangle box":
                    print("---------------------------------------------------------")
                    print("Its well anchored to the table, cant be removed.")
                elif second.lower() == "open triangle box" or second.lower() == "open the triangle box":
                    print("---------------------------------------------------------")
                    print("There is a four digit console on the box enter right combination to open,"
                          " it shall open from "
                          "left to right")
                    triangle_open = input("Do you want to enter the digits\n")
                    if triangle_open.lower() == "yes":
                        print("Enter code...\n")
                        triangle_code = input()
                        if triangle_code == str(code.code_box_1_digit_collect_1_2_3_4):
                            print("The code is correct!! You hear a clicking sound an the box opens")
                            print("You pick up a piece of an object, that looks like a half of a key")
                            print("There is also 300 gil in the box, share them wise")
                            cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [150]})
                            elena.x_add_items_inventory({"Item": ["gil"], "QTY": [150]})
                            cloud.key_item_2 = 1
                            cloud.key_item_2_name = "half of a key"
                            if cloud.key_item_1 == 1 and cloud.key_item_2 == 1:
                                print("Looks like you have both pieces")
                                print("if put together... do you want to try?\n")
                                key_input = input("yes or no\n")
                                if key_input.lower() == "yes":
                                    cloud.key_item_3 = 1
                                    cloud.key_item_3_name = "Key of the serpent"
                                    print("You know have the key of the serpent")
                                    print("---------------------------------------------------------")
                                else:
                                    print("Even if you don't want to, the pieces are put together")
                                    cloud.key_item_3 = 1
                                    cloud.key_item_3_name = "Key of the serpent"
                                    print("You know have the key of the serpent")
                                    print("---------------------------------------------------------")
                            else:
                                print("You know have the first part of the key,"
                                      " and you need to find the second part")
                                print("---------------------------------------------------------")
                        else:
                            print("When you entered the code", triangle_code, "a mysterious voice calls your name")
                            print("\ryou", cloud.name, end="")
                            time.sleep(time_short_wait)
                            print("\rhave entered the wrong code...", end="")
                            time.sleep(time_short_wait)
                            print("\rprepare your self...", end="")
                            time.sleep(time_short_wait)
                            YourHero.x_battle(cloud, 0, cloud, elena, "no")
                    elif triangle_open.lower() == "no":
                        print("---------------------------------------------------------")
                        print("wise choice if you don't know the password")
                elif second.lower() == "walk to the door" or second.lower() == "walk to door":
                    print("---------------------------------------------------------")
                    print("You walk to the door, and see that the door has an encryption written on it"
                          " 'only the snake can open the door' ")
                elif second.lower() == "use key of the serpent" or second.lower() == "use key" \
                        or second.lower() == \
                        "open" or second.lower() == "open door" or second.lower() == "open the door":
                    if cloud.key_item_3 == 1:
                        print("---------------------------------------------------------")
                        print("You use the key of the serpent and open the door")
                        print("The key has no more purpose and vanish")
                        cloud.key_item_1 = 0
                        cloud.key_item_2 = 0
                        cloud.key_item_3 = 0
                        cloud.key_item_4 = 0
                        cloud.key_item_1_name = ""
                        cloud.key_item_2_name = ""
                        cloud.key_item_3_name = ""
                        cloud.key_item_4_name = ""
                        elena.help = 0
                        cloud.chapter_1_round_room_count += 1
                        elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[4])
                        ask_save()
                        elena.x_story_chapter_count()
                    elif cloud.key_item_3 == 0:
                        print("---------------------------------------------------------")
                        print("You try to open, but without the key its impossible, "
                              "and a strange fog fills the room...they are coming!!")
                        YourHero.x_battle(cloud, 0, cloud, elena, "no")
                elif second.lower() == "help":
                    if elena.help == 0:
                        elena.help += 1
                        print("---------------------------------------------------------")
                        print("Walk around and look, if your are not interested in art, maybe today you should...")
                        print("This was your first clue, be careful to use them to much")
                    elif elena.help == 1:
                        elena.help += 1
                        print("---------------------------------------------------------")
                        print("The boxes on the table opens more together then them self's...")
                        print("This was your second clue, I have warned you....one more clue and....")
                    elif elena.help == 2:
                        elena.help += 1
                        print("---------------------------------------------------------")
                        print("Last clue!!\nThe bird and hero holds the key...")
                        print("As you receive the last clue,"
                              " the candles in the room blows out by a power ful wind\n"
                              "they are coming!!!!!")
                        YourHero.x_battle(cloud, 0, cloud, elena, "no")
                elif "chandelier" in second:
                    print("---------------------------------------------------------")
                    print("A beautiful hand-made chandelier, with candles that lights up the whole room. "
                          "Hope they don't burn out")
                elif "burn paper" in second or second.lower() == "throw paper at fire" or second.lower() == "burn paper with candles" \
                        or second.lower() == "burn paper with chandelier" or second.lower() == "throw paper at candles":
                    if cloud.key_item_4 == 1:
                        print("---------------------------------------------------------")
                        print("You throw the paper from the clay hand into the fire of the chandelier,"
                              "\na golden smog fills the room"
                              "\nmaking your eyes blind for a few seconds, then you hear a voice saying: ")
                        print("---------------------------------------------------------")
                        statue_voice = "Follow me down, i have something to you, but don't cut yourself on my axe\n"
                        for char in statue_voice:
                            sys.stdout.write(char)
                            sys.stdout.flush()
                            time.sleep(time_writing)
                    else:
                        print("What are you trying to do?")
                elif "statue" in second:
                    print("---------------------------------------------------------")
                    print("Its an old statue, made of iron. The statue has a axe in its left hand "
                          "and the right hand is making a fist, like its holding something valuable")
                elif second.lower() == "look at the axe":
                    print("---------------------------------------------------------")
                    print("Its made to kill, absolutely a fantastic weapon,"
                          " if it strikes down, it will tare a whole...")
                elif "left hand" in second:
                    if cloud.key_item_4 == 1:
                        print("---------------------------------------------------------")
                        print("You move the left hand down, carefully without touching the axe, "
                              "a clicking sound echoes out in the room")
                        print("And a hidden hatch in the floor, just next to you opens up,"
                              " you can now see a letter going down")
                        cloud.key_item_4 = 0
                        cloud.key_item_4_name = ""
                        print("Do you want to climbed down the letter?\n Yes\n No")
                        hatch_answer = input()
                        if hatch_answer.lower() == "yes" or hatch_answer == "y":
                            print("---------------------------------------------------------")
                            print("You start to climbed down the latter in to the darkness")
                            print("As you set your fot on the ground, the hatch closes")
                            elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[7])
                            ask_save()
                            story1_basement()
                        elif hatch_answer.lower() == "no" or hatch_answer == "n":
                            print("You decide to stay in the room, looking for more clues")
                    else:
                        print("The hand just want to move, 'smoke has not yet been seen' ")
                elif second.lower() == "touch the axe" or second.lower() == "grab the axe" \
                        or second.lower() == "take the axe":
                    print("---------------------------------------------------------")
                    print("As you try to", second.lower(), "you immediately feel the "
                                                           "sharpness of its blade, cutting your hand,"
                                                           " loosing 1 HP")
                    cloud.hp = cloud.hp - 1
                    elena.hp = elena.hp - 1
                    if elena.hp < 0:
                        YourHero.x_game_over(elena)
                    elif cloud.hp < 0:
                        YourHero.x_game_over(elena)
                    elif cloud.hp and elena.hp < 0:
                        YourHero.x_game_over(cloud and elena)
                elif "right hand" in second:
                    print("---------------------------------------------------------")
                    print("There is something strange with the right hand, "
                          "the material is different from the rest of the statue")
                    print("It looks like it's made of clay and could easily fall of")
                elif second.lower() == "take the clay" or second.lower() == "take clay" or second.lower() == "clay":
                    print("---------------------------------------------------------")
                    print("As you reach for the right hand you feel the wind on your back, and voice saying: "
                          "\n'where to hide, a piece of word, in clay they walked, we searched their trial")
                elif second.lower() == "search clay" or second.lower() == "investigate clay" \
                        or second.lower() == "search right hand" or second.lower() == "search clay hand":
                    print("---------------------------------------------------------")
                    print("You follow the trail...and as you start investigating the clay,"
                          " you feel a piece of paper")
                elif "paper" in second:
                    print("---------------------------------------------------------")
                    print("There is an very old note with a few lines of text: 'when fire burns i will vanish' ")
                    print("You have received key item, secret words")
                    cloud.key_item_4 = 1
                    cloud.key_item_4_name = "Secret words"
                elif second.lower() == "status screen" or second.lower() == "stats":
                    print(elena.item_list)
                    print("---------------------------------------------------------")
                    print(cloud.item_list)
                    print("---------------------------------------------------------")
                    input("Press enter to continue")
                    print(elena.total_stat)
                    print("---------------------------------------------------------")
                    print(cloud.total_stat)
                    input("Press enter to continue\n")
                elif second.lower() == "exit()":
                    cloud.x_exit_game()
                else:
                    print("---------------------------------------------------------")
            elif room_time >= 25:
                print("A strange cold wind, pass your feet, and moves up to the big chandelier,"
                      "\nbefore you even had a chance to think,"
                      "\nit blows out the lights, and the room turns pitch black"
                      "\ndon't mind to remember the the codes you have found, they are now useless!!"
                      "\n....They are coming!!!")
                time.sleep(time_long_wait)
                YourHero.x_battle(cloud, 0, cloud, elena, "no")




def story1_basement():
    while True:
        play.music_loop(play.sound_chapter_1[12])
        room = 2
        # second = ""
        while room == 2:
            if room == 2:
                print("---------------------------------------------------------")
                print("The darkness have never been darken then down here, where is here???")
                second = input("\nAfter the hatch closed behind you, you feel that you are not alone,"
                               "\nsomeone or something is down there with you\n"
                               "Infront of you you see two tunnels one to the right and one to the left\n"
                                "What will you do?")
                print("---------------------------------------------------------")
                if second.lower() == "shout at the person you heard" or second.lower() == "hello" \
                        or second.lower() == "somebody there":
                    print("As you shout out in the dark, you hear a noice coming from the left tunnel!!!")
                    print("---------------------------------------------------------")

                elif second.lower() == "left" or second.lower() == "walk to the left":
                    print("The walls in the tunnel are brown-red with spots coverd in green slime")
                    print("The air is surprisingly fresh, and a  bit cold")
                    print("As you go further in to the tunnel you hear the noice from something or someone better")
                    print("As further you go, you see a light hanging above a door")
                    print("Without a doubt you open the door....")
                    print("There in front of you is two chairs, there is a young lady sitting in one of them")
                    input("Press enter")
                    lady_in_chair = f"Hello {cloud.name} and {elena.name}, I have been waiting for you...\n"
                    for char in lady_in_chair:
                        sys.stdout.write(char)
                        sys.stdout.flush()
                        time.sleep(time_writing)
                    lady_in_chair_2 = f"I have something you want, but you need to earn it"
                    for char in lady_in_chair_2:
                        sys.stdout.write(char)
                        sys.stdout.flush()
                        time.sleep(time_writing)
                    game_question = input("\nDo you want to play a game?")
                    if game_question.lower() == "yes":
                        print("Wise choice, lets start....")
                        game_tick_tack.start_game()
                        if game_tick_tack.win >= 1:
                            game_tick_tack.win = 0
                            print("You are good\n")
                            print("You will have the one thing you need! Bye Bye")
                            print("The mist takes over the room, blinding your eyes....and you return to the ally")
                            cloud.x_add_items_inventory({"Item": ["Key world item 2"], "QTY": ["Iron Gloves"]})
                            print("You receive the", cloud.x_see_inventory_name_qty("item", "Key world item 2"))
                            cloud.key_item_1 = 0
                            cloud.key_item_2 = 0
                            cloud.key_item_3 = 0
                            cloud.key_item_4 = 0
                            cloud.key_item_1_name = ""
                            cloud.key_item_2_name = ""
                            cloud.key_item_3_name = ""
                            cloud.key_item_4_name = ""
                            elena.help = 0
                            cloud.chapter_1_round_room_count += 1
                            cloud.key_world_item_2 = 1
                            elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[4])
                            ask_save()
                            elena.x_story_chapter_count()
                        elif game_tick_tack.win == 0:
                            print("No good, i will punish you, they will come")
                            print("But return to me if you wish to play again....")
                            YourHero.x_battle(cloud, 0, cloud, elena, "no")
                            print("---------------------------------------------------------")

                elif second.lower() == "right" or second.lower() == "walk to the right":
                    print("---------------------------------------------------------")
                    print("As you start walking in to the right tunnel you suddenly slip and lands on your back\n")
                    print("Something is coming for you...!!")
                    print("---------------------------------------------------------")
                    time.sleep(time_short_wait)
                    escape_choice = input("Do you want to try to escape before they are coming??\n")
                    print("---------------------------------------------------------")
                    if escape_choice == "yes" or escape_choice == "Yes":
                        escape_chance = random.randint(0, 5)
                        if escape_chance >= 3:
                            print(
                                "You manage to get up from the ground and escape and return\n")
                            print(" to the entrence of the cave\n")

                            print("---------------------------------------------------------")

                            story1_basement()
                        elif escape_chance <= 2:
                            print("You desperate try to get back on your feet, but fails,"
                                  " you see the shadows on the wall, they are coming!!!")
                            print("---------------------------------------------------------")
                            YourHero.x_battle(cloud, 0, cloud, elena, "no")
                            print("As you win the battle, they have left a chest for you\n")
                            open_chest = input("Do you want to open the chest yes/no?")
                            print("---------------------------------------------------------")
                            if open_chest.lower() == "yes":
                                money_give = random.randint(10, 150)
                                print(f"You recive {money_give} gil each")
                                print("And walk back to the cave entrance")
                                print("---------------------------------------------------------")
                                cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [money_give]})
                                elena.x_add_items_inventory({"Item": ["gil"], "QTY": [money_give]})
                            elif open_chest.lower() == "no":
                                print("You walk back to the cave entrance, leaving the chest alone")
                                print("---------------------------------------------------------")
                            else:
                                print("You walk back to the cave entrance, leaving the chest alone")
                                print("---------------------------------------------------------")
                    else:
                        print("Please choose a valid answer, yes, or no.")
                        print("---------------------------------------------------------")




                elif second.lower() == "exit()":
                    cloud.x_exit_game()
                else:
                    print("---------------------------------------------------------")


def story1_mountains_in_north():
    play.music_loop(play.sound_chapter_1[10])
    print("---------------------------------------------------------")
    tprint("Mountains in north")
    print("---------------------------------------------------------")
    print("The cliffs are sharp and steep, no way to climb without 'iron gloves'")
    if "Iron Gloves" in cloud.x_see_inventory_name_qty("item", "Key world item 2"):
        print("You start to climb")
        print("---------------------------------------------------------")
        while True:
            vertical = 1
            # second = ""
            while vertical == 1:
                print("---------------------------------------------------------")
                print("As you hold on to the hard cliffs you see a light coming out from the mountain to the left\n"
                      "To the right you can see the shadows moves in a mysterious way")
                where_to_climb = input("Where do you want to climb?")
                if "left" in where_to_climb:
                    vertical = 2
                elif "right" in where_to_climb:
                    vertical = 3

            while vertical == 2:
                print("You choose the left way....")
                input("Press enter")
                story1_first_crossing()

            while vertical == 3:
                print("You choose the right way....")
                input("Press enter")
                story1_first_crossing()
                # YourHero.x_battle(cloud, 1, cloud, elena, "yes")
    else:
        print("Without the gloves its impossible to climb, you fall down while trying, hopelessly returning back")
        print("---------------------------------------------------------")
        input("press enter")
        story1_first_crossing()


def story1_forest_in_east():
    reward = 0
    while True:
        while reward <= 5:
            play.music_loop(play.sound_chapter_1[8])
            print("---------------------------------------------------------")
            tprint("Forest in east")
            print("---------------------------------------------------------")
            print("The forest have many tall trees, reaching all the way to the sky")
            print("There is a rumor about this forest, have you heard about it?")
            print("---------------------------------------------------------")
            rumors_choice = input("Would you like to share a rumor??")
            print("---------------------------------------------------------")
            if rumors_choice.lower() == "yes" or rumors_choice.lower() == "YES":
                rumors_talk = input("What have you heard??")
                print("---------------------------------------------------------")
                if rumors_talk in cloud.x_see_inventory_name_qty("item", "Key world item 3"):
                    print("You remember correctly, a reward may be in your hands")
                    print("---------------------------------------------------------")
                    rumors_reward = random.randint(0, 6)
                    print("Let the trees decide...")
                    if rumors_reward >= 3:
                        reward_money = random.randint(2, 30)
                        cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [reward_money]})
                        elena.x_add_items_inventory({"Item": ["gil"], "QTY": [reward_money]})
                        print("You are lucky, they have decided to give you each", reward_money, "GIL")
                        reward += 1
                        input("press enter")
                    elif rumors_reward <= 2:
                        take_money = random.randint(1, 5)
                        cloud.x_remove_items_inventory({"Item": ["gil"], "QTY": [take_money]}, "no", cloud, "yes")
                        elena.x_remove_items_inventory({"Item": ["gil"], "QTY": [take_money]}, "no", elena, "yes")
                        print("Bad luck, the forest will take", take_money, "Gil, from each of you")
                        input("press enter")
                        reward += 1
                else:
                    print("That is not what i have heard, come back later")
                    print("---------------------------------------------------------")
                    # cloud.key_world_item_3 = 1
                    time.sleep(time_short_wait)
                    story1_first_crossing()
                    break
            else:
                print("You choose not to share a rumor, and are going back to the crossing")
                print("---------------------------------------------------------")
                # cloud.key_world_item_3 = 1
                time.sleep(time_short_wait)
                story1_first_crossing()
                break



def story1_snow_in_south():
    elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[8])
    play.music_loop(play.sound_chapter_1[9])
    escape = 0
    print("---------------------------------------------------------")
    tprint("Snow in south")
    print("---------------------------------------------------------")
    if "Snow Boots" in cloud.x_see_inventory_name_qty("item", "Key world item"):
        print("You put your snow boots on, and start walk, feeling the snow crunch under neath you")
        print("---------------------------------------------------------")
    else:
        while True:
            if escape == 0:
                print("Without the snow boots...\nIts impossible to walk")
                print("You take some steps but falls down in the deep snow and get stuck,"
                      " something is moving in the snow!!!")
                print("---------------------------------------------------------")
                time.sleep(time_short_wait)
                escape_choice = input("Do you want to try to escape before they are coming??\n")
                print("---------------------------------------------------------")
                if escape_choice == "yes" or escape_choice == "Yes":
                    escape_chance = random.randint(0, 5)
                    if escape_chance >= 3:
                        print("You manage to get up from the snow, and crawl your way back before they come")
                        print("---------------------------------------------------------")
                        escape += 1
                        story1_first_crossing()
                    elif escape_chance <= 2:
                        print("You desperate try to crawl your self up from the snow, but fail,"
                              " you see the shadows in the snow getting closer, they are coming!!!")
                        print("---------------------------------------------------------")
                        YourHero.x_battle(cloud, 0, cloud, elena, "no")
                else:
                    print("Please choose a valid answer, yes, or no.")
                    print("---------------------------------------------------------")
            elif escape == 2:
                print("You manage to get up from the snow, and crawl your way back before they come")
                print("---------------------------------------------------------")
                story1_first_crossing()


def story1_abrehiem_town():
    elena.music_name = ""
    town = 1
    print("---------------------------------------------------------")
    tprint("Town in west")
    print("---------------------------------------------------------")
    time.sleep(time_short_wait)
    while True:
        play.music_loop(play.sound_chapter_1[1])
        while town == 1:
            if town == 1:
                print("---------------------------------------------------------")
                tprint("Welcome to:")
                tprint("Abreheim")
                print("---------------------------------------------------------")
                print("The small town reminds you of home, the cobblestones are covering the ground and leading the"
                      " way throughout the whole town")
                time.sleep(time_short_wait)
                print("The towns-people are walking to different stores, talking to each-other")
                print("The houses are tiny but cozy, every one of them seems to say welcome stranger")
                print("---------------------------------------------------------")
                time.sleep(time_short_wait)
                print("There is an item shop, a weapon shop and a materia shop ")
                print("You can stay at an inn 'where you can rest'.")
                print("There is a smart looking girl with an umbrella looking at you, as you walk in to the town")
                print("or walk to the strange ally with dim lights and shady people. Or do you want to leave the town?")
                walk_choice = input("Where do you want to go??")
                print("---------------------------------------------------------")
                if "item shop" in walk_choice:
                    print("You start walking to the item shop, you feel a calling wind from the strange ally")
                    print("The door to the shop is a robust door made of solid wood, you open the door and enter")
                    print("It's very warm inside, hardwood floors, decorated walls with paintings of old towns people")
                    print("A bookshelf is standing tall, proud of the books it keeps")
                    print("When you come to the counter and friendly clerk greets you")
                    print("---------------------------------------------------------")
                    input("press enter to continue")
                    play.music_loop(play.sound_chapter_1[2])
                    YourHero.x_shop(elena, "Abreheim's items", "Lisa", "Abreheim's items_default", "yes")
                    YourHero.x_shop(cloud, "Abreheim's items", "Lisa", "Abreheim's items_default", "no")
                    play.music_loop(play.sound_chapter_1[1])
                    # print("Something more")
                    input("Press enter to continue")
                elif "materia shop" in walk_choice:
                    # need to add music to this shop
                    play.music_loop(play.sound_chapter_1[11])
                    YourHero.x_materia_shop(elena, "Abreheim's materia", "Mystica")
                    YourHero.x_materia_shop(cloud, "Abreheim's materia", "Mystica")
                    play.music_loop(play.sound_chapter_1[1])
                elif "weapon shop" in walk_choice:
                    play.music_loop(play.sound_chapter_1[6])
                    YourHero.x_weapon_armory_shop(elena, "Abreheim's weapons", "Ophelia")
                    YourHero.x_weapon_armory_shop(cloud, "Abreheim's weapons", "Ophelia")
                    play.music_loop(play.sound_chapter_1[1])
                    input("press enter to continue")
                elif "umbrella" in walk_choice:
                    print("Hei there ")
                    talk_choice = input("Wanna know a secret?")
                    if talk_choice == "yes":
                        print("Ok, good.....")
                        print("Talk with the woman in a green hat in the ally!")
                        print("Do do so you need this...")
                        cloud.x_remove_items_inventory({"Item": ["Key world item 3"], "QTY": [""]}, "no", cloud, "no")
                        cloud.x_add_items_inventory({"Item": ["Key world item 3"], "QTY": ["rumor"]})
                        print("You received the Key world item 3",
                              cloud.x_see_inventory_name_qty("item", "Key world item 3"))
                        input("press enter")
                        print("---------------------------------------------------------")
                    else:
                        print("See you late, I hope!")
                        input("press enter")
                        print("---------------------------------------------------------")
                elif "ally" in walk_choice:
                    story1_ally_in_abreheim(1)
                elif "inn" in walk_choice:
                    story1_inn_at_abreheim_town()
                elif "leave" in walk_choice:
                    story1_first_crossing()
                elif walk_choice.lower() == "inventory":
                    YourHero.x_inventory(elena)
                    YourHero.x_inventory(cloud)
                elif walk_choice == "exit()":
                    cloud.x_exit_game()
                else:
                    print("please make up your mind!")
                    print("---------------------------------------------------------")


def story1_inn_at_abreheim_town():
    print("---------------------------------------------------------")
    tprint("Abreheim Inn")
    print("---------------------------------------------------------")
    while True:
        stranger_at_bar = 0
        while stranger_at_bar == 0:
            if stranger_at_bar == 0:
                play.music_loop(play.sound_chapter_1[5])
                print("When you enter the Abreheims Inn, like a warm hug the place takes you in")
                print("In front of you is a woman standing at the counter and smiling")
                print("To the left is the library and to the right is the restaurant")
                print("---------------------------------------------------------")
                walk_choice = input("Where do you want to go?\n")
                if "right" in walk_choice or "restaurant" in walk_choice:
                    print("Excuse me, im sorry but you need to book a table in order to get to the restaurant")
                    print("---------------------------------------------------------")
                elif "left" in walk_choice or "library" in walk_choice:
                    if cloud.sleep_abreheim_inn == 0:
                        woman_in_white = random.randint(0, 8)
                        if woman_in_white >= 3:
                            while woman_in_white >= 3:
                                print("The library is absolutely fabulous, it's a grand room"
                                      "\nwith hand made bookshelf surrounding the room")
                                print("A beautiful chandelier is hanging in the middle of the room")
                                print("The warm light from the candles embraces the room")
                                print("There is many places to sit, a few arm-chairs are placed sofas")
                                print("As you enter the library, a strange woman in a gray"
                                      "\noutfit pass you by on her way out")
                                print("There is a lady in all white sitting in one of the sofas, and reading a book")
                                print("Another woman is sitting in one of the arm-chairs and writing her journal")
                                print("or leave the library")
                                print("---------------------------------------------------------")
                                library_choice = input("What is on your mind?\n")
                                if "sofa" in library_choice or "lady" in library_choice or "white" in library_choice:
                                    print("Hello stranger, "
                                          "the lady that just passed you had something interesting to tell")
                                    print("But you look tired, maybe you should sleep here for a night,"
                                          "and we can talk")
                                    print("---------------------------------------------------------")
                                elif "arm-chair" in library_choice or "journal" in \
                                        library_choice or "woman" in library_choice:
                                    print("Sorry im busy at the moment")
                                    print("---------------------------------------------------------")
                                elif "leave" in library_choice:
                                    story1_inn_at_abreheim_town()
                                elif walk_choice == "exit()":
                                    cloud.x_exit_game()
                                else:
                                    print("Try talking to someone in here")
                                    print("---------------------------------------------------------")
                        elif woman_in_white <= 2:
                            while woman_in_white <= 2:
                                print("The library is absolutely fabulous, it's a grand room"
                                      "\nwith hand made bookshelf surrounding the room")
                                print("A beautiful chandelier is hanging in the middle of the room")
                                print("The warm light from the candles embraces the room")
                                print("There is many places to sit, a few arm-chairs are placed sofas")
                                print("A woman is sitting in one of the arm-chairs and writing her journal")
                                print("or leave the library")
                                print("---------------------------------------------------------")
                                library_choice = input("What is on your mind?\n")
                                if "arm-chair" in library_choice or "journal" in \
                                        library_choice or "woman" in library_choice:
                                    print("Good day stranger, come back another time when the lady in white is here")
                                    print("Talk to her, she has something interesting to tell you")
                                    print("---------------------------------------------------------")
                                elif "leave" in library_choice:
                                    story1_inn_at_abreheim_town()
                                elif walk_choice == "exit()":
                                    cloud.x_exit_game()
                                else:
                                    print("Try talking to someone in here")
                                    print("---------------------------------------------------------")
                    elif cloud.sleep_abreheim_inn == 1:
                        woman_in_white = random.randint(0, 8)
                        if woman_in_white >= 3:
                            while woman_in_white >= 3:
                                print("The library is absolutely fabulous, it's a grand room"
                                      "\nwith hand made bookshelf surrounding the room")
                                print("A beautiful chandelier is hanging in the middle of the room")
                                print("The warm light from the candles embraces the room")
                                print("There is many places to sit, a few arm-chairs are placed sofas")
                                print("As you enter the library, a strange woman in a gray"
                                      "\noutfit pass you by on her way out")
                                print("There is a lady in all white sitting in one of the sofas, and reading a book")
                                print("Another woman is sitting in one of the arm-chairs and writing her journal")
                                print("or leave the library")
                                print("---------------------------------------------------------")
                                library_choice = input("What is on your mind?\n")
                                if "sofa" in library_choice or "lady" in library_choice or "white" in library_choice:
                                    print("How was your night, hope you could sleep alright")
                                    print("As promised i will tell you what she said to me")
                                    print("Walk to the ally, talk to the lady facing the wall, and then....")
                                    print("The arm of the statue, will lead the way down....")
                                    print("---------------------------------------------------------")
                                elif "arm-chair" in library_choice or "journal" in \
                                        library_choice or "woman" in library_choice:
                                    print("Sorry im busy at the moment")
                                    print("---------------------------------------------------------")
                                elif "leave" in library_choice:
                                    story1_inn_at_abreheim_town()
                                elif walk_choice == "exit()":
                                    cloud.x_exit_game()
                                else:
                                    print("Try talking to someone in here")
                        elif woman_in_white <= 2:
                            while woman_in_white <= 2:
                                print("The library is absolutely fabulous, it's a grand room"
                                      "\nwith hand made bookshelf surrounding the room")
                                print("A beautiful chandelier is hanging in the middle of the room")
                                print("The warm light from the candles embraces the room")
                                print("There is many places to sit, a few arm-chairs are placed sofas")
                                print("A woman is sitting in one of the arm-chairs and writing her journal")
                                print("or leave the library")
                                print("---------------------------------------------------------")
                                library_choice = input("What is on your mind?\n")
                                if "arm-chair" in library_choice or "journal" in \
                                        library_choice or "woman" in library_choice:
                                    print("Good day stranger, come back another time when the lady in white is here")
                                    print("Talk to her, she has something interesting to tell you")
                                    print("---------------------------------------------------------")
                                elif "leave" in library_choice:
                                    story1_inn_at_abreheim_town()
                                elif walk_choice == "exit()":
                                    cloud.x_exit_game()
                                else:
                                    print("Try talking to someone in here")
                                    print("---------------------------------------------------------")
                elif "reception" in walk_choice or "counter" in walk_choice or "front" in walk_choice:
                    print("As you walk to the reception, the smiling woman greets you")
                    cloud.x_sleep_inn("Abreheim Inn", "Michelle", "small room",
                                      "grand room", 100, 300, "Lasagne", 20, "Fish & rice", 40)
                    if cloud.inn_room_1 == 1:
                        stranger_at_bar = 1
                    elif cloud.inn_room_2 == 1:
                        stranger_at_bar = 2
                    elif cloud.food_1 == 1:
                        stranger_at_bar = 3
                    elif cloud.food_2 == 1:
                        stranger_at_bar = 4
                    else:
                        print("---------------------------------------------------------")
                elif "leave" in walk_choice or "town" in walk_choice:
                    story1_abrehiem_town()
                elif walk_choice == "exit()":
                    cloud.x_exit_game()
                else:
                    print("Where do you want to go?")
                    print("---------------------------------------------------------")
        while stranger_at_bar == 1:
            if stranger_at_bar == 1:
                cloud.sleep_abreheim_inn = 1
                cloud.inn_room_1 = 0
                closet_findings = 0
                drawer_findings = 0
                print("---------------------------------------------------------")
                tprint("Sleeping")
                print("---------------------------------------------------------")
                cloud.x_initials_stats()
                elena.x_initials_stats()
                cloud.x_tent_use()
                elena.x_tent_use()
                play.music_loop(play.music_for_all[0])
                sleep_text = "zzzz....\n"
                now = time.time()
                future = now + 15
                while time.time() < future:
                    for char in sleep_text:
                        sys.stdout.write(char)
                        sys.stdout.flush()
                        time.sleep(time_save)
                print("---------------------------------------------------------")
                rest_over = 0
                while rest_over == 0:
                    if rest_over == 0:
                        play.music_loop(play.music_for_all[1])
                        print("You wake up in one of the cozy beds at the inn")
                        print("It is a kind of a small room with two beds, a drawer,"
                              "\na closet and and a window.")
                        print("The door out to the stair way and the lobby is made of wood")
                        print("---------------------------------------------------------")
                        room_choice = input("What do you want to do?")
                        if "closet" in room_choice:
                            print("You walk over to the closet, and you see there is a knob on it")
                            print("That seems to been turned before")
                            print("---------------------------------------------------------")
                        elif "drawer" in room_choice:
                            print("You open the drawer...")
                            print("---------------------------------------------------------")
                            open_drawer = random.randint(0, 10)
                            if open_drawer >= 6:
                                if drawer_findings == 0:
                                    print("Received 50 gil each!")
                                    print("---------------------------------------------------------")
                                    cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [50]})
                                    elena.x_add_items_inventory({"Item": ["gil"], "QTY": [50]})
                                    drawer_findings = 1
                                    print("---------------------------------------------------------")
                                else:
                                    print("Drawer is empty")
                                    print("---------------------------------------------------------")
                            elif open_drawer <= 4:
                                if open_drawer == 0:
                                    print("The drawer is empty")
                                    print("---------------------------------------------------------")
                                    drawer_findings = 1
                                else:
                                    print("Drawer is empty")
                                    print("---------------------------------------------------------")
                            elif open_drawer == 5:
                                if open_drawer == 0:
                                    cloud.x_finding_weapon_armory_on_journey("The mark of the serpents")
                                    elena.x_finding_weapon_armory_on_journey("The mark of the serpents")
                                    drawer_findings = 1
                                    print("You found the", cloud.x_see_inventory_name_qty("ring", 1)[0])
                                    print("---------------------------------------------------------")
                                else:
                                    print("Drawer is empty")
                                    print("---------------------------------------------------------")
                        elif "closet" in room_choice or "knob" in room_choice:
                            print("You open the closet...")
                            open_closet = random.randint(0, 10)
                            if open_closet >= 5:
                                if closet_findings == 0:
                                    print("You found two potion!")
                                    print("Received two potions each!")
                                    elena.x_add_items_inventory({"Item": ["potion"], "QTY": [2]})
                                    cloud.x_add_items_inventory({"Item": ["potion"], "QTY": [2]})
                                    closet_findings = 1
                                    print("---------------------------------------------------------")
                                else:
                                    print("Closet is empty")
                            elif open_closet <= 4:
                                if open_closet == 0:
                                    print("The closet is empty")
                                    closet_findings = 1
                                else:
                                    print("Closet is empty")
                        elif "leave" in room_choice or "out" in room_choice:
                            print("You walk out of the inn and back to the town")
                            print("---------------------------------------------------------")
                            input("press enter")
                            closet_findings = 0
                            drawer_findings = 0
                            story1_abrehiem_town()
                        elif "window" in room_choice:
                            print("The view over the city is beautiful")
                            print("---------------------------------------------------------")
                        else:
                            print("What do you want to do?")
                            print("---------------------------------------------------------")
        while stranger_at_bar == 2:
            if stranger_at_bar == 2:
                cloud.sleep_abreheim_inn = 1
                cloud.inn_room_2 = 0
                print("Sleeping..... in room 2")
                print("---------------------------------------------------------")
                input("press enter")
                story1_abrehiem_town()
        while stranger_at_bar == 3:
            if stranger_at_bar == 3:
                cloud.food_1 = 0
                print("Eating, food 1")
                print("---------------------------------------------------------")
                input("press enter")
                story1_abrehiem_town()
        while stranger_at_bar == 4:
            if stranger_at_bar == 4:
                cloud.food_2 = 0
                print("eating food, 2")
                print("---------------------------------------------------------")
                input("press enter")
                story1_abrehiem_town()


def story1_ally_in_abreheim(ally_count):
    while True:
        ally = ally_count
        while ally == 1:
            if ally == 1:
                play.music_loop(play.sound_chapter_1[3])
                elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[4])
                tprint("The  Ally")
                print("---------------------------------------------------------")
                print("Your heart starts to beat hard")
                time.sleep(time_short_wait)
                print("It's like your soul is telling you to stop, but you keep on going...")
                time.sleep(time_short_wait)
                print("The daylight disappear...")
                time.sleep(time_short_wait)
                print("When you finally is one with the mist that surrounds the ally, the town is far behind you")
                print("The only thing you can see in the mist, is silhouettes of what may be four different people,"
                      "\na small door and the mist that seems to go further and deeper in")
                print("From what you can tell, one woman has a green hat, another woman facing the wall\n"
                      "The other two women you see, are talking to each other but seems upset.\n"
                      "You can always walk back to the town")
                print("---------------------------------------------------------")
                ally_choice = input("What will you do")
                if "green hat" in ally_choice or "hat" in ally_choice:
                    print("You walk over to the woman in the green hat,"
                          " curios what she could help you with on this journey")
                    woman_green_hat = "Hello stranger, haven't seen you around before"
                    for char in woman_green_hat:
                        sys.stdout.write(char)
                        sys.stdout.flush()
                        time.sleep(time_writing)
                    if "rumor" in cloud.x_see_inventory_name_qty("item", "Key world item 3") \
                            and cloud.chapter_1_ally_woman_green_hat_away == 0:
                        while True:
                            woman_green_hat_talks = 1
                            while woman_green_hat_talks == 1:
                                print("Your body shivers and are surprised how kind her voice is")
                                print("---------------------------------------------------------")
                                time.sleep(time_short_wait)
                                woman_green_hat = "You are seeking for something"
                                for char in woman_green_hat:
                                    sys.stdout.write(char)
                                    sys.stdout.flush()
                                    time.sleep(time_writing)
                                print("Do you need items, weapons or a good night sleep,\n"
                                      "or talk to somebody else in the ally?")
                                print("Or are you seeking the answer of the rumor?")
                                talk_back = input("What do you seek?")
                                print("---------------------------------------------------------")
                                if talk_back == "items":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif talk_back == "weapons":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif talk_back == "good night sleep":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif "answer" in talk_back or "rumor" in talk_back:
                                    cloud.chapter_1_ally_woman_green_hat_away = 1
                                    cloud.x_remove_items_inventory({"Item": ["Key world item 3"], "QTY": ["rumor"]},
                                                                   "no",
                                                                   cloud, "no")
                                    cloud.x_add_items_inventory({"Item": ["Key world item 3"],
                                                                 "QTY": ["Tall trees blooms in winter"]})
                                    rumor = cloud.x_see_inventory_name_qty("item", "Key world item 3")
                                    woman_green_hat = "The legend says," \
                                                      " listen carefully\nthey say: ", \
                                                      rumor, ".\nIs what they are seeking"
                                    for char in woman_green_hat:
                                        sys.stdout.write(char)
                                        sys.stdout.flush()
                                        time.sleep(time_writing)
                                    print("\nThe woman walks away, and you walk back to the town")
                                    print("\n---------------------------------------------------------")
                                    input("press enter")
                                    story1_abrehiem_town()
                                    woman_green_hat_talks = 0
                                elif talk_back.lower() == "talk to somebody else in the ally" \
                                        or talk_back.lower() == "talk to somebody else":
                                    story1_ally_in_abreheim(1)
                                else:
                                    print("What did you say?")
                    elif "rumor" not in cloud.x_see_inventory_name_qty("item", "Key world item 3"):
                        while True:
                            woman_green_hat_talks = 1
                            while woman_green_hat_talks == 1:
                                print("Your body shivers and are surprised how kind her voice is")
                                print("---------------------------------------------------------")
                                time.sleep(time_short_wait)
                                woman_green_hat = "You are seeking for something"
                                for char in woman_green_hat:
                                    sys.stdout.write(char)
                                    sys.stdout.flush()
                                    time.sleep(time_writing)
                                print("---------------------------------------------------------")
                                print("Do you need items, weapons or a good night sleep,\n"
                                      "or talk to somebody else in the ally?")
                                talk_back = input("What do you seek?")
                                print("---------------------------------------------------------")
                                if talk_back == "items":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif talk_back == "weapons":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif talk_back == "good night sleep":
                                    print("There is a good store in the town...")
                                    item_store = input("do you want to go there?")
                                    print("---------------------------------------------------------")
                                    if item_store.lower() == "yes":
                                        story1_abrehiem_town()
                                    else:
                                        print("ok, so whats on your mind?")
                                        print("---------------------------------------------------------")
                                elif talk_back.lower() == "talk to somebody else in the ally" \
                                        or talk_back.lower() == "talk to somebody else":
                                    story1_ally_in_abreheim(1)
                                    print("What did you say?")
                                    print("---------------------------------------------------------")
                    elif "Tall trees blooms in winter" in cloud.x_see_inventory_name_qty("item",
                                                                                         "Key world item 3") and cloud.chapter_1_ally_woman_green_hat_away == 1:
                        print("The woman is not to be seen anymore")
                        print("The only thing left of her is her silhouette")
                        print("---------------------------------------------------------")
                        input("press enter to continue")
                        story1_ally_in_abreheim(1)
                elif "wall" in ally_choice or "woman facing" in ally_choice:
                    while True:
                        woman_facing_the_wall = 1
                        while woman_facing_the_wall == 1:
                            if woman_facing_the_wall == 1:
                                item_list = pd.DataFrame(cloud.inventory)
                                if cloud.chapter_1_round_room_count >= 1 and "Iron Gloves" \
                                        in cloud.x_see_inventory_name_qty("item", "Key world item 2") \
                                        and cloud.chapter_1_ally_woman_wall_away == 0:
                                    print("After finding the",
                                          cloud.x_see_inventory_name_qty("item", "Key world item 2"),
                                          "you decide to walk over to the woman again")
                                    print("Her eyes are shining her smile is big, she feels happy,"
                                          " that she could help you")
                                    print("'Very well, you found the",
                                          cloud.x_see_inventory_name_qty("item", "Key world item 2"),
                                          "now you can climb")
                                    print("She turns away, and walk away into the mist")
                                    cloud.chapter_1_ally_woman_wall_away = 1
                                elif "Iron Gloves" in cloud.x_see_inventory_name_qty("item", "Key world item 2") \
                                        and cloud.chapter_1_ally_woman_wall_away == 1:
                                    print("The woman has left in to the mist, and can no more be seen")
                                    print("Only the silhouette of her can be seen in the mist")
                                    print("---------------------------------------------------------")
                                    input("press enter")
                                    time.sleep(time_short_wait)
                                    print("You turn around and see if there is anyone else to talk to")
                                    print("---------------------------------------------------------")
                                    input("press enter")
                                    story1_ally_in_abreheim(1)
                                elif cloud.chapter_1_round_room_count == 1 \
                                        and "Iron Gloves" not in cloud.x_see_inventory_name_qty("item",
                                                                                                "Key world item 2"):
                                    print("You walk over to the mysterious woman again, "
                                          "she is now already facing you")
                                    print("'So you are back, but i sense you didn't find the secret....")
                                    answer_choice = input("'Do you want to try to find it again???")
                                    print("---------------------------------------------------------")
                                    if answer_choice.lower() == "yes" or answer_choice.lower() == "YES":
                                        print("'Don't cut your self, but try to move the hand...'")
                                        print("---------------------------------------------------------")
                                        time.sleep(time_short_wait)
                                        elena.progress = story.story_name_loop(story.chapter_names[0],
                                                                               story.chap_1_names[6])
                                        ask_save()
                                        story1_round_room()
                                    else:
                                        print("Your choice, and that is a bad one.....")
                                elif cloud.chapter_1_round_room_count >= 2 \
                                        and "Iron Gloves" not in cloud.x_see_inventory_name_qty("item",
                                                                                                "Key world item 2"):
                                    print("Once again you walk over to the mysterious woman, "
                                          "her eyes are closed, then opens up when you come closer")
                                    print("'Still no luck....'")
                                    print("'You did not find it....")
                                    answer_choice = input("'Do you want to try to find it again???")
                                    print("---------------------------------------------------------")
                                    if answer_choice.lower() == "yes" or answer_choice.lower() == "YES":
                                        print("'Don't cut your self, but try to move the hand...'")
                                        print("'The one who always stands still,"
                                              " but the left arm is more interesting, then anything else...'")
                                        print("---------------------------------------------------------")
                                        time.sleep(time_short_wait)
                                        elena.progress = story.story_name_loop(story.chapter_names[0],
                                                                               story.chap_1_names[6])
                                        ask_save()
                                        story1_round_room()
                                    else:
                                        print("Your choice, and that is a bad one.....")
                                elif cloud.chapter_1_round_room_count == 0:
                                    print("You walk over to the mysterious woman, that is looking in to the wall")
                                    time.sleep(time_short_wait)
                                    woman_at_wall = "One with a flag, one with a broken wing....."
                                    for char in woman_at_wall:
                                        sys.stdout.write(char)
                                        sys.stdout.flush()
                                        time.sleep(time_writing)
                                    print("\n", cloud.name, ": 'Pardon me ma'am'")
                                    print("---------------------------------------------------------")
                                    input("press enter")
                                    woman_at_wall = "One with a flag, one with a broken wing....."
                                    for char in woman_at_wall:
                                        sys.stdout.write(char)
                                        sys.stdout.flush()
                                        time.sleep(time_writing)
                                    print("You try to get in contact with her, but she seems to repeating her self")
                                    print("---------------------------------------------------------")
                                    print("You wounder what she is saying and what it means")
                                    talk_back = input("What can you say to make her hear you?\n")
                                    if "flag" in talk_back:
                                        print("..........")
                                        print(" 'So your name is", cloud.name, " and yours is", elena.name, " '")
                                        print("The woman has now turned around and are looking at you,"
                                              " her face tells a story")
                                        print("It tells that she has seen a lot,"
                                              " and that a secret is deep within her")
                                        print("'I have something that may interest you.....'")
                                        print("Dirt can be clean, secrets will reveal")
                                        print("---------------------------------------------------------")
                                        answer_choice = input("Do you want to try to find it???")
                                        print("---------------------------------------------------------")
                                        if answer_choice.lower() == "yes" or answer_choice.lower() == "YES":
                                            elena.progress = story.story_name_loop(story.chapter_names[0],
                                                                                   story.chap_1_names[6])
                                            ask_save()
                                            story1_round_room()
                                        else:
                                            print("She turns around and start talking to her self again")
                                            print("---------------------------------------------------------")
                                    elif "wing" in talk_back:
                                        print("..........")
                                        print(" 'So your name is", cloud.name, " and yours is", elena.name)
                                        print("The woman has now turned around and are looking at you,"
                                              " her face tells a story")
                                        print("It tells that she has seen a lot, and that a secret is deep within her")
                                        print("'I have something that may interest you.....'")
                                        print("Pay attention to how many there is from left to right,"
                                              " it should then be four")
                                        print("---------------------------------------------------------")
                                        answer_choice = input("Do you want to try to find it???")
                                        print("---------------------------------------------------------")
                                        if answer_choice.lower() == "yes" or answer_choice.lower() == "YES":
                                            elena.progress = story.story_name_loop(story.chapter_names[0],
                                                                                   story.chap_1_names[6])
                                            ask_save()
                                            story1_round_room()
                                        else:
                                            print("She turns around and start talking to her self again")
                                            print("---------------------------------------------------------")

                                    elif talk_back.lower() == "what are you saying":
                                        woman_at_wall = "One with a flag, one with broken wings....."
                                        for char in woman_at_wall:
                                            sys.stdout.write(char)
                                            sys.stdout.flush()
                                            time.sleep(time_writing)
                                            print("---------------------------------------------------------")
                                    else:
                                        print("Nobody seems to hear you.....")
                                        print("---------------------------------------------------------")
                elif "women" in ally_choice or "couple" in ally_choice:
                    print("When you come close the the two women, they look at you for a second and then push you away")
                    time.sleep(time_short_wait)
                    print("---------------------------------------------------------")
                elif "town" in ally_choice:
                    print("The ally creeps you out, and you decide to leave")
                    print("---------------------------------------------------------")
                    story1_abrehiem_town()
                elif "small door" in ally_choice or "door" in ally_choice:
                    print("There you are in the mist, in front of a small door")
                    print("The door has no key-hole or any handle, seems like a door, you can leave but not enter")
                    print("---------------------------------------------------------")
                elif "deep" in ally_choice or "further" in ally_choice:
                    print("You decide to walk deeper in to the mist....")
                    print("---------------------------------------------------------")
                    elena.progress = story.story_name_loop(story.chapter_names[0],
                                                           story.chap_1_names[5])
                    ask_save()
                    ally = 2
                elif ally_choice.lower() == "inventory":
                    YourHero.x_inventory(elena)
                    YourHero.x_inventory(cloud)
                elif ally_choice == "exit()":
                    cloud.x_exit_game()
                else:
                    print("The mist must have put a spell on your fingers, cant read what you wrote!")
                    print("---------------------------------------------------------")
                    input("press enter")
        while ally == 2:
            if ally == 2:
                play.music_loop(play.sound_chapter_1[7])
                tprint("The dark Ally")
                print("---------------------------------------------------------")
                print("Will you survive???")
                survive = input("Walk back to ally in abreheim and leave the dark ally??")
                print("---------------------------------------------------------")
                if survive.lower() == "yes":
                    elena.progress = story.story_name_loop(story.chapter_names[0],
                                                           story.chap_1_names[4])
                    ally = 1
                else:
                    print("Lucky for you, i haven't coded more for this part, go back!!!")
                    print("---------------------------------------------------------")
                    elena.progress = story.story_name_loop(story.chapter_names[0],
                                                           story.chap_1_names[4])
                    ally = 1


def intro():

    elena.name = input("Welcome hero what is your name? ")
    cloud.name = input("Greetings young hero what is your name? ")
    cloud.x_initials_stats()
    elena.x_initials_stats()
    enemy.ex_initials_stats()
    enemy_2.ex_initials_stats()
    enemy_3.ex_initials_stats()
    play.music_loop(play.music_main[0])
    while True:
        skip_intro = input("Do you want to skip intro?")
        if skip_intro == "no" or skip_intro == "No":
            intro_text = "After many decades of miss using the planet.\nOur planet has now become hard to live on\n" \
                         "The wind is uneven, the rain falls like waterfalls in the desserts\n" \
                         "and snow are covering the rain forest.\n"
            for char in intro_text:
                sys.stdout.write(char)
                sys.stdout.flush()
                time.sleep(time_writing)
            input("Press enter to continue")
            intro_text_2 = "There is one way to stop this madness\nbut it will take a lot of courage to do it.\n" \
                           "Hopefully you two can give the planet\nthe humanity\n" \
                           "a new place to live on......\n"
            for char in intro_text_2:
                sys.stdout.write(char)
                sys.stdout.flush()
                time.sleep(time_writing)
            time.sleep(time_short_wait)
            input("Press enter to continue")
            intro_text_3 = "Cause you have been chosen!!\nStanding in front of the portal" \
                           " of time\nthe portal that will lead you to the" \
                           " worlds of times\nand will be able to change the" \
                           " destiny of our planet!!\n" \
                           "Will you enter with your mind or heart?\n"
            for char in intro_text_3:
                sys.stdout.write(char)
                sys.stdout.flush()
                time.sleep(time_writing)
            time.sleep(time_short_wait)
            intro_2()
        elif skip_intro == "yes" or skip_intro == "Yes":
            intro_2()
        else:
            print("please choose a valid choice")
            print("---------------------------------------------------------")


def intro_2():
    while True:

        print("---------------------------------------------------------")
        choice = input("Select your destiny: \n1. Mind\n2. Heart\nWhat is your choice?")
        print("---------------------------------------------------------")
        if choice.lower() == "1" or choice.lower() == "mind":
            gil_chance = random.randint(1, 2)
            if gil_chance == 1:
                cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [200]})
                elena.x_add_items_inventory({"Item": ["gil"], "QTY": [200]})
                print("---------------------------------------------------------")
                print("You receive: 200 gil")
                print("---------------------------------------------------------")
                input("Press enter to continue")
                cloud.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                ask_save()
                story1()
            elif gil_chance == 2:
                cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [500]})
                elena.x_add_items_inventory({"Item": ["gil"], "QTY": [500]})
                print("---------------------------------------------------------")
                print("You receive: 500 gil")
                print("---------------------------------------------------------")
                input("Press enter to continue")
                cloud.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                ask_save()
                story1()
        elif choice.lower() == "2" or choice.lower() == "heart":
            serpent_ring_chance = random.randint(1, 5)
            if serpent_ring_chance <= 2:
                cloud.x_finding_weapon_armory_on_journey("The mark of the serpents")
                elena.x_finding_weapon_armory_on_journey("The mark of the serpents")
                print("---------------------------------------------------------")
                print("You receive the ring: ", cloud.x_see_inventory_name_qty("ring", 1)[0])
                print("---------------------------------------------------------")
                input("Press enter to continue")
                cloud.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                ask_save()
                story1()
            elif serpent_ring_chance >= 3:
                cloud.status_poison = 1
                elena.status_poison = 1
                print("---------------------------------------------------------")
                print("Bad luck, the evil from the words of times have poisoned you.")
                print("---------------------------------------------------------")
                input("Press enter to continue")
                cloud.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                elena.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[1])
                ask_save()
                story1()
        else:
            print("Make a choice")
            print("---------------------------------------------------------")


def story1():
    chapter_1 = "CHAPTER ONE\n"
    print("---------------------------------------------------------")
    for char in chapter_1:
        sys.stdout.write(char)
        sys.stdout.flush()
        time.sleep(time_writing)
    time.sleep(time_short_wait)
    print("---------------------------------------------------------")
    round_room_text = "You open your eyes......\n"
    for char in round_room_text:
        sys.stdout.write(char)
        sys.stdout.flush()
        time.sleep(time_writing)
    print("---------------------------------------------------------")
    time.sleep(time_short_wait)
    story1_first_crossing()


def story1_first_crossing():
    while True:
        choice = 1
        while choice == 1:
            if choice == 1:
                story.progress = story.story_name_loop(story.chapter_names[0], story.chap_1_names[2])
                play.music_loop(play.sound_chapter_1[0])
                cloud.open_chest_count += 1
                print("---------------------------------------------------------")
                tprint("The  Crossing")
                print("---------------------------------------------------------")
                print("You are standing in a beautiful landscape, there is mountains to the north,\n"
                      "a forest to the east, a town located west,\n"
                      "south is a place of deep snow and deep troubles.\n"
                      "or you can stay and explore the area.")
                print("---------------------------------------------------------")
                walk = input("Where do you want to go?")
                print("---------------------------------------------------------")

                if "mountains" in walk or "north" in walk:
                    story1_mountains_in_north()
                    break
                elif "forest" in walk or "east" in walk:
                    story1_forest_in_east()
                    break
                elif "south" in walk or "deep snow" in walk: \
                        story1_snow_in_south()
                elif "west" in walk or "town" in walk:
                    story1_abrehiem_town()
                elif "explore" in walk:
                    print("You find a chest, in one of the bushes")
                    open_chest = input("Do you want to open?")
                    print("---------------------------------------------------------")
                    if open_chest.lower() == "yes" or open_chest.lower() == "YES":
                        if cloud.open_chest_count == 1:
                            money_chance = random.randint(50, 200)
                            cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [money_chance]})
                            elena.x_add_items_inventory({"Item": ["gil"], "QTY": [money_chance]})
                            print("You found", money_chance, "GIL, each")
                            print("---------------------------------------------------------")
                        elif cloud.open_chest_count == 3:
                            cloud.x_add_items_inventory({"Item": ["potion"], "QTY": [1]})
                            elena.x_add_items_inventory({"Item": ["potion"], "QTY": [1]})
                            print("You received 1 potion each")
                            print("---------------------------------------------------------")
                        elif cloud.open_chest_count == 2 or cloud.open_chest_count == 4 \
                                or cloud.open_chest_count == 5:
                            print("The chest is empty")
                            print("They are coming!!!!")
                            print("---------------------------------------------------------")
                            YourHero.x_battle(cloud, 0, cloud, elena, "no")
                        elif cloud.open_chest_count == 6:
                            cloud.open_chest_count = 0
                            print("Nothing inside")
                            print("---------------------------------------------------------")
                    else:
                        print("You decided to leave the chest untouched")
                        print("---------------------------------------------------------")
                elif walk.lower() == "inventory":
                    YourHero.x_inventory(elena)
                    YourHero.x_inventory(cloud)
                    print("---------------------------------------------------------")
                elif walk.lower() == "save":
                    ask_save()
                elif walk.lower() == "exit()":
                    cloud.x_exit_game()
                elif walk.lower() == "code_chris83":
                    secret_menu()
                elif walk.lower() == "status screen" or walk.lower() == "stats":
                    print(cloud.item_list)
                    print("---------------------------------------------------------")
                    print(elena.item_list)
                    print("---------------------------------------------------------")
                    input("Press enter to continue")
                    print(cloud.total_stat)
                    print("---------------------------------------------------------")
                    print(elena.total_stat)
                    input("Press enter to continue\n")

                else:
                    print("Choose a valid option")
                    print("---------------------------------------------------------")


def story2():
    chapter_2 = "CHAPTER TWO\n"
    # elena.story_number += 1
    for char in chapter_2:
        sys.stdout.write(char)
        sys.stdout.flush()
        time.sleep(time_writing)
    time.sleep(time_short_wait)
    print("When you leave the room of mystery you are blinded by a strong light, they are coming!!!")
    input("Press enter")
    YourHero.x_battle(cloud, 0, cloud, elena, "no")
    YourHero.x_battle(cloud, 0, cloud, elena, "no")
    elena.x_story_chapter_count()


def story3():
    print("Story Chapter", elena.story_number)
    elena.story_number += 1
    elena.x_story_chapter_count()


def story4():
    print("Story Chapter", elena.story_number)
    print(elena.stat)
    print(cloud.stat)
    input("Press enter")
    exit()


def secret_menu():
    print("Some secret")
    cloud.x_add_items_inventory({"Item": ["gil"], "QTY": [100000]})
    elena.x_add_items_inventory({"Item": ["gil"], "QTY": [100000]})
    input("press enter")
    print("---------------------------------------------------------")

def find_files(filename):
    result = []
    for root, dirs, files in os.walk(CURR_DIR_PATH):
        for name in files:
            if name == filename:
                result.append(CURR_DIR_PATH + "\\" + filename)
    # print(result)
    return result


def installing_game():
    finding = find_files("lost_shadow_installed.txt")
    if finding == []:

        # run make_excel_files.py
        make_excel_files.run_main()
        main_folder.get_folder_name()
        main_folder.get_music_name_id()
        important.install_game()
        # open("lost_shadow_installed.txt", "w+")
        dir_path.save_path = important.dir_path[important.save]
        dir_path.enemy_path = important.dir_path[important.enemy]
        dir_path.inventory_items_path = important.dir_path[important.inventory_items]
        dir_path.magic_path = important.dir_path[important.magic]
        dir_path.weapons_armor_path = important.dir_path[important.weapons_armor]
        dir_path.level_path = important.dir_path[important.level]
        dir_path.music_main_intro = important.dir_path[important.main_intro_intro]
        dir_path.chapter_1_path = important.dir_path[important.chapter_1_path]
        dir_path.for_all_path = important.dir_path[important.for_all_path]
        dir_path.battle_path = important.dir_path[important.battle_path]
        initiate_game()
    else:
        # file_open = open("lost_shadow_installed.txt", "r")
        read = CURR_DIR_PATH  # file_open.read()
        use_pdf_input = read.replace("\\", "\\\\")
        extra_backslash = "\\\\"
        for i in important.folder_name:
            important_directory = use_pdf_input + extra_backslash + i + extra_backslash
            important.dir_path.append(important_directory)
        important.music_path_2()
        important.dir_path.pop(4)
        # file_open.close()
        dir_path.save_path = important.dir_path[important.save]
        dir_path.enemy_path = important.dir_path[important.enemy]
        dir_path.inventory_items_path = important.dir_path[important.inventory_items]
        dir_path.magic_path = important.dir_path[important.magic]
        dir_path.weapons_armor_path = important.dir_path[important.weapons_armor]
        dir_path.level_path = important.dir_path[important.level]
        dir_path.music_main_intro = important.dir_path[important.main_intro_intro]
        dir_path.chapter_1_path = important.dir_path[important.chapter_1_path]
        dir_path.for_all_path = important.dir_path[important.for_all_path]
        dir_path.battle_path = important.dir_path[important.battle_path]
        # print(dir_path.save_path)
        initiate_game()


def initiate_game():
    # file names_important files

    my_timer.start_time()
    cloud.armory_inventory = {"Weapon/Armor": ["wind storm"], "Magic Slots": [1], "weapon power": [6],
                              "Protection": ["none"], "Equipped": ["True"], "Type": ["weapon"]}
    cloud.magic_spells = {"Magic": ["fire"], "mp cost": [4], "AP": [0], "Level": [1],
                          "Equipped": ["True"], "Magic Type": ["black magic"]}
    cloud.magic_slots += 1
    cloud.magic_slots_max += 1
    cloud.weapon_equipped = 1
    cloud.inventory = {"Item": ["gil", "potion", "ether", "antidote", "phoenix down", "silver dust", "tent",
                                "Key world item", "Key world item 2", "Key world item 3"],
                       "QTY": [100, 2, 0, 0, 0, 0, 0, "", "", ""]}
    elena.armory_inventory = {"Weapon/Armor": ["sand storm"], "Magic Slots": [1], "weapon power": [6],
                              "Protection": ["none"], "Equipped": ["True"], "Type": ["weapon"]}
    elena.magic_spells = {"Magic": ["fire"], "mp cost": [4], "AP": [0], "Level": [1],
                          "Equipped": ["True"], "Magic Type": ["black magic"]}
    elena.magic_slots += 1
    elena.magic_slots_max += 1
    elena.weapon_equipped = 1
    elena.inventory = {"Item": ["gil", "potion", "ether", "antidote", "phoenix down", "silver dust", "tent",
                                "Key world item", "Key world item 2", "Key world item 3"],
                       "QTY": [100, 2, 0, 0, 0, 0, 0, "", "", ""]}

    main_game()


installing_game()

'''
Buying habits, when bought over atleast 10 then you can have discount, easy to code, check with excel file,
compare >= 10 then....discount, maybe a count, for 1 next discount at 30 sold?? and so on...

using "in" to search for the input of the users text to navigate in the game, change this in the whole game!!
 if "mountains" in walk or "north" in walk:
    luck = 1
    # Rings & amulets
    # Enemy fought count
    enemy_kill = 0
    # Adding a special attack or something that will unlock when level 5 reached
    if boss:
    YourHero.x_battle(cloud, 1, cloud, elena, "yes")


'''
